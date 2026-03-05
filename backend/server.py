from fastapi import FastAPI, APIRouter, Depends, HTTPException, status, UploadFile, File, Query, Request
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import Response, FileResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func as func_sql, and_, or_, update, delete, text, extract
from sqlalchemy.orm import selectinload
from dotenv import load_dotenv
from pathlib import Path
from datetime import datetime, timezone, timedelta, date
from zoneinfo import ZoneInfo
from decimal import Decimal
from typing import List, Optional
import os
import logging
import httpx
import bcrypt
import jwt
import uuid
import shutil
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch, cm, mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT

# Local imports
from database import get_db, init_db, engine, Base, async_session_maker
from models import (
    Empresa, Usuario, Rol, Permiso, RolPermiso, UsuarioRol,
    Cliente, CreditoCliente, PagoCredito, Proveedor, ProveedorProducto, DeudaProveedor,
    Categoria, Marca, Producto, MateriaLaboratorio, EstadoMateria,
    Almacen, StockActual, MovimientoStock, TipoMovimientoStock,
    Venta, VentaItem, EstadoVenta, TipoPago,
    Funcionario, AdelantoSalario, CicloSalario,
    Vehiculo, TipoVehiculo, Entrega, EstadoEntrega,
    Factura, DocumentoElectronico, PreferenciaUsuario,
    DocumentoTemporal, TipoDocumento,
    Gasto,
)
from schemas import (
    EmpresaCreate, EmpresaResponse,
    UsuarioCreate, UsuarioResponse, UsuarioLogin, TokenResponse,
    RolCreate, RolResponse, PermisoResponse,
    ClienteCreate, ClienteResponse, CreditoClienteCreate, CreditoClienteResponse,
    ProveedorCreate, ProveedorResponse, DeudaProveedorCreate, DeudaProveedorResponse,
    CategoriaCreate, CategoriaResponse, MarcaCreate, MarcaResponse,
    GastoCreate, GastoUpdate, GastoResponse,
    ProductoCreate, ProductoResponse, ProductoConStock,
    MateriaLaboratorioCreate, MateriaLaboratorioResponse,
    AlmacenCreate, AlmacenResponse, StockActualCreate, StockActualResponse, StockConDetalles,
    MovimientoStockCreate, MovimientoStockResponse, TraspasoStockCreate, SalidaStockCreate,
    EntradaStockHistorialResponse,
    VentaCreate, VentaUpdate, VentaResponse, VentaConDetalles, VentaItemResponse,
    FuncionarioCreate, FuncionarioResponse,
    AdelantoSalarioCreate, AdelantoSalarioResponse,
    CicloSalarioCreate, CicloSalarioResponse,
    VehiculoCreate, VehiculoResponse,
    EntregaCreate, EntregaResponse, EntregaConDetalles, AsignarEntrega,
    FacturaCreate, FacturaResponse,
    PreferenciaUsuarioCreate, PreferenciaUsuarioResponse,
    DashboardStats, VentasPorHora, StockBajo, CotizacionDivisa, Alerta,
    DocumentoTemporalCreate, DocumentoTemporal as DocumentoTemporalSchema
)

# ==================== TIMEZONE CONFIGURATION ====================
# Paraguay timezone: America/Asuncion (GMT-4 standard, GMT-3 during DST)
PARAGUAY_TZ = ZoneInfo("America/Asuncion")

def now_paraguay():
    """Obtiene la fecha y hora actual en zona horaria de Paraguay"""
    return datetime.now(PARAGUAY_TZ)

def today_paraguay():
    """Obtiene la fecha actual (solo fecha) en Paraguay"""
    return now_paraguay().date()

def get_day_range_paraguay(date_obj):
    """Obtiene el rango completo de un día en Paraguay (00:00:01 - 23:59:59)"""
    # Start: 00:00:01 of the day in Paraguay
    day_start = datetime.combine(date_obj, datetime.min.time()).replace(tzinfo=PARAGUAY_TZ)
    # End: 23:59:59 of the day in Paraguay
    day_end = datetime.combine(date_obj, datetime.max.time()).replace(tzinfo=PARAGUAY_TZ)
    return day_start, day_end

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# Create uploads directory
UPLOADS_DIR = ROOT_DIR / 'uploads' / 'productos'
UPLOADS_DIR.mkdir(parents=True, exist_ok=True)

# JWT Config
JWT_SECRET = os.environ.get('JWT_SECRET', 'luzbrill-secret-key-change-in-production')
JWT_ALGORITHM = "HS256"
JWT_EXPIRATION_HOURS = 8

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Manual currency rates storage
MANUAL_CURRENCY_RATES = {
    'usd_pyg': None,
    'brl_pyg': None,
    'manual': False,
    'updated_at': None
}

# Create FastAPI app
app = FastAPI(title="Luz Brill ERP API", version="1.0.0")

# CORS Configuration
cors_origins = os.environ.get('CORS_ORIGINS', '*')
if cors_origins == '*':
    allowed_origins = ['*']
    allow_credentials = False  # Can't use credentials with wildcard
else:
    allowed_origins = [origin.strip() for origin in cors_origins.split(',')]
    allow_credentials = True

logger.info(f"CORS allowed origins: {allowed_origins}")
logger.info(f"CORS allow credentials: {allow_credentials}")

app.add_middleware(
    CORSMiddleware,
    allow_origins=allowed_origins,
    allow_credentials=allow_credentials,
    allow_methods=["*"],
    allow_headers=["*"],
    expose_headers=["*"]
)

# Mount static files for uploads
app.mount("/uploads", StaticFiles(directory=str(ROOT_DIR / 'uploads')), name="uploads")

# API Router
api_router = APIRouter(prefix="/api")

# ==================== AUTH HELPERS ====================
def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

def verify_password(password: str, hashed: str) -> bool:
    return bcrypt.checkpw(password.encode('utf-8'), hashed.encode('utf-8'))

# Security scheme for JWT Bearer tokens
security_scheme = HTTPBearer(auto_error=False)

def create_token(usuario_id: int) -> str:
    payload = {
        "sub": str(usuario_id),
        "exp": now_paraguay() + timedelta(hours=JWT_EXPIRATION_HOURS)
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

def decode_token(token: str) -> dict:
    """Decode and validate a JWT token. Raises HTTPException on failure."""
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        return payload
    except jwt.ExpiredSignatureError:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Token expirado. Por favor inicie sesión nuevamente."
        )
    except jwt.InvalidTokenError:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Token inválido."
        )

async def get_current_usuario(
    credentials: HTTPAuthorizationCredentials = Depends(security_scheme),
    db: AsyncSession = Depends(get_db)
) -> Usuario:
    """Dependency that validates JWT and returns the current user."""
    if not credentials:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Token de autenticación requerido."
        )
    payload = decode_token(credentials.credentials)
    usuario_id = int(payload.get("sub", 0))
    if not usuario_id:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Token inválido: sin identificador de usuario."
        )
    result = await db.execute(select(Usuario).where(Usuario.id == usuario_id))
    usuario = result.scalar_one_or_none()
    if not usuario:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Usuario no encontrado."
        )
    if not usuario.activo:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Usuario inactivo."
        )
    return usuario

def numero_a_letras(numero):
    """Convert number to Spanish words for receipts"""
    unidades = ['', 'uno', 'dos', 'tres', 'cuatro', 'cinco', 'seis', 'siete', 'ocho', 'nueve']
    decenas = ['', 'diez', 'veinte', 'treinta', 'cuarenta', 'cincuenta', 'sesenta', 'setenta', 'ochenta', 'noventa']
    especiales = {11: 'once', 12: 'doce', 13: 'trece', 14: 'catorce', 15: 'quince', 
                  16: 'dieciséis', 17: 'diecisiete', 18: 'dieciocho', 19: 'diecinueve'}
    centenas = ['', 'ciento', 'doscientos', 'trescientos', 'cuatrocientos', 'quinientos',
                'seiscientos', 'setecientos', 'ochocientos', 'novecientos']
    
    n = int(numero)
    if n == 0:
        return 'cero'
    if n == 100:
        return 'cien'
    
    resultado = ''
    
    if n >= 1000000:
        millones = n // 1000000
        resultado += ('un millón ' if millones == 1 else numero_a_letras(millones) + ' millones ')
        n %= 1000000
    
    if n >= 1000:
        miles = n // 1000
        resultado += ('mil ' if miles == 1 else numero_a_letras(miles) + ' mil ')
        n %= 1000
    
    if n >= 100:
        resultado += centenas[n // 100] + ' '
        n %= 100
    
    if n in especiales:
        resultado += especiales[n]
    elif n >= 10:
        resultado += decenas[n // 10]
        if n % 10:
            resultado += ' y ' + unidades[n % 10]
    elif n > 0:
        resultado += unidades[n]
    
    return resultado.strip()

# ==================== ROOT ====================
@api_router.get("/")
async def root():
    return {"message": "Luz Brill ERP API v1.0"}

@api_router.get("/health")
async def health():
    return {"status": "healthy", "timestamp": now_paraguay().isoformat()}

# ==================== EMPRESA ====================
@api_router.post("/empresas", response_model=EmpresaResponse)
async def crear_empresa(data: EmpresaCreate, db: AsyncSession = Depends(get_db)):
    empresa = Empresa(**data.model_dump())
    db.add(empresa)
    await db.commit()
    await db.refresh(empresa)
    return empresa

@api_router.get("/empresas", response_model=List[EmpresaResponse])
async def listar_empresas(db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Empresa).where(Empresa.estado == True))
    return result.scalars().all()

@api_router.get("/empresas/{empresa_id}", response_model=EmpresaResponse)
async def obtener_empresa(empresa_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Empresa).where(Empresa.id == empresa_id))
    empresa = result.scalar_one_or_none()
    if not empresa:
        raise HTTPException(status_code=404, detail="Empresa no encontrada")
    return empresa

# ==================== AUTH ====================
@api_router.post("/auth/register", response_model=TokenResponse)
async def registrar_usuario(data: UsuarioCreate, db: AsyncSession = Depends(get_db)):
    existing = await db.execute(select(Usuario).where(Usuario.email == data.email))
    if existing.scalar_one_or_none():
        raise HTTPException(status_code=400, detail="Email ya registrado")
    
    usuario = Usuario(
        email=data.email,
        password_hash=hash_password(data.password),
        nombre=data.nombre,
        apellido=data.apellido,
        telefono=data.telefono,
        empresa_id=data.empresa_id
    )
    db.add(usuario)
    await db.commit()
    await db.refresh(usuario)
    
    token = create_token(usuario.id)
    return TokenResponse(access_token=token, usuario=UsuarioResponse.model_validate(usuario))

@api_router.post("/auth/login", response_model=TokenResponse)
async def login(data: UsuarioLogin, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Usuario).where(Usuario.email == data.email))
    usuario = result.scalar_one_or_none()
    
    if not usuario or not verify_password(data.password, usuario.password_hash):
        raise HTTPException(status_code=401, detail="Credenciales inválidas")
    
    if not usuario.activo:
        raise HTTPException(status_code=401, detail="Usuario inactivo")
    
    token = create_token(usuario.id)
    return TokenResponse(access_token=token, usuario=UsuarioResponse.model_validate(usuario))

@api_router.get("/auth/me", response_model=UsuarioResponse)
async def obtener_usuario_actual(usuario: Usuario = Depends(get_current_usuario)):
    """Validate JWT token and return current user data."""
    return usuario

# ==================== USUARIOS ====================
@api_router.get("/usuarios", response_model=List[UsuarioResponse])
async def listar_usuarios(empresa_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(Usuario)
        .options(selectinload(Usuario.roles))
        .where(Usuario.empresa_id == empresa_id, Usuario.activo == True)
    )
    usuarios = result.scalars().all()
    
    # Agregar rol_id desde la relación usuario_roles
    usuarios_response = []
    for usuario in usuarios:
        usuario_dict = {
            "id": usuario.id,
            "empresa_id": usuario.empresa_id,
            "email": usuario.email,
            "nombre": usuario.nombre,
            "apellido": usuario.apellido,
            "telefono": usuario.telefono,
            "activo": usuario.activo,
            "creado_en": usuario.creado_en,
            "rol_id": usuario.roles[0].rol_id if usuario.roles else None
        }
        usuarios_response.append(usuario_dict)
    
    return usuarios_response

@api_router.get("/usuarios/{usuario_id}", response_model=UsuarioResponse)
async def obtener_usuario(usuario_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Usuario).where(Usuario.id == usuario_id))
    usuario = result.scalar_one_or_none()
    if not usuario:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    return usuario

@api_router.put("/usuarios/{usuario_id}", response_model=UsuarioResponse)
async def actualizar_usuario(usuario_id: int, data: dict, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Usuario).where(Usuario.id == usuario_id))
    usuario = result.scalar_one_or_none()
    if not usuario:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    
    for key, value in data.items():
        if key == 'password' and value:
            usuario.password_hash = hash_password(value)
        elif hasattr(usuario, key) and key not in ['id', 'password_hash', 'password']:
            setattr(usuario, key, value)
    
    await db.commit()
    await db.refresh(usuario)
    return usuario

@api_router.delete("/usuarios/{usuario_id}")
async def eliminar_usuario(usuario_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Usuario).where(Usuario.id == usuario_id))
    usuario = result.scalar_one_or_none()
    if not usuario:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    
    usuario.activo = False
    await db.commit()
    return {"message": "Usuario desactivado"}

# ==================== ROLES Y PERMISOS ====================
@api_router.post("/roles", response_model=RolResponse)
async def crear_rol(data: RolCreate, db: AsyncSession = Depends(get_db)):
    rol = Rol(**data.model_dump())
    db.add(rol)
    await db.commit()
    await db.refresh(rol)
    return rol

@api_router.get("/roles", response_model=List[RolResponse])
async def listar_roles(empresa_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Rol).where(Rol.empresa_id == empresa_id))
    return result.scalars().all()

@api_router.get("/permisos", response_model=List[PermisoResponse])
async def listar_permisos(db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Permiso))
    return result.scalars().all()

@api_router.post("/permisos", response_model=PermisoResponse)
async def crear_permiso(data: dict, db: AsyncSession = Depends(get_db)):
    permiso = Permiso(clave=data['clave'], descripcion=data.get('descripcion'))
    db.add(permiso)
    await db.commit()
    await db.refresh(permiso)
    return permiso

@api_router.post("/roles/{rol_id}/permisos/{permiso_id}")
async def asignar_permiso_rol(rol_id: int, permiso_id: int, db: AsyncSession = Depends(get_db)):
    # Check if already exists
    existing = await db.execute(
        select(RolPermiso).where(RolPermiso.rol_id == rol_id, RolPermiso.permiso_id == permiso_id)
    )
    if existing.scalar_one_or_none():
        return {"message": "Permiso ya asignado"}
    
    rol_permiso = RolPermiso(rol_id=rol_id, permiso_id=permiso_id)
    db.add(rol_permiso)
    await db.commit()
    return {"message": "Permiso asignado"}

@api_router.delete("/roles/{rol_id}/permisos/{permiso_id}")
async def quitar_permiso_rol(rol_id: int, permiso_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(RolPermiso).where(RolPermiso.rol_id == rol_id, RolPermiso.permiso_id == permiso_id)
    )
    rol_permiso = result.scalar_one_or_none()
    if rol_permiso:
        await db.delete(rol_permiso)
        await db.commit()
    return {"message": "Permiso quitado"}

@api_router.get("/roles/{rol_id}/permisos")
async def obtener_permisos_rol(rol_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(Permiso)
        .join(RolPermiso, Permiso.id == RolPermiso.permiso_id)
        .where(RolPermiso.rol_id == rol_id)
    )
    return result.scalars().all()

@api_router.post("/usuarios/{usuario_id}/roles/{rol_id}")
async def asignar_rol_usuario(usuario_id: int, rol_id: int, db: AsyncSession = Depends(get_db)):
    # Check if already exists
    existing = await db.execute(
        select(UsuarioRol).where(UsuarioRol.usuario_id == usuario_id, UsuarioRol.rol_id == rol_id)
    )
    if existing.scalar_one_or_none():
        return {"message": "Rol ya asignado"}
    
    usuario_rol = UsuarioRol(usuario_id=usuario_id, rol_id=rol_id)
    db.add(usuario_rol)
    await db.commit()
    return {"message": "Rol asignado"}

@api_router.delete("/usuarios/{usuario_id}/roles/{rol_id}")
async def quitar_rol_usuario(usuario_id: int, rol_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(UsuarioRol).where(UsuarioRol.usuario_id == usuario_id, UsuarioRol.rol_id == rol_id)
    )
    usuario_rol = result.scalar_one_or_none()
    if usuario_rol:
        await db.delete(usuario_rol)
        await db.commit()
    return {"message": "Rol quitado"}

@api_router.get("/usuarios/{usuario_id}/permisos")
async def obtener_permisos_usuario(usuario_id: int, db: AsyncSession = Depends(get_db)):
    """Get all permissions for a user through their roles"""
    result = await db.execute(
        select(Permiso)
        .join(RolPermiso, Permiso.id == RolPermiso.permiso_id)
        .join(UsuarioRol, RolPermiso.rol_id == UsuarioRol.rol_id)
        .where(UsuarioRol.usuario_id == usuario_id)
        .distinct()
    )
    permisos = result.scalars().all()
    return [{"id": p.id, "clave": p.clave, "descripcion": p.descripcion} for p in permisos]

# ==================== CLIENTES ====================
@api_router.post("/clientes", response_model=ClienteResponse)
async def crear_cliente(data: ClienteCreate, db: AsyncSession = Depends(get_db)):
    cliente = Cliente(**data.model_dump())
    db.add(cliente)
    await db.commit()
    await db.refresh(cliente)
    return cliente

@api_router.get("/clientes", response_model=List[ClienteResponse])
async def listar_clientes(empresa_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(Cliente).where(Cliente.empresa_id == empresa_id, Cliente.estado == True)
    )
    return result.scalars().all()

@api_router.get("/clientes/{cliente_id}", response_model=ClienteResponse)
async def obtener_cliente(cliente_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Cliente).where(Cliente.id == cliente_id))
    cliente = result.scalar_one_or_none()
    if not cliente:
        raise HTTPException(status_code=404, detail="Cliente no encontrado")
    return cliente

@api_router.put("/clientes/{cliente_id}", response_model=ClienteResponse)
async def actualizar_cliente(cliente_id: int, data: dict, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Cliente).where(Cliente.id == cliente_id))
    cliente = result.scalar_one_or_none()
    if not cliente:
        raise HTTPException(status_code=404, detail="Cliente no encontrado")
    
    for key, value in data.items():
        if hasattr(cliente, key) and key != 'id':
            setattr(cliente, key, value)
    
    await db.commit()
    await db.refresh(cliente)
    return cliente

@api_router.delete("/clientes/{cliente_id}")
async def eliminar_cliente(cliente_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Cliente).where(Cliente.id == cliente_id))
    cliente = result.scalar_one_or_none()
    if not cliente:
        raise HTTPException(status_code=404, detail="Cliente no encontrado")
    
    cliente.estado = False
    await db.commit()
    return {"message": "Cliente desactivado"}

# ==================== CREDITOS CLIENTES ====================
@api_router.get("/clientes/{cliente_id}/credito-disponible")
async def obtener_credito_disponible(cliente_id: int, db: AsyncSession = Depends(get_db)):
    """Obtiene el crédito usado y disponible de un cliente"""
    result = await db.execute(select(Cliente).where(Cliente.id == cliente_id))
    cliente = result.scalar_one_or_none()
    if not cliente:
        raise HTTPException(status_code=404, detail="Cliente no encontrado")
    
    # Calcular crédito usado (suma de monto_pendiente de créditos no pagados)
    creditos_result = await db.execute(
        select(func_sql.coalesce(func_sql.sum(CreditoCliente.monto_pendiente), 0))
        .where(CreditoCliente.cliente_id == cliente_id, CreditoCliente.pagado == False)
    )
    credito_usado = creditos_result.scalar() or Decimal('0')
    
    limite = cliente.limite_credito or Decimal('0')
    disponible = max(Decimal('0'), limite - credito_usado)
    
    return {
        "cliente_id": cliente_id,
        "limite_credito": float(limite),
        "credito_usado": float(credito_usado),
        "credito_disponible": float(disponible)
    }

@api_router.post("/clientes/{cliente_id}/creditos", response_model=CreditoClienteResponse)
async def crear_credito_cliente(cliente_id: int, data: dict, db: AsyncSession = Depends(get_db)):
    """Registra una nueva transacción a crédito"""
    monto = Decimal(str(data.get('monto_original', 0)))
    
    # Verificar límite de crédito
    result = await db.execute(select(Cliente).where(Cliente.id == cliente_id))
    cliente = result.scalar_one_or_none()
    if not cliente:
        raise HTTPException(status_code=404, detail="Cliente no encontrado")
    
    # Calcular crédito actual
    creditos_result = await db.execute(
        select(func_sql.coalesce(func_sql.sum(CreditoCliente.monto_pendiente), 0))
        .where(CreditoCliente.cliente_id == cliente_id, CreditoCliente.pagado == False)
    )
    credito_usado = creditos_result.scalar() or Decimal('0')
    
    limite = cliente.limite_credito or Decimal('0')
    # Si el límite es 0, no se permite crédito
    if limite <= 0:
        raise HTTPException(
            status_code=400,
            detail="El cliente no tiene crédito habilitado. Asigne un límite de crédito mayor a 0 para habilitar compras a crédito."
        )
    if (credito_usado + monto) > limite:
        raise HTTPException(
            status_code=400, 
            detail=f"Crédito excede el límite. Disponible: {float(limite - credito_usado):,.0f} Gs"
        )
    
    credito = CreditoCliente(
        cliente_id=cliente_id,
        venta_id=data.get('venta_id'),
        monto_original=monto,
        monto_pendiente=monto,
        descripcion=data.get('descripcion'),
        fecha_venta=today_paraguay(),
        creado_en=now_paraguay()
    )
    db.add(credito)
    await db.commit()
    await db.refresh(credito)
    return credito

@api_router.get("/clientes/{cliente_id}/creditos")
async def listar_creditos_cliente(cliente_id: int, solo_pendientes: bool = False, db: AsyncSession = Depends(get_db)):
    """Lista todos los créditos de un cliente con sus pagos"""
    query = select(CreditoCliente).where(CreditoCliente.cliente_id == cliente_id).order_by(CreditoCliente.fecha_venta.desc())
    if solo_pendientes:
        query = query.where(CreditoCliente.pagado == False)
    
    result = await db.execute(query)
    creditos = result.scalars().all()
    
    response = []
    for credito in creditos:
        # Get pagos for this credito
        pagos_result = await db.execute(
            select(PagoCredito).where(PagoCredito.credito_id == credito.id).order_by(PagoCredito.fecha_pago.desc())
        )
        pagos = pagos_result.scalars().all()
        
        response.append({
            "id": credito.id,
            "cliente_id": credito.cliente_id,
            "venta_id": credito.venta_id,
            "monto_original": float(credito.monto_original),
            "monto_pendiente": float(credito.monto_pendiente),
            "descripcion": credito.descripcion,
            "fecha_venta": credito.fecha_venta.isoformat() if credito.fecha_venta else None,
            "pagado": credito.pagado,
            "creado_en": credito.creado_en.isoformat() if credito.creado_en else None,
            "pagos": [
                {
                    "id": p.id,
                    "monto": float(p.monto),
                    "fecha_pago": p.fecha_pago.isoformat() if p.fecha_pago else None,
                    "observacion": p.observacion
                }
                for p in pagos
            ]
        })
    
    return response

@api_router.post("/creditos/{credito_id}/pagar")
async def pagar_credito(credito_id: int, data: dict, db: AsyncSession = Depends(get_db)):
    """Registra un pago parcial o total a un crédito"""
    result = await db.execute(select(CreditoCliente).where(CreditoCliente.id == credito_id))
    credito = result.scalar_one_or_none()
    if not credito:
        raise HTTPException(status_code=404, detail="Crédito no encontrado")
    
    if credito.pagado:
        raise HTTPException(status_code=400, detail="Este crédito ya está pagado")
    
    monto_pago = Decimal(str(data.get('monto', 0)))
    if monto_pago <= 0:
        raise HTTPException(status_code=400, detail="El monto debe ser mayor a 0")
    
    if monto_pago > credito.monto_pendiente:
        raise HTTPException(status_code=400, detail=f"El monto excede la deuda pendiente de {float(credito.monto_pendiente):,.0f} Gs")
    
    # Create payment record
    pago = PagoCredito(
        credito_id=credito_id,
        monto=monto_pago,
        observacion=data.get('observacion'),
        fecha_pago=now_paraguay()
    )
    db.add(pago)
    
    # Update credito
    credito.monto_pendiente = credito.monto_pendiente - monto_pago
    if credito.monto_pendiente <= 0:
        credito.pagado = True
        credito.monto_pendiente = Decimal('0')
    
    await db.commit()
    await db.refresh(credito)
    
    return {
        "message": "Pago registrado exitosamente",
        "credito_id": credito_id,
        "monto_pagado": float(monto_pago),
        "monto_pendiente": float(credito.monto_pendiente),
        "pagado": credito.pagado
    }

# ==================== PROVEEDORES ====================
@api_router.post("/proveedores", response_model=ProveedorResponse)
async def crear_proveedor(data: ProveedorCreate, db: AsyncSession = Depends(get_db)):
    proveedor = Proveedor(**data.model_dump())
    db.add(proveedor)
    await db.commit()
    await db.refresh(proveedor)
    return proveedor

@api_router.get("/proveedores", response_model=List[ProveedorResponse])
async def listar_proveedores(empresa_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(Proveedor).where(Proveedor.empresa_id == empresa_id, Proveedor.estado == True)
    )
    return result.scalars().all()

@api_router.get("/proveedores/{proveedor_id}", response_model=ProveedorResponse)
async def obtener_proveedor(proveedor_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Proveedor).where(Proveedor.id == proveedor_id))
    proveedor = result.scalar_one_or_none()
    if not proveedor:
        raise HTTPException(status_code=404, detail="Proveedor no encontrado")
    return proveedor

@api_router.put("/proveedores/{proveedor_id}", response_model=ProveedorResponse)
async def actualizar_proveedor(proveedor_id: int, data: dict, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Proveedor).where(Proveedor.id == proveedor_id))
    proveedor = result.scalar_one_or_none()
    if not proveedor:
        raise HTTPException(status_code=404, detail="Proveedor no encontrado")
    
    for key, value in data.items():
        if hasattr(proveedor, key) and key != 'id':
            setattr(proveedor, key, value)
    
    await db.commit()
    await db.refresh(proveedor)
    return proveedor

@api_router.delete("/proveedores/{proveedor_id}")
async def eliminar_proveedor(proveedor_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Proveedor).where(Proveedor.id == proveedor_id))
    proveedor = result.scalar_one_or_none()
    if not proveedor:
        raise HTTPException(status_code=404, detail="Proveedor no encontrado")
    
    proveedor.estado = False
    await db.commit()
    return {"message": "Proveedor desactivado"}

# ==================== DEUDAS PROVEEDORES ====================
@api_router.post("/proveedores/{proveedor_id}/deudas", response_model=DeudaProveedorResponse)
async def crear_deuda_proveedor(proveedor_id: int, data: dict, db: AsyncSession = Depends(get_db)):
    # Parse fecha_limite if provided as string
    fecha_limite_parsed = None
    if data.get('fecha_limite'):
        try:
            fecha_limite_parsed = date.fromisoformat(data['fecha_limite'])
        except (ValueError, TypeError):
            fecha_limite_parsed = None
    
    deuda = DeudaProveedor(
        proveedor_id=proveedor_id,
        monto=Decimal(str(data['monto'])),
        descripcion=data.get('descripcion'),
        fecha_emision=today_paraguay(),
        fecha_limite=fecha_limite_parsed,
        pagado=False
    )
    db.add(deuda)
    await db.commit()
    await db.refresh(deuda)
    return deuda

@api_router.get("/proveedores/{proveedor_id}/deudas", response_model=List[DeudaProveedorResponse])
async def listar_deudas_proveedor(proveedor_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(DeudaProveedor)
        .where(DeudaProveedor.proveedor_id == proveedor_id)
        .order_by(DeudaProveedor.creado_en.desc())
    )
    return result.scalars().all()

@api_router.put("/deudas/{deuda_id}/pagar")
async def pagar_deuda(deuda_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(DeudaProveedor).where(DeudaProveedor.id == deuda_id))
    deuda = result.scalar_one_or_none()
    if not deuda:
        raise HTTPException(status_code=404, detail="Deuda no encontrada")
    
    deuda.pagado = True
    deuda.fecha_pago = today_paraguay()
    await db.commit()
    return {"message": "Deuda marcada como pagada"}

@api_router.delete("/deudas/{deuda_id}")
async def eliminar_deuda(deuda_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(DeudaProveedor).where(DeudaProveedor.id == deuda_id))
    deuda = result.scalar_one_or_none()
    if not deuda:
        raise HTTPException(status_code=404, detail="Deuda no encontrada")
    await db.delete(deuda)
    await db.commit()
    return {"message": "Deuda eliminada"}

# ==================== CATEGORIAS ====================
@api_router.post("/categorias", response_model=CategoriaResponse)
async def crear_categoria(data: CategoriaCreate, db: AsyncSession = Depends(get_db)):
    categoria = Categoria(**data.model_dump())
    db.add(categoria)
    await db.commit()
    await db.refresh(categoria)
    return categoria

@api_router.get("/categorias", response_model=List[CategoriaResponse])
async def listar_categorias(empresa_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Categoria).where(Categoria.empresa_id == empresa_id))
    return result.scalars().all()

@api_router.put("/categorias/{categoria_id}", response_model=CategoriaResponse)
async def actualizar_categoria(categoria_id: int, data: dict, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Categoria).where(Categoria.id == categoria_id))
    categoria = result.scalar_one_or_none()
    if not categoria:
        raise HTTPException(status_code=404, detail="Categoría no encontrada")
    if 'nombre' in data:
        categoria.nombre = data['nombre']
    await db.commit()
    await db.refresh(categoria)
    return categoria

@api_router.delete("/categorias/{categoria_id}")
async def eliminar_categoria(categoria_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Categoria).where(Categoria.id == categoria_id))
    categoria = result.scalar_one_or_none()
    if categoria:
        await db.delete(categoria)
        await db.commit()
    return {"message": "Categoría eliminada"}

# ==================== MARCAS ====================
@api_router.post("/marcas", response_model=MarcaResponse)
async def crear_marca(data: MarcaCreate, db: AsyncSession = Depends(get_db)):
    marca = Marca(**data.model_dump())
    db.add(marca)
    await db.commit()
    await db.refresh(marca)
    return marca

@api_router.get("/marcas", response_model=List[MarcaResponse])
async def listar_marcas(empresa_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Marca).where(Marca.empresa_id == empresa_id))
    return result.scalars().all()

@api_router.delete("/marcas/{marca_id}")
async def eliminar_marca(marca_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Marca).where(Marca.id == marca_id))
    marca = result.scalar_one_or_none()
    if marca:
        await db.delete(marca)
        await db.commit()
    return {"message": "Marca eliminada"}

@api_router.put("/marcas/{marca_id}", response_model=MarcaResponse)
async def actualizar_marca(marca_id: int, data: dict, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Marca).where(Marca.id == marca_id))
    marca = result.scalar_one_or_none()
    if not marca:
        raise HTTPException(status_code=404, detail="Marca no encontrada")
    
    if 'nombre' in data:
        marca.nombre = data['nombre']
    
    await db.commit()
    await db.refresh(marca)
    return marca

# ==================== PRODUCTOS ====================
@api_router.post("/productos", response_model=ProductoResponse)
async def crear_producto(data: ProductoCreate, db: AsyncSession = Depends(get_db)):
    dump = data.model_dump()
    # Convert empty strings to NULL for unique-constrained fields
    if not dump.get('codigo_barra') or not str(dump['codigo_barra']).strip():
        dump['codigo_barra'] = None
    else:
        dump['codigo_barra'] = str(dump['codigo_barra']).strip()
    if not dump.get('imagen_url') or not str(dump['imagen_url']).strip():
        dump['imagen_url'] = None
    producto = Producto(**dump)
    db.add(producto)
    await db.commit()
    await db.refresh(producto)
    return producto

@api_router.get("/productos", response_model=List[ProductoConStock])
async def listar_productos(empresa_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(Producto, Categoria, Marca, Proveedor)
        .outerjoin(Categoria, Producto.categoria_id == Categoria.id)
        .outerjoin(Marca, Producto.marca_id == Marca.id)
        .outerjoin(Proveedor, Producto.proveedor_id == Proveedor.id)
        .where(Producto.empresa_id == empresa_id, Producto.activo == True)
    )
    productos = []
    for row in result.all():
        producto, categoria, marca, proveedor = row
        stock_result = await db.execute(
            select(func_sql.coalesce(func_sql.sum(StockActual.cantidad), 0))
            .where(StockActual.producto_id == producto.id)
        )
        stock_total = stock_result.scalar() or 0
        
        prod_dict = ProductoResponse.model_validate(producto).model_dump()
        prod_dict['stock_total'] = stock_total
        prod_dict['categoria_nombre'] = categoria.nombre if categoria else None
        prod_dict['marca_nombre'] = marca.nombre if marca else None
        prod_dict['proveedor_nombre'] = proveedor.nombre if proveedor else None
        productos.append(ProductoConStock(**prod_dict))
    
    return productos

@api_router.get("/productos/{producto_id}", response_model=ProductoConStock)
async def obtener_producto(producto_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(Producto, Categoria, Marca, Proveedor)
        .outerjoin(Categoria, Producto.categoria_id == Categoria.id)
        .outerjoin(Marca, Producto.marca_id == Marca.id)
        .outerjoin(Proveedor, Producto.proveedor_id == Proveedor.id)
        .where(Producto.id == producto_id)
    )
    row = result.first()
    if not row:
        raise HTTPException(status_code=404, detail="Producto no encontrado")
    
    producto, categoria, marca, proveedor = row
    stock_result = await db.execute(
        select(func_sql.coalesce(func_sql.sum(StockActual.cantidad), 0))
        .where(StockActual.producto_id == producto.id)
    )
    stock_total = stock_result.scalar() or 0
    
    prod_dict = ProductoResponse.model_validate(producto).model_dump()
    prod_dict['stock_total'] = stock_total
    prod_dict['categoria_nombre'] = categoria.nombre if categoria else None
    prod_dict['marca_nombre'] = marca.nombre if marca else None
    prod_dict['proveedor_nombre'] = proveedor.nombre if proveedor else None
    return ProductoConStock(**prod_dict)

@api_router.get("/productos/codigo/{codigo_barra}", response_model=ProductoConStock)
async def buscar_producto_por_codigo(codigo_barra: str, db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(Producto, Categoria, Marca, Proveedor)
        .outerjoin(Categoria, Producto.categoria_id == Categoria.id)
        .outerjoin(Marca, Producto.marca_id == Marca.id)
        .outerjoin(Proveedor, Producto.proveedor_id == Proveedor.id)
        .where(Producto.codigo_barra == codigo_barra)
    )
    row = result.first()
    if not row:
        raise HTTPException(status_code=404, detail="Producto no encontrado")
    
    producto, categoria, marca, proveedor = row
    stock_result = await db.execute(
        select(func_sql.coalesce(func_sql.sum(StockActual.cantidad), 0))
        .where(StockActual.producto_id == producto.id)
    )
    stock_total = stock_result.scalar() or 0
    
    prod_dict = ProductoResponse.model_validate(producto).model_dump()
    prod_dict['stock_total'] = stock_total
    prod_dict['categoria_nombre'] = categoria.nombre if categoria else None
    prod_dict['marca_nombre'] = marca.nombre if marca else None
    prod_dict['proveedor_nombre'] = proveedor.nombre if proveedor else None
    return ProductoConStock(**prod_dict)

@api_router.put("/productos/{producto_id}", response_model=ProductoResponse)
async def actualizar_producto(producto_id: int, data: dict, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Producto).where(Producto.id == producto_id))
    producto = result.scalar_one_or_none()
    if not producto:
        raise HTTPException(status_code=404, detail="Producto no encontrado")
    
    for key, value in data.items():
        if hasattr(producto, key) and key != 'id':
            # Handle codigo_barra specially - treat empty strings as NULL
            if key == 'codigo_barra':
                if value == '' or value is None:
                    value = None
                elif value and value.strip():
                    # Check for uniqueness of non-empty codigo_barra
                    existing = await db.execute(
                        select(Producto).where(
                            Producto.codigo_barra == value.strip(),
                            Producto.id != producto_id
                        )
                    )
                    if existing.scalar_one_or_none():
                        raise HTTPException(status_code=400, detail="Código de barra ya existe")
                    value = value.strip()
            setattr(producto, key, value)
    
    await db.commit()
    await db.refresh(producto)
    return producto

@api_router.delete("/productos/{producto_id}")
async def eliminar_producto(producto_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Producto).where(Producto.id == producto_id))
    producto = result.scalar_one_or_none()
    if not producto:
        raise HTTPException(status_code=404, detail="Producto no encontrado")
    
    # Eliminar registros de stock asociados
    await db.execute(delete(StockActual).where(StockActual.producto_id == producto_id))
    
    # Eliminar movimientos de stock asociados
    await db.execute(delete(MovimientoStock).where(MovimientoStock.producto_id == producto_id))
    
    # Desactivar el producto
    producto.activo = False
    await db.commit()
    return {"message": "Producto y stock asociado eliminados"}

# ==================== UPLOAD IMAGEN PRODUCTO ====================
@api_router.post("/productos/{producto_id}/imagen")
async def subir_imagen_producto(producto_id: int, file: UploadFile = File(...), db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Producto).where(Producto.id == producto_id))
    producto = result.scalar_one_or_none()
    if not producto:
        raise HTTPException(status_code=404, detail="Producto no encontrado")
    
    # Generate unique filename
    ext = file.filename.split('.')[-1] if '.' in file.filename else 'jpg'
    filename = f"{uuid.uuid4()}.{ext}"
    filepath = UPLOADS_DIR / filename
    
    # Save file
    with open(filepath, 'wb') as f:
        shutil.copyfileobj(file.file, f)
    
    # Update product
    producto.imagen_url = f"/uploads/productos/{filename}"
    await db.commit()
    await db.refresh(producto)
    
    return {"imagen_url": producto.imagen_url}

# ==================== MATERIAS LABORATORIO ====================
import uuid as _uuid

@api_router.post("/materias-laboratorio", response_model=MateriaLaboratorioResponse)
async def crear_materia_laboratorio(data: MateriaLaboratorioCreate, db: AsyncSession = Depends(get_db)):
    # Auto-generate unique barcode if not provided
    if not data.codigo_barra:
        while True:
            codigo = f"LAB{_uuid.uuid4().hex[:8].upper()}"
            dup_mat = await db.execute(
                select(MateriaLaboratorio).where(MateriaLaboratorio.codigo_barra == codigo)
            )
            dup_prod = await db.execute(
                select(Producto).where(Producto.codigo_barra == codigo)
            )
            if not dup_mat.scalar_one_or_none() and not dup_prod.scalar_one_or_none():
                data.codigo_barra = codigo
                break
    else:
        existing = await db.execute(
            select(MateriaLaboratorio).where(MateriaLaboratorio.codigo_barra == data.codigo_barra)
        )
        if existing.scalar_one_or_none():
            raise HTTPException(status_code=400, detail="Código de barra ya existe")
        
        existing_prod = await db.execute(
            select(Producto).where(Producto.codigo_barra == data.codigo_barra)
        )
        if existing_prod.scalar_one_or_none():
            raise HTTPException(status_code=400, detail="Código de barra ya existe en productos")
    
    materia = MateriaLaboratorio(**data.model_dump(), creado_en=now_paraguay())
    db.add(materia)
    await db.commit()
    await db.refresh(materia)
    return materia

@api_router.get("/materias-laboratorio", response_model=List[MateriaLaboratorioResponse])
async def listar_materias_laboratorio(empresa_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(MateriaLaboratorio).where(MateriaLaboratorio.empresa_id == empresa_id)
    )
    return result.scalars().all()

@api_router.get("/materias-laboratorio/disponibles", response_model=List[MateriaLaboratorioResponse])
async def listar_materias_disponibles(empresa_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(MateriaLaboratorio).where(
            MateriaLaboratorio.empresa_id == empresa_id,
            MateriaLaboratorio.estado == EstadoMateria.DISPONIBLE
        )
    )
    return result.scalars().all()

# ==================== ALMACENES ====================
@api_router.post("/almacenes", response_model=AlmacenResponse)
async def crear_almacen(data: AlmacenCreate, db: AsyncSession = Depends(get_db)):
    almacen = Almacen(**data.model_dump())
    db.add(almacen)
    await db.commit()
    await db.refresh(almacen)
    return almacen

@api_router.get("/almacenes", response_model=List[AlmacenResponse])
async def listar_almacenes(empresa_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Almacen).where(Almacen.empresa_id == empresa_id))
    return result.scalars().all()

# ==================== STOCK ====================
@api_router.post("/stock", response_model=StockActualResponse)
async def crear_actualizar_stock(data: StockActualCreate, db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(StockActual).where(
            StockActual.producto_id == data.producto_id,
            StockActual.almacen_id == data.almacen_id
        )
    )
    stock = result.scalar_one_or_none()
    
    if stock:
        stock.cantidad = data.cantidad
        if data.alerta_minima is not None:
            stock.alerta_minima = data.alerta_minima
    else:
        stock = StockActual(**data.model_dump())
        db.add(stock)
    
    await db.commit()
    await db.refresh(stock)
    return stock

@api_router.get("/stock", response_model=List[StockConDetalles])
async def listar_stock(empresa_id: int, almacen_id: Optional[int] = None, db: AsyncSession = Depends(get_db)):
    query = (
        select(StockActual, Producto, Almacen)
        .join(Producto, StockActual.producto_id == Producto.id)
        .join(Almacen, StockActual.almacen_id == Almacen.id)
        .where(Producto.empresa_id == empresa_id)
    )
    if almacen_id:
        query = query.where(StockActual.almacen_id == almacen_id)
    
    result = await db.execute(query)
    stocks = []
    for row in result.all():
        stock, producto, almacen = row
        stock_dict = StockActualResponse.model_validate(stock).model_dump()
        stock_dict['producto_nombre'] = producto.nombre
        stock_dict['almacen_nombre'] = almacen.nombre
        stocks.append(StockConDetalles(**stock_dict))
    
    return stocks


async def _consume_entrada_fifo(db: AsyncSession, producto_id: int, almacen_id: int, cantidad_consumida: int):
    """Reduce cantidad_restante on ENTRADA batches using FIFO (oldest batch first)."""
    result = await db.execute(
        select(MovimientoStock)
        .where(
            MovimientoStock.producto_id == producto_id,
            MovimientoStock.almacen_id == almacen_id,
            MovimientoStock.tipo == TipoMovimientoStock.ENTRADA,
            MovimientoStock.cantidad_restante > 0
        )
        .order_by(MovimientoStock.creado_en.asc())
    )
    batches = result.scalars().all()
    remaining = cantidad_consumida
    for batch in batches:
        if remaining <= 0:
            break
        consumable = min(batch.cantidad_restante, remaining)
        batch.cantidad_restante -= consumable
        remaining -= consumable


async def _restore_entrada_fifo(db: AsyncSession, producto_id: int, almacen_id: int, cantidad_restaurar: int):
    """Restore cantidad_restante on ENTRADA batches — newest first (FIFO reversal)."""
    result = await db.execute(
        select(MovimientoStock)
        .where(
            MovimientoStock.producto_id == producto_id,
            MovimientoStock.almacen_id == almacen_id,
            MovimientoStock.tipo == TipoMovimientoStock.ENTRADA,
            MovimientoStock.cantidad_restante < MovimientoStock.cantidad
        )
        .order_by(MovimientoStock.creado_en.desc())
    )
    batches = result.scalars().all()
    remaining = cantidad_restaurar
    for batch in batches:
        if remaining <= 0:
            break
        can_restore = min(batch.cantidad - batch.cantidad_restante, remaining)
        batch.cantidad_restante += can_restore
        remaining -= can_restore


async def _recalculate_precio_costo_fifo(db: AsyncSession, producto_id: int):
    """After a consumption, recalculate producto.precio_costo based on remaining active ENTRADA batches.
    - 1 active batch  → use its real costo_unitario
    - Multiple batches → weighted average CPP across all active batches
    - 0 active batches → no change
    """
    result = await db.execute(
        select(MovimientoStock)
        .where(
            MovimientoStock.producto_id == producto_id,
            MovimientoStock.tipo == TipoMovimientoStock.ENTRADA,
            MovimientoStock.cantidad_restante > 0,
            MovimientoStock.costo_unitario.isnot(None),
        )
        .order_by(MovimientoStock.creado_en.asc())
    )
    active_batches = result.scalars().all()

    if not active_batches:
        return  # nothing to recalculate

    if len(active_batches) == 1:
        new_costo = active_batches[0].costo_unitario
    else:
        total_qty = sum(b.cantidad_restante for b in active_batches)
        if total_qty == 0:
            return
        total_value = sum(
            Decimal(str(b.costo_unitario)) * Decimal(str(b.cantidad_restante))
            for b in active_batches
        )
        new_costo = round(total_value / Decimal(str(total_qty)))

    prod_result = await db.execute(select(Producto).where(Producto.id == producto_id))
    producto_obj = prod_result.scalar_one_or_none()
    if producto_obj:
        producto_obj.precio_costo = new_costo


@api_router.post("/stock/entrada")
async def entrada_stock(data: MovimientoStockCreate, db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(StockActual).where(
            StockActual.producto_id == data.producto_id,
            StockActual.almacen_id == data.almacen_id
        )
    )
    stock = result.scalar_one_or_none()
    
    if stock:
        stock.cantidad += data.cantidad
    else:
        stock = StockActual(
            producto_id=data.producto_id,
            almacen_id=data.almacen_id,
            cantidad=data.cantidad
        )
        db.add(stock)
    
    # Actualizar precio_costo y/o proveedor del producto si se proporcionan
    if data.costo_unitario is not None or data.costo_ponderado is not None or data.proveedor_id is not None or data.precio_venta is not None:
        prod_result = await db.execute(select(Producto).where(Producto.id == data.producto_id))
        producto_obj = prod_result.scalar_one_or_none()
        if producto_obj:
            # Use CPP if provided, otherwise fall back to real unit cost
            costo_para_producto = data.costo_ponderado if data.costo_ponderado is not None else data.costo_unitario
            if costo_para_producto is not None:
                producto_obj.precio_costo = costo_para_producto
            if data.proveedor_id is not None:
                producto_obj.proveedor_id = data.proveedor_id
            if data.precio_venta is not None:
                producto_obj.precio_venta = data.precio_venta

    movimiento = MovimientoStock(
        producto_id=data.producto_id,
        almacen_id=data.almacen_id,
        tipo=TipoMovimientoStock.ENTRADA,
        cantidad=data.cantidad,
        referencia_tipo=data.referencia_tipo,
        referencia_id=data.referencia_id,
        proveedor_id=data.proveedor_id,
        costo_unitario=data.costo_unitario,
        costo_ponderado=data.costo_ponderado,
        condicion_pago=data.condicion_pago,
        fecha_limite_pago=data.fecha_limite_pago,
        notas=data.notas,
        cantidad_restante=data.cantidad,
        creado_en=now_paraguay()
    )
    db.add(movimiento)

    # Si la condición de pago es crédito, crear deuda al proveedor
    if data.condicion_pago == 'credito' and data.proveedor_id and data.costo_unitario is not None:
        total_deuda = Decimal(str(data.costo_unitario)) * Decimal(str(data.cantidad))
        # Obtener nombre del producto para la descripción
        prod_res = await db.execute(select(Producto.nombre).where(Producto.id == data.producto_id))
        prod_nombre = prod_res.scalar_one_or_none() or 'Producto'
        deuda = DeudaProveedor(
            proveedor_id=data.proveedor_id,
            monto=total_deuda,
            descripcion=f"Entrada de stock: {data.cantidad} x {prod_nombre}" + (f" - {data.notas}" if data.notas else ""),
            fecha_emision=now_paraguay().date(),
            fecha_limite=data.fecha_limite_pago,
            pagado=False
        )
        db.add(deuda)
    
    await db.commit()
    await db.refresh(movimiento)
    return movimiento

@api_router.post("/stock/traspaso")
async def traspasar_stock(data: TraspasoStockCreate, db: AsyncSession = Depends(get_db)):
    result_origen = await db.execute(
        select(StockActual).where(
            StockActual.producto_id == data.producto_id,
            StockActual.almacen_id == data.almacen_origen_id
        )
    )
    stock_origen = result_origen.scalar_one_or_none()
    
    if not stock_origen or stock_origen.cantidad < data.cantidad:
        raise HTTPException(status_code=400, detail="Stock insuficiente en almacén origen")
    
    stock_origen.cantidad -= data.cantidad
    
    result_destino = await db.execute(
        select(StockActual).where(
            StockActual.producto_id == data.producto_id,
            StockActual.almacen_id == data.almacen_destino_id
        )
    )
    stock_destino = result_destino.scalar_one_or_none()
    
    if stock_destino:
        stock_destino.cantidad += data.cantidad
    else:
        stock_destino = StockActual(
            producto_id=data.producto_id,
            almacen_id=data.almacen_destino_id,
            cantidad=data.cantidad
        )
        db.add(stock_destino)
    
    mov_salida = MovimientoStock(
        producto_id=data.producto_id,
        almacen_id=data.almacen_origen_id,
        tipo=TipoMovimientoStock.TRASPASO,
        cantidad=-data.cantidad,
        creado_en=now_paraguay()
    )
    mov_entrada = MovimientoStock(
        producto_id=data.producto_id,
        almacen_id=data.almacen_destino_id,
        tipo=TipoMovimientoStock.TRASPASO,
        cantidad=data.cantidad,
        creado_en=now_paraguay()
    )
    db.add(mov_salida)
    db.add(mov_entrada)
    
    await db.commit()
    return {"message": "Traspaso realizado correctamente"}

@api_router.post("/stock/salida")
async def registrar_salida_stock(data: SalidaStockCreate, db: AsyncSession = Depends(get_db)):
    """Registra una salida/eliminación de stock"""
    
    result = await db.execute(
        select(StockActual).where(
            StockActual.producto_id == data.producto_id,
            StockActual.almacen_id == data.almacen_id
        )
    )
    stock = result.scalar_one_or_none()
    
    if not stock:
        raise HTTPException(status_code=404, detail="No hay stock disponible en este almacén")
    
    if stock.cantidad < data.cantidad:
        raise HTTPException(
            status_code=400, 
            detail=f"Stock insuficiente. Disponible: {stock.cantidad}, solicitado: {data.cantidad}"
        )
    
    stock.cantidad -= data.cantidad
    
    # Apply FIFO: consume from oldest active ENTRADA batches
    await _consume_entrada_fifo(db, data.producto_id, data.almacen_id, data.cantidad)
    await _recalculate_precio_costo_fifo(db, data.producto_id)

    # Register movement
    movimiento = MovimientoStock(
        producto_id=data.producto_id,
        almacen_id=data.almacen_id,
        tipo=TipoMovimientoStock.SALIDA,
        cantidad=-data.cantidad,
        creado_en=now_paraguay()
    )
    db.add(movimiento)

    await db.commit()
    return {"message": "Salida registrada correctamente", "stock_restante": stock.cantidad}

@api_router.put("/stock/{stock_id}/alerta", response_model=StockActualResponse)
async def configurar_alerta_stock(stock_id: int, alerta_minima: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(StockActual).where(StockActual.id == stock_id))
    stock = result.scalar_one_or_none()
    if not stock:
        raise HTTPException(status_code=404, detail="Stock no encontrado")
    
    stock.alerta_minima = alerta_minima
    await db.commit()
    await db.refresh(stock)
    return stock

@api_router.get("/stock/historial/{producto_id}", response_model=List[EntradaStockHistorialResponse])
async def historial_stock_producto(producto_id: int, db: AsyncSession = Depends(get_db)):
    """Retorna el historial de movimientos de stock para un producto específico"""
    result = await db.execute(
        select(MovimientoStock)
        .where(MovimientoStock.producto_id == producto_id)
        .order_by(MovimientoStock.creado_en.desc())
    )
    movimientos = result.scalars().all()

    response = []
    for mov in movimientos:
        # Obtener nombre del proveedor
        proveedor_nombre = None
        if mov.proveedor_id:
            prov_result = await db.execute(select(Proveedor.nombre).where(Proveedor.id == mov.proveedor_id))
            proveedor_nombre = prov_result.scalar_one_or_none()

        # Obtener nombre del almacén
        almacen_nombre = None
        if mov.almacen_id:
            alm_result = await db.execute(select(Almacen.nombre).where(Almacen.id == mov.almacen_id))
            almacen_nombre = alm_result.scalar_one_or_none()

        # Calcular total de compra
        total_compra = None
        if mov.costo_unitario is not None:
            total_compra = Decimal(str(mov.costo_unitario)) * Decimal(str(abs(mov.cantidad)))

        response.append(EntradaStockHistorialResponse(
            id=mov.id,
            producto_id=mov.producto_id,
            almacen_id=mov.almacen_id,
            tipo=mov.tipo,
            cantidad=mov.cantidad,
            costo_unitario=mov.costo_unitario,
            costo_ponderado=mov.costo_ponderado,
            total_compra=total_compra,
            condicion_pago=mov.condicion_pago,
            fecha_limite_pago=mov.fecha_limite_pago,
            notas=mov.notas,
            proveedor_id=mov.proveedor_id,
            proveedor_nombre=proveedor_nombre,
            almacen_nombre=almacen_nombre,
            referencia_tipo=mov.referencia_tipo,
            referencia_id=mov.referencia_id,
            creado_en=mov.creado_en,
            cantidad_restante=mov.cantidad_restante,
            estado=(
                "Activo" if (mov.cantidad_restante or 0) > 0
                else "Agotado" if mov.cantidad_restante == 0
                else "Activo"  # old records without tracking
            ) if mov.tipo == TipoMovimientoStock.ENTRADA else None
        ))

    return response

# ==================== VENTAS ====================
@api_router.post("/ventas", response_model=VentaResponse)
async def crear_venta(data: VentaCreate, crear_pendiente: bool = False, db: AsyncSession = Depends(get_db)):
    # Get client
    cliente_result = await db.execute(select(Cliente).where(Cliente.id == data.cliente_id))
    cliente = cliente_result.scalar_one_or_none()
    if not cliente:
        raise HTTPException(status_code=404, detail="Cliente no encontrado")
    
    # Si hay representante, usar sus privilegios (descuento y crédito)
    cliente_privilegios = cliente
    if data.representante_cliente_id:
        rep_result = await db.execute(select(Cliente).where(Cliente.id == data.representante_cliente_id))
        representante = rep_result.scalar_one_or_none()
        if representante:
            cliente_privilegios = representante
    
    # Si es venta pendiente, no validar stock ni crédito todavía
    if not crear_pendiente:
        # Validate cheque payment - usar privilegios del cliente efectivo
        if data.tipo_pago == TipoPago.CHEQUE and not cliente_privilegios.acepta_cheque:
            raise HTTPException(status_code=400, detail="Este cliente no tiene habilitado el pago con cheque")
    
    descuento_porcentaje = cliente_privilegios.descuento_porcentaje or Decimal('0')
    
    subtotal = Decimal('0')
    costo_total = Decimal('0')
    items_data = []
    
    # Solo validar stock si no es venta pendiente
    for item in data.items:
        item_total = Decimal(str(item.cantidad)) * item.precio_unitario
        subtotal += item_total

        # Fetch precio_costo for product items
        precio_costo = Decimal('0')
        if item.producto_id:
            prod_costo_res = await db.execute(
                select(Producto.precio_costo).where(Producto.id == item.producto_id)
            )
            raw_costo = prod_costo_res.scalar_one_or_none()
            precio_costo = Decimal(str(raw_costo or 0))
        costo_total += Decimal(str(item.cantidad)) * precio_costo

        items_data.append({
            **item.model_dump(),
            'total': item_total,
            'precio_costo': precio_costo,
        })
        
        if not crear_pendiente and item.producto_id:
            stock_result = await db.execute(
                select(func_sql.coalesce(func_sql.sum(StockActual.cantidad), 0))
                .where(StockActual.producto_id == item.producto_id)
            )
            stock_total = stock_result.scalar() or 0
            if stock_total < item.cantidad:
                prod_result = await db.execute(select(Producto).where(Producto.id == item.producto_id))
                prod = prod_result.scalar_one_or_none()
                raise HTTPException(
                    status_code=400, 
                    detail=f"Stock insuficiente para {prod.nombre if prod else 'producto'}. Disponible: {stock_total}"
                )
    
    descuento = subtotal * descuento_porcentaje / Decimal('100')
    subtotal_con_descuento = subtotal - descuento
    iva = subtotal_con_descuento * Decimal('10') / Decimal('110')
    total = subtotal_con_descuento
    ganancia = total - costo_total
    
    # Validar crédito si es venta a crédito y no es pendiente
    if not crear_pendiente and data.tipo_pago == TipoPago.CREDITO:
        # Calcular crédito usado del cliente que tiene los privilegios
        creditos_result = await db.execute(
            select(func_sql.coalesce(func_sql.sum(CreditoCliente.monto_pendiente), 0))
            .where(CreditoCliente.cliente_id == cliente_privilegios.id, CreditoCliente.pagado == False)
        )
        credito_usado = creditos_result.scalar() or Decimal('0')
        limite = cliente_privilegios.limite_credito or Decimal('0')
        
        if limite > 0 and (credito_usado + total) > limite:
            disponible = max(Decimal('0'), limite - credito_usado)
            raise HTTPException(
                status_code=400, 
                detail=f"Crédito insuficiente. Límite: {float(limite):,.0f}, Usado: {float(credito_usado):,.0f}, Disponible: {float(disponible):,.0f} Gs"
            )
    
    venta = Venta(
        empresa_id=data.empresa_id,
        cliente_id=data.cliente_id,
        usuario_id=data.usuario_id,
        representante_cliente_id=data.representante_cliente_id,
        total=total,
        iva=iva,
        descuento=descuento,
        costo_total=costo_total,
        ganancia=ganancia,
        tipo_pago=data.tipo_pago,
        es_delivery=data.es_delivery,
        estado=EstadoVenta.PENDIENTE if crear_pendiente else EstadoVenta.BORRADOR,
        creado_en=now_paraguay()
    )
    db.add(venta)
    await db.flush()
    
    for item_data in items_data:
        venta_item = VentaItem(venta_id=venta.id, **item_data)
        db.add(venta_item)
    
    await db.commit()
    await db.refresh(venta)
    return venta

@api_router.post("/ventas/{venta_id}/confirmar", response_model=VentaResponse)
async def confirmar_venta(venta_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(Venta).options(selectinload(Venta.items)).where(Venta.id == venta_id)
    )
    venta = result.scalar_one_or_none()
    if not venta:
        raise HTTPException(status_code=404, detail="Venta no encontrada")
    
    if venta.estado != EstadoVenta.BORRADOR:
        raise HTTPException(status_code=400, detail="La venta ya fue procesada")
    
    for item in venta.items:
        if item.producto_id:
            stock_result = await db.execute(
                select(StockActual)
                .where(StockActual.producto_id == item.producto_id, StockActual.cantidad > 0)
                .order_by(StockActual.cantidad.desc())
            )
            stocks = stock_result.scalars().all()
            
            cantidad_restante = item.cantidad
            for stock in stocks:
                if cantidad_restante <= 0:
                    break
                    
                a_descontar = min(stock.cantidad, cantidad_restante)
                stock.cantidad -= a_descontar
                cantidad_restante -= a_descontar
                
                await _consume_entrada_fifo(db, item.producto_id, stock.almacen_id, a_descontar)
                
                mov = MovimientoStock(
                    producto_id=item.producto_id,
                    almacen_id=stock.almacen_id,
                    tipo=TipoMovimientoStock.SALIDA,
                    cantidad=-a_descontar,
                    referencia_tipo='venta',
                    referencia_id=venta.id,
                    creado_en=now_paraguay()
                )
                db.add(mov)
        
            await _recalculate_precio_costo_fifo(db, item.producto_id)

        elif item.materia_laboratorio_id:
            materia_result = await db.execute(
                select(MateriaLaboratorio).where(MateriaLaboratorio.id == item.materia_laboratorio_id)
            )
            materia = materia_result.scalar_one_or_none()
            if materia:
                materia.estado = EstadoMateria.VENDIDO
    
    venta.estado = EstadoVenta.CONFIRMADA
    
    # Si es venta a crédito, crear registro de crédito
    if venta.tipo_pago == TipoPago.CREDITO:
        # Determinar el cliente que tiene los privilegios de crédito
        cliente_credito_id = venta.representante_cliente_id or venta.cliente_id
        
        credito = CreditoCliente(
            cliente_id=cliente_credito_id,
            venta_id=venta.id,
            monto_original=venta.total,
            monto_pendiente=venta.total,
            descripcion=f"Venta #{venta.id}",
            fecha_venta=today_paraguay()
        )
        db.add(credito)
    
    # Si es delivery, crear entrega automáticamente en estado PENDIENTE
    if venta.es_delivery:
        # Verificar si ya existe entrega para esta venta
        entrega_existente = await db.execute(
            select(Entrega).where(Entrega.venta_id == venta.id)
        )
        if not entrega_existente.scalar_one_or_none():
            entrega = Entrega(
                venta_id=venta.id,
                vehiculo_id=None,
                responsable_usuario_id=None,
                estado=EstadoEntrega.PENDIENTE
            )
            db.add(entrega)
    
    await db.commit()
    await db.refresh(venta)
    return venta

@api_router.put("/ventas/{venta_id}", response_model=VentaResponse)
async def actualizar_venta_pendiente(venta_id: int, data: VentaUpdate, db: AsyncSession = Depends(get_db)):
    """Permite editar una venta PENDIENTE: cambiar cliente, items, descuentos, tipo de pago"""
    result = await db.execute(
        select(Venta).options(selectinload(Venta.items)).where(Venta.id == venta_id)
    )
    venta = result.scalar_one_or_none()
    if not venta:
        raise HTTPException(status_code=404, detail="Venta no encontrada")
    
    if venta.estado != EstadoVenta.PENDIENTE:
        raise HTTPException(status_code=400, detail="Solo se pueden editar ventas PENDIENTES")
    
    # Update cliente if provided
    if data.cliente_id is not None:
        cliente_result = await db.execute(select(Cliente).where(Cliente.id == data.cliente_id))
        cliente = cliente_result.scalar_one_or_none()
        if not cliente:
            raise HTTPException(status_code=404, detail="Cliente no encontrado")
        venta.cliente_id = data.cliente_id
    
    # Update representante if provided
    if data.representante_cliente_id is not None:
        venta.representante_cliente_id = data.representante_cliente_id
    
    # Update tipo_pago if provided
    if data.tipo_pago is not None:
        venta.tipo_pago = data.tipo_pago
    
    # Update es_delivery if provided
    if data.es_delivery is not None:
        venta.es_delivery = data.es_delivery
    
    # Update items if provided
    if data.items is not None:
        # Delete existing items
        for item in venta.items:
            await db.delete(item)
        await db.flush()
        
        # Get cliente para descuento
        cliente_result = await db.execute(select(Cliente).where(Cliente.id == venta.cliente_id))
        cliente = cliente_result.scalar_one_or_none()
        
        # Si hay representante, usar sus privilegios
        cliente_privilegios = cliente
        if venta.representante_cliente_id:
            rep_result = await db.execute(select(Cliente).where(Cliente.id == venta.representante_cliente_id))
            representante = rep_result.scalar_one_or_none()
            if representante:
                cliente_privilegios = representante
        
        descuento_porcentaje = cliente_privilegios.descuento_porcentaje or Decimal('0')
        
        # Recalculate totals with new items
        subtotal = Decimal('0')
        costo_total = Decimal('0')
        for item in data.items:
            item_total = Decimal(str(item.cantidad)) * item.precio_unitario
            subtotal += item_total

            # Fetch precio_costo for product items
            precio_costo = Decimal('0')
            if item.producto_id:
                prod_costo_res = await db.execute(
                    select(Producto.precio_costo).where(Producto.id == item.producto_id)
                )
                raw_costo = prod_costo_res.scalar_one_or_none()
                precio_costo = Decimal(str(raw_costo or 0))
            costo_total += Decimal(str(item.cantidad)) * precio_costo
            
            venta_item = VentaItem(
                venta_id=venta.id,
                producto_id=item.producto_id,
                materia_laboratorio_id=item.materia_laboratorio_id,
                cantidad=item.cantidad,
                precio_unitario=item.precio_unitario,
                precio_costo=precio_costo,
                total=item_total,
                observaciones=item.observaciones
            )
            db.add(venta_item)
        
        descuento = subtotal * descuento_porcentaje / Decimal('100')
        subtotal_con_descuento = subtotal - descuento
        iva = subtotal_con_descuento * Decimal('10') / Decimal('110')
        
        venta.descuento = descuento
        venta.iva = iva
        venta.total = subtotal_con_descuento
        venta.costo_total = costo_total
        venta.ganancia = subtotal_con_descuento - costo_total
    
    await db.commit()
    await db.refresh(venta)
    return venta

@api_router.post("/ventas/{venta_id}/confirmar-pendiente", response_model=VentaResponse)
async def confirmar_venta_pendiente(venta_id: int, db: AsyncSession = Depends(get_db)):
    """Confirma una venta PENDIENTE: valida stock, genera movimientos, crea crédito si aplica"""
    result = await db.execute(
        select(Venta).options(selectinload(Venta.items)).where(Venta.id == venta_id)
    )
    venta = result.scalar_one_or_none()
    if not venta:
        raise HTTPException(status_code=404, detail="Venta no encontrada")
    
    if venta.estado != EstadoVenta.PENDIENTE:
        raise HTTPException(status_code=400, detail="Solo se pueden confirmar ventas PENDIENTES")
    
    # Get cliente para validaciones
    cliente_result = await db.execute(select(Cliente).where(Cliente.id == venta.cliente_id))
    cliente = cliente_result.scalar_one_or_none()
    
    # Si hay representante, usar sus privilegios
    cliente_privilegios = cliente
    if venta.representante_cliente_id:
        rep_result = await db.execute(select(Cliente).where(Cliente.id == venta.representante_cliente_id))
        representante = rep_result.scalar_one_or_none()
        if representante:
            cliente_privilegios = representante
    
    # Validate cheque payment
    if venta.tipo_pago == TipoPago.CHEQUE and not cliente_privilegios.acepta_cheque:
        raise HTTPException(status_code=400, detail="Este cliente no tiene habilitado el pago con cheque")
    
    # Validate stock for each item
    for item in venta.items:
        if item.producto_id:
            stock_result = await db.execute(
                select(func_sql.coalesce(func_sql.sum(StockActual.cantidad), 0))
                .where(StockActual.producto_id == item.producto_id)
            )
            stock_total = stock_result.scalar() or 0
            if stock_total < item.cantidad:
                prod_result = await db.execute(select(Producto).where(Producto.id == item.producto_id))
                prod = prod_result.scalar_one_or_none()
                raise HTTPException(
                    status_code=400, 
                    detail=f"Stock insuficiente para {prod.nombre if prod else 'producto'}. Disponible: {stock_total}"
                )
    
    # Validate crédito if applicable
    if venta.tipo_pago == TipoPago.CREDITO:
        creditos_result = await db.execute(
            select(func_sql.coalesce(func_sql.sum(CreditoCliente.monto_pendiente), 0))
            .where(CreditoCliente.cliente_id == cliente_privilegios.id, CreditoCliente.pagado == False)
        )
        credito_usado = creditos_result.scalar() or Decimal('0')
        limite = cliente_privilegios.limite_credito or Decimal('0')
        
        if limite > 0 and (credito_usado + venta.total) > limite:
            disponible = max(Decimal('0'), limite - credito_usado)
            raise HTTPException(
                status_code=400, 
                detail=f"Crédito insuficiente. Límite: {float(limite):,.0f}, Usado: {float(credito_usado):,.0f}, Disponible: {float(disponible):,.0f} Gs"
            )
    
    # Process stock movements
    for item in venta.items:
        if item.producto_id:
            stock_result = await db.execute(
                select(StockActual)
                .where(StockActual.producto_id == item.producto_id, StockActual.cantidad > 0)
                .order_by(StockActual.cantidad.desc())
            )
            stocks = stock_result.scalars().all()
            
            cantidad_restante = item.cantidad
            for stock in stocks:
                if cantidad_restante <= 0:
                    break
                    
                a_descontar = min(stock.cantidad, cantidad_restante)
                stock.cantidad -= a_descontar
                cantidad_restante -= a_descontar
                
                await _consume_entrada_fifo(db, item.producto_id, stock.almacen_id, a_descontar)
                
                mov = MovimientoStock(
                    producto_id=item.producto_id,
                    almacen_id=stock.almacen_id,
                    tipo=TipoMovimientoStock.SALIDA,
                    cantidad=-a_descontar,
                    referencia_tipo='venta',
                    referencia_id=venta.id,
                    creado_en=now_paraguay()
                )
                db.add(mov)

            await _recalculate_precio_costo_fifo(db, item.producto_id)

        elif item.materia_laboratorio_id:
            materia_result = await db.execute(
                select(MateriaLaboratorio).where(MateriaLaboratorio.id == item.materia_laboratorio_id)
            )
            materia = materia_result.scalar_one_or_none()
            if materia:
                materia.estado = EstadoMateria.VENDIDO
    
    venta.estado = EstadoVenta.CONFIRMADA
    
    # Si es venta a crédito, crear registro de crédito
    if venta.tipo_pago == TipoPago.CREDITO:
        cliente_credito_id = venta.representante_cliente_id or venta.cliente_id
        
        credito = CreditoCliente(
            cliente_id=cliente_credito_id,
            venta_id=venta.id,
            monto_original=venta.total,
            monto_pendiente=venta.total,
            descripcion=f"Venta #{venta.id}",
            fecha_venta=today_paraguay()
        )
        db.add(credito)
    
    # Si es delivery, crear entrega automáticamente en estado PENDIENTE
    if venta.es_delivery:
        entrega_existente = await db.execute(
            select(Entrega).where(Entrega.venta_id == venta.id)
        )
        if not entrega_existente.scalar_one_or_none():
            entrega = Entrega(
                venta_id=venta.id,
                vehiculo_id=None,
                responsable_usuario_id=None,
                estado=EstadoEntrega.PENDIENTE
            )
            db.add(entrega)
    
    await db.commit()
    await db.refresh(venta)
    return venta

@api_router.get("/ventas", response_model=List[VentaConDetalles])
async def listar_ventas(
    empresa_id: int,
    fecha_desde: Optional[str] = None,
    fecha_hasta: Optional[str] = None,
    cliente_id: Optional[int] = None,
    usuario_id: Optional[int] = None,
    estado: Optional[str] = None,
    monto_min: Optional[float] = None,
    monto_max: Optional[float] = None,
    db: AsyncSession = Depends(get_db)
):
    query = (
        select(Venta, Cliente)
        .join(Cliente, Venta.cliente_id == Cliente.id)
        .where(Venta.empresa_id == empresa_id)
        .order_by(Venta.creado_en.desc())
    )
    
    if fecha_desde:
        query = query.where(Venta.creado_en >= datetime.fromisoformat(fecha_desde))
    if fecha_hasta:
        query = query.where(Venta.creado_en <= datetime.fromisoformat(fecha_hasta))
    if cliente_id:
        query = query.where(Venta.cliente_id == cliente_id)
    if usuario_id:
        query = query.where(Venta.usuario_id == usuario_id)
    if estado:
        try:
            query = query.where(Venta.estado == EstadoVenta[estado])
        except KeyError:
            raise HTTPException(status_code=400, detail=f"Estado '{estado}' no válido")
    if monto_min:
        query = query.where(Venta.total >= monto_min)
    if monto_max:
        query = query.where(Venta.total <= monto_max)
    
    result = await db.execute(query)
    ventas = []
    for row in result.all():
        venta, cliente = row
        items_result = await db.execute(
            select(VentaItem).where(VentaItem.venta_id == venta.id)
        )
        raw_items = items_result.scalars().all()
        
        # Enrich items with product/materia names
        items = []
        for item in raw_items:
            item_dict = VentaItemResponse.model_validate(item).model_dump()
            
            # Get product name
            if item.producto_id:
                prod_result = await db.execute(select(Producto).where(Producto.id == item.producto_id))
                producto = prod_result.scalar_one_or_none()
                if producto:
                    item_dict['producto_nombre'] = producto.nombre
                    item_dict['descripcion'] = producto.nombre
            
            # Get materia name
            elif item.materia_laboratorio_id:
                mat_result = await db.execute(select(MateriaLaboratorio).where(MateriaLaboratorio.id == item.materia_laboratorio_id))
                materia = mat_result.scalar_one_or_none()
                if materia:
                    item_dict['materia_nombre'] = materia.nombre
                    item_dict['descripcion'] = f"{materia.nombre} - {materia.descripcion or ''}"
            
            items.append(item_dict)
        
        venta_dict = VentaResponse.model_validate(venta).model_dump()
        venta_dict['items'] = items
        venta_dict['cliente_nombre'] = f"{cliente.nombre} {cliente.apellido or ''}"
        venta_dict['cliente_ruc'] = cliente.ruc
        ventas.append(VentaConDetalles(**venta_dict))
    
    return ventas

@api_router.get("/ventas/{venta_id}", response_model=VentaConDetalles)
async def obtener_venta(venta_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(Venta, Cliente)
        .join(Cliente, Venta.cliente_id == Cliente.id)
        .where(Venta.id == venta_id)
    )
    row = result.first()
    if not row:
        raise HTTPException(status_code=404, detail="Venta no encontrada")
    
    venta, cliente = row
    items_result = await db.execute(
        select(VentaItem).where(VentaItem.venta_id == venta.id)
    )
    raw_items = items_result.scalars().all()
    
    # Enrich items with product/materia names
    items = []
    for item in raw_items:
        item_dict = VentaItemResponse.model_validate(item).model_dump()
        
        # Get product name
        if item.producto_id:
            prod_result = await db.execute(select(Producto).where(Producto.id == item.producto_id))
            producto = prod_result.scalar_one_or_none()
            if producto:
                item_dict['producto_nombre'] = producto.nombre
                item_dict['descripcion'] = producto.nombre
        
        # Get materia name
        elif item.materia_laboratorio_id:
            mat_result = await db.execute(select(MateriaLaboratorio).where(MateriaLaboratorio.id == item.materia_laboratorio_id))
            materia = mat_result.scalar_one_or_none()
            if materia:
                item_dict['materia_nombre'] = materia.nombre
                item_dict['descripcion'] = f"{materia.nombre} - {materia.descripcion or ''}"
        
        items.append(item_dict)
    
    venta_dict = VentaResponse.model_validate(venta).model_dump()
    venta_dict['items'] = items
    venta_dict['cliente_nombre'] = f"{cliente.nombre} {cliente.apellido or ''}"
    venta_dict['cliente_ruc'] = cliente.ruc
    return VentaConDetalles(**venta_dict)

@api_router.post("/ventas/{venta_id}/anular", response_model=VentaResponse)
async def anular_venta(venta_id: int, db: AsyncSession = Depends(get_db)):
    try:
        result = await db.execute(select(Venta).where(Venta.id == venta_id))
        venta = result.scalar_one_or_none()
        if not venta:
            raise HTTPException(status_code=404, detail="Venta no encontrada")
        
        if venta.estado == EstadoVenta.ANULADA:
            raise HTTPException(status_code=400, detail="La venta ya está anulada")
        
        # 1. Devolver stock de productos
        items_result = await db.execute(
            select(VentaItem).where(VentaItem.venta_id == venta_id)
        )
        items = items_result.scalars().all()
        
        for item in items:
            if item.producto_id:
                # Buscar los movimientos de VENTA/SALIDA de esta venta para devolver al mismo almacén
                movimientos_salida_result = await db.execute(
                    select(MovimientoStock).where(
                        MovimientoStock.referencia_tipo == 'venta',
                        MovimientoStock.referencia_id == venta_id,
                        MovimientoStock.producto_id == item.producto_id,
                        MovimientoStock.tipo == TipoMovimientoStock.SALIDA
                    )
                )
                movimientos_salida = movimientos_salida_result.scalars().all()
                
                # Devolver stock a los almacenes originales
                for mov_salida in movimientos_salida:
                    # Actualizar stock
                    stock_result = await db.execute(
                        select(StockActual).where(
                            StockActual.producto_id == item.producto_id,
                            StockActual.almacen_id == mov_salida.almacen_id
                        )
                    )
                    stock = stock_result.scalar_one_or_none()
                    
                    cantidad_devolver = abs(mov_salida.cantidad)
                    
                    if stock:
                        stock.cantidad += cantidad_devolver
                    else:
                        # Crear stock si no existe
                        stock = StockActual(
                            producto_id=item.producto_id,
                            almacen_id=mov_salida.almacen_id,
                            cantidad=cantidad_devolver
                        )
                        db.add(stock)
                    
                    # Restore FIFO batches
                    await _restore_entrada_fifo(db, item.producto_id, mov_salida.almacen_id, cantidad_devolver)
                    
                    # Registrar movimiento de ENTRADA (devolución)
                    movimiento = MovimientoStock(
                        almacen_id=mov_salida.almacen_id,
                        producto_id=item.producto_id,
                        tipo=TipoMovimientoStock.ENTRADA,
                        cantidad=cantidad_devolver,
                        referencia_tipo="ANULACION_VENTA",
                        referencia_id=venta_id,
                        cantidad_restante=cantidad_devolver,
                        creado_en=now_paraguay()
                    )
                    db.add(movimiento)
            
            elif item.materia_laboratorio_id:
                # Devolver materia de laboratorio a estado DISPONIBLE
                materia_result = await db.execute(
                    select(MateriaLaboratorio).where(MateriaLaboratorio.id == item.materia_laboratorio_id)
                )
                materia = materia_result.scalar_one_or_none()
                if materia:
                    materia.estado = EstadoMateria.DISPONIBLE
        
        # 2. Devolver crédito si fue venta a crédito
        if venta.tipo_pago == TipoPago.CREDITO:
            credito_result = await db.execute(
                select(CreditoCliente).where(CreditoCliente.venta_id == venta_id)
            )
            credito = credito_result.scalar_one_or_none()
            if credito:
                # Eliminar el crédito (o marcarlo como anulado si tiene pagos)
                pagos_result = await db.execute(
                    select(PagoCredito).where(PagoCredito.credito_id == credito.id)
                )
                pagos = pagos_result.scalars().all()
                
                if pagos:
                    # Si tiene pagos, marcar como anulado pero mantener el registro
                    credito.monto_pendiente = 0
                    credito.descripcion += " (ANULADO)"
                else:
                    # Si no tiene pagos, eliminar el crédito
                    await db.delete(credito)
        
        # 3. Marcar venta como anulada
        venta.estado = EstadoVenta.ANULADA
        
        # 4. Si tiene entrega asociada, siempre eliminarla (al anular venta)
        entrega_result = await db.execute(select(Entrega).where(Entrega.venta_id == venta_id))
        entrega = entrega_result.scalar_one_or_none()
        if entrega:
            await db.delete(entrega)
        
        await db.commit()
        await db.refresh(venta)
        return venta
    except HTTPException:
        raise
    except Exception as e:
        await db.rollback()
        import traceback
        error_detail = f"Error anulando venta: {str(e)}\n{traceback.format_exc()}"
        print(error_detail)  # Esto aparecerá en los logs
        raise HTTPException(status_code=500, detail=error_detail)

# ==================== BOLETA Y FACTURA ====================
@api_router.get("/ventas/{venta_id}/boleta")
async def generar_boleta(venta_id: int, db: AsyncSession = Depends(get_db)):
    """Generate boleta data for printing"""
    try:
        result = await db.execute(
            select(Venta, Cliente, Usuario, Empresa)
            .join(Cliente, Venta.cliente_id == Cliente.id)
            .join(Usuario, Venta.usuario_id == Usuario.id)
            .join(Empresa, Venta.empresa_id == Empresa.id)
            .where(Venta.id == venta_id)
        )
        row = result.first()
        if not row:
            logger.warning(f"Venta {venta_id} no encontrada o datos relacionados eliminados")
            raise HTTPException(status_code=404, detail="Venta no encontrada o datos relacionados faltantes")
        
        venta, cliente, usuario, empresa = row
        
        items_result = await db.execute(
            select(VentaItem, Producto, MateriaLaboratorio)
            .outerjoin(Producto, VentaItem.producto_id == Producto.id)
            .outerjoin(MateriaLaboratorio, VentaItem.materia_laboratorio_id == MateriaLaboratorio.id)
            .where(VentaItem.venta_id == venta_id)
        )
        
        items = []
        for item_row in items_result.all():
            item, producto, materia = item_row
            
            if producto:
                codigo = producto.codigo_barra or str(producto.id)
                descripcion = producto.nombre
            elif materia:
                codigo = f"LAB-{materia.id}"
                descripcion = f"{materia.nombre} - {materia.descripcion or ''}"
            else:
                codigo = 'N/A'
                descripcion = 'Item'
            
            items.append({
                'codigo': codigo,
                'cantidad': float(item.cantidad),
                'descripcion': descripcion,
                'iva': 10,
                'precio': float(item.precio_unitario),
                'total': float(item.total)
            })
        
        return {
            'tipo': 'BOLETA',
            'numero': venta.id,
            'fecha': venta.creado_en.strftime('%d/%m/%Y %I:%M %p'),
            'tipo_pago': venta.tipo_pago.value if venta.tipo_pago else 'CONTADO',
            'empresa': {
                'nombre': empresa.nombre,
                'ruc': empresa.ruc,
                'telefono': empresa.telefono,
                'direccion': empresa.direccion
            },
            'cliente': {
                'nombre': f"{cliente.nombre} {cliente.apellido or ''}".strip(),
                'ruc': cliente.ruc or '0',
                'direccion': cliente.direccion or '0',
                'telefono': cliente.telefono or '0'
            },
            'vendedor': f"{usuario.nombre} {usuario.apellido or ''}".strip(),
            'items': items,
            'subtotal_sin_descuento': float(venta.total + venta.descuento),
            'descuento': float(venta.descuento),
            'descuento_porcentaje': float(cliente.descuento_porcentaje) if cliente.descuento_porcentaje else 0,
            'subtotal': float(venta.total + venta.descuento),
            'iva': float(venta.iva),
            'total': float(venta.total),
            'total_letras': numero_a_letras(int(venta.total)) + ' Guaraníes'
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error generando datos de boleta para venta {venta_id}: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Error al generar boleta: {str(e)}")

@api_router.get("/ventas/{venta_id}/factura")
async def generar_factura(venta_id: int, db: AsyncSession = Depends(get_db)):
    """Generate factura data for printing"""
    try:
        result = await db.execute(
            select(Venta, Cliente, Usuario, Empresa)
            .join(Cliente, Venta.cliente_id == Cliente.id)
            .join(Usuario, Venta.usuario_id == Usuario.id)
            .join(Empresa, Venta.empresa_id == Empresa.id)
            .where(Venta.id == venta_id)
        )
        row = result.first()
        if not row:
            logger.warning(f"Venta {venta_id} no encontrada o datos relacionados eliminados")
            raise HTTPException(status_code=404, detail="Venta no encontrada o datos relacionados faltantes")
        
        venta, cliente, usuario, empresa = row
        
        # Check if client has RUC for factura
        if not cliente.ruc:
            raise HTTPException(status_code=400, detail="El cliente no tiene RUC para emitir factura")
        
        items_result = await db.execute(
            select(VentaItem, Producto, MateriaLaboratorio)
            .outerjoin(Producto, VentaItem.producto_id == Producto.id)
            .outerjoin(MateriaLaboratorio, VentaItem.materia_laboratorio_id == MateriaLaboratorio.id)
            .where(VentaItem.venta_id == venta_id)
        )
        
        items = []
        for item_row in items_result.all():
            item, producto, materia = item_row
            
            if producto:
                descripcion = producto.nombre
            elif materia:
                descripcion = f"{materia.nombre} - {materia.descripcion or ''}"
            else:
                descripcion = 'Item'
            
            items.append({
                'cantidad': float(item.cantidad),
                'descripcion': descripcion,
                'precio_unitario': float(item.precio_unitario),
                'exenta': 0,
                'iva_5': 0,
                'iva_10': float(item.total)
            })
        
        # Calculate IVA breakdown
        base_imponible = float(venta.total) / 1.10
        iva_10 = float(venta.total) - base_imponible
        
        return {
            'tipo': 'FACTURA',
            'numero': f"{venta.id:07d}",
            'fecha': venta.creado_en.strftime('%d de %B de %Y'),
            'condicion': venta.tipo_pago.value if venta.tipo_pago else 'CONTADO',
            'empresa': {
                'nombre': empresa.nombre,
                'ruc': empresa.ruc,
                'telefono': empresa.telefono,
                'direccion': empresa.direccion
            },
            'cliente': {
                'nombre': f"{cliente.nombre} {cliente.apellido or ''}".strip(),
                'ruc': cliente.ruc,
                'direccion': cliente.direccion or '',
                'telefono': cliente.telefono or ''
            },
            'items': items,
            'subtotal_exenta': 0,
            'subtotal_iva_5': 0,
            'subtotal_iva_10': float(venta.total + venta.descuento),
            'descuento': float(venta.descuento),
            'descuento_porcentaje': float(cliente.descuento_porcentaje) if cliente.descuento_porcentaje else 0,
            'iva_10': round(iva_10, 0),
            'total': float(venta.total),
            'total_letras': numero_a_letras(int(venta.total)),
            'creado_en': venta.creado_en.isoformat(),
            'liquidacion_iva': {
                'iva_5': 0,
                'iva_10': round(iva_10, 0),
                'total_iva': round(iva_10, 0)
            }
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error generando datos de factura para venta {venta_id}: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Error al generar factura: {str(e)}")

# ==================== DOCUMENTOS TEMPORALES ====================
# Configuración de directorio para documentos temporales
# IMPORTANTE: Usar directorio persistente dentro de uploads/ para que los enlaces
# funcionen por 30 días. NO usar /tmp ya que se limpia automáticamente.
DOCS_DIR = ROOT_DIR / 'uploads' / 'documentos'
DOCS_DIR.mkdir(parents=True, exist_ok=True)
logger.info(f"Directorio de documentos temporales (persistente): {DOCS_DIR}")

@api_router.post("/ventas/{venta_id}/generar-enlace")
async def generar_enlace_documento(
    venta_id: int,
    tipo_documento: TipoDocumento = Query(...),
    db: AsyncSession = Depends(get_db)
):
    """
    Genera un PDF del documento (boleta o factura), lo guarda en disco
    y retorna un enlace público de descarga válido por 30 días
    """
    try:
        logger.info(f"Iniciando generación de enlace para venta {venta_id}, tipo {tipo_documento}")
        
        # Verificar que la venta exista y esté confirmada
        result = await db.execute(
            select(Venta).where(Venta.id == venta_id)
        )
        venta = result.scalar_one_or_none()
        if not venta:
            logger.warning(f"Venta {venta_id} no encontrada")
            raise HTTPException(status_code=404, detail="Venta no encontrada")
        
        logger.info(f"Venta encontrada con estado: {venta.estado}")
        
        if venta.estado != EstadoVenta.CONFIRMADA:
            logger.warning(f"Venta {venta_id} no está confirmada, estado: {venta.estado}")
            raise HTTPException(status_code=400, detail="Solo se pueden generar enlaces para ventas confirmadas")
        
        logger.info("Verificando documentos existentes...")
        
        # Verificar si ya existe un documento temporal para esta venta y tipo
        existing = await db.execute(
            select(DocumentoTemporal).where(
                and_(
                    DocumentoTemporal.venta_id == venta_id,
                    DocumentoTemporal.tipo_documento == tipo_documento,
                    DocumentoTemporal.fecha_expiracion > now_paraguay()
                )
            )
        )
        existing_doc = existing.scalar_one_or_none()
        
        # Si existe y no ha expirado, verificar que el archivo también exista en disco
        if existing_doc:
            file_exists = Path(existing_doc.file_path).exists() if existing_doc.file_path else False
            if file_exists:
                logger.info(f"Documento existente encontrado con archivo en disco: {existing_doc.token}")
                base_url = os.getenv("BACKEND_URL", "http://localhost:8000")
                return {
                    "url": f"{base_url}/api/documentos/{existing_doc.token}",
                    "token": existing_doc.token,
                    "tipo_documento": tipo_documento.value,
                    "fecha_expiracion": existing_doc.fecha_expiracion,
                    "ya_existia": True
                }
            else:
                # El registro existe pero el archivo fue eliminado (Render limpió el filesystem)
                # Eliminar el registro viejo para regenerar abajo
                logger.warning(f"Documento {existing_doc.token} existe en BD pero archivo no encontrado en disco. Regenerando...")
                await db.delete(existing_doc)
                await db.commit()
        
        logger.info("Generando nuevo documento PDF...")
        
        # Generar datos del documento según tipo
        if tipo_documento == TipoDocumento.BOLETA:
            # Reutilizar lógica del endpoint /ventas/{venta_id}/boleta
            result = await db.execute(
                select(Venta, Cliente, Usuario, Empresa)
                .join(Cliente, Venta.cliente_id == Cliente.id)
                .join(Usuario, Venta.usuario_id == Usuario.id)
                .join(Empresa, Venta.empresa_id == Empresa.id)
                .where(Venta.id == venta_id)
            )
            row = result.first()
            if not row:
                raise HTTPException(status_code=404, detail="Error al cargar datos de venta")
            
            venta_obj, cliente, vendedor, empresa = row
            pdf_content = await generar_pdf_boleta(venta_obj, cliente, vendedor, empresa, db)
            filename_prefix = f"boleta_{venta_id}"
        else:  # FACTURA
            result = await db.execute(
                select(Venta, Cliente, Usuario, Empresa)
                .join(Cliente, Venta.cliente_id == Cliente.id)
                .join(Usuario, Venta.usuario_id == Usuario.id)
                .join(Empresa, Venta.empresa_id == Empresa.id)
                .where(Venta.id == venta_id)
            )
            row = result.first()
            if not row:
                raise HTTPException(status_code=404, detail="Error al cargar datos de venta")
            
            venta_obj, cliente, vendedor, empresa = row
            
            if not cliente.ruc:
                raise HTTPException(status_code=400, detail="El cliente no tiene RUC para emitir factura")
            
            pdf_content = await generar_pdf_factura(venta_obj, cliente, vendedor, empresa, db)
            filename_prefix = f"factura_{venta_id}"
        
        logger.info(f"PDF generado exitosamente, tamaño: {len(pdf_content)} bytes")
        
        # Generar token único
        token = str(uuid.uuid4())
        filename = f"{filename_prefix}_{token}.pdf"
        file_path = DOCS_DIR / filename
        
        logger.info(f"Guardando PDF en: {file_path}")
        
        # Guardar PDF en disco
        with open(file_path, "wb") as f:
            f.write(pdf_content)
        
        logger.info(f"PDF guardado exitosamente")
        
        # Crear registro en base de datos
        fecha_expiracion = now_paraguay() + timedelta(days=30)
        logger.info(f"Creando registro en base de datos...")
        
        nuevo_doc = DocumentoTemporal(
            token=token,
            venta_id=venta_id,
            tipo_documento=tipo_documento,
            file_path=str(file_path),
            fecha_creacion=now_paraguay(),
            fecha_expiracion=fecha_expiracion,
            empresa_id=venta.empresa_id
        )
        
        db.add(nuevo_doc)
        await db.commit()
        await db.refresh(nuevo_doc)
        
        logger.info(f"Documento guardado en BD con ID: {nuevo_doc.id}, token: {token}")
        
        # Retornar URL pública
        base_url = os.getenv("BACKEND_URL", "http://localhost:8000")
        logger.info(f"Enlace generado: {base_url}/api/documentos/{token}")
        
        return {
            "url": f"{base_url}/api/documentos/{token}",
            "token": token,
            "tipo_documento": tipo_documento.value,
            "fecha_expiracion": fecha_expiracion,
            "ya_existia": False
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error generando enlace de documento: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Error al generar documento: {str(e)}")


@api_router.get("/documentos/{token}")
async def descargar_documento(token: str, db: AsyncSession = Depends(get_db)):
    """
    Endpoint público para descargar documentos mediante token único.
    No requiere autenticación - cualquiera con el enlace puede descargar.
    Si el archivo no existe en disco (ej: Render limpió el filesystem),
    se regenera automáticamente desde los datos de la venta en la BD.
    """
    # Buscar documento por token
    result = await db.execute(
        select(DocumentoTemporal, Venta)
        .join(Venta, DocumentoTemporal.venta_id == Venta.id)
        .where(DocumentoTemporal.token == token)
    )
    row = result.first()
    
    if not row:
        raise HTTPException(status_code=404, detail="Documento no encontrado")
    
    documento, venta = row
    
    # Verificar que no haya expirado
    if documento.fecha_expiracion < now_paraguay():
        raise HTTPException(status_code=410, detail="Este enlace ha expirado")
    
    # Verificar que el archivo existe; si no, regenerarlo desde los datos de la venta
    file_path = Path(documento.file_path)
    if not file_path.exists():
        logger.warning(f"Archivo no encontrado en disco: {file_path}. Regenerando PDF desde datos de venta...")
        try:
            # Cargar datos necesarios para regenerar el PDF
            regen_result = await db.execute(
                select(Venta, Cliente, Usuario, Empresa)
                .join(Cliente, Venta.cliente_id == Cliente.id)
                .join(Usuario, Venta.usuario_id == Usuario.id)
                .join(Empresa, Venta.empresa_id == Empresa.id)
                .where(Venta.id == documento.venta_id)
            )
            regen_row = regen_result.first()
            if not regen_row:
                raise HTTPException(status_code=404, detail="No se pudieron cargar los datos de la venta para regenerar el documento")
            
            venta_obj, cliente, vendedor, empresa = regen_row
            
            if documento.tipo_documento == TipoDocumento.BOLETA:
                pdf_content = await generar_pdf_boleta(venta_obj, cliente, vendedor, empresa, db)
            else:  # FACTURA
                pdf_content = await generar_pdf_factura(venta_obj, cliente, vendedor, empresa, db)
            
            # Asegurar que el directorio existe (puede haberse perdido también)
            DOCS_DIR.mkdir(parents=True, exist_ok=True)
            
            # Guardar el PDF regenerado en disco
            with open(file_path, "wb") as f:
                f.write(pdf_content)
            
            logger.info(f"PDF regenerado exitosamente y guardado en: {file_path} ({len(pdf_content)} bytes)")
        except HTTPException:
            raise
        except Exception as e:
            logger.error(f"Error al regenerar PDF para token {token}: {str(e)}", exc_info=True)
            raise HTTPException(status_code=500, detail="Error al regenerar el documento. Intente generar un nuevo enlace.")
    
    # Incrementar contador de descargas
    documento.descargas += 1
    await db.commit()
    
    # Servir el archivo
    tipo_doc = "Boleta" if documento.tipo_documento == TipoDocumento.BOLETA else "Factura"
    filename = f"{tipo_doc}_Venta_{venta.id}.pdf"
    
    return FileResponse(
        path=str(file_path),
        media_type="application/pdf",
        filename=filename
    )


@api_router.delete("/documentos/limpiar-expirados")
async def limpiar_documentos_expirados(
    db: AsyncSession = Depends(get_db)
):
    """
    Tarea de limpieza: Elimina documentos expirados tanto de la BD como del disco.
    Endpoint para ser llamado por cronjobs o tareas programadas.
    """
    # Buscar documentos expirados
    result = await db.execute(
        select(DocumentoTemporal).where(
            DocumentoTemporal.fecha_expiracion < now_paraguay()
        )
    )
    documentos_expirados = result.scalars().all()
    
    eliminados_bd = 0
    eliminados_disco = 0
    errores = []
    
    for doc in documentos_expirados:
        try:
            # Eliminar archivo del disco
            file_path = Path(doc.file_path)
            if file_path.exists():
                file_path.unlink()
                eliminados_disco += 1
        except Exception as e:
            errores.append(f"Error al eliminar {doc.file_path}: {str(e)}")
        
        # Eliminar registro de BD
        await db.delete(doc)
        eliminados_bd += 1
    
    await db.commit()
    
    return {
        "mensaje": "Limpieza completada",
        "documentos_eliminados_bd": eliminados_bd,
        "archivos_eliminados_disco": eliminados_disco,
        "errores": errores
    }


# Funciones auxiliares para generar PDFs
async def generar_pdf_boleta(venta: Venta, cliente: Cliente, vendedor: Usuario, empresa: Empresa, db: AsyncSession) -> bytes:
    """Genera PDF de boleta con formato similar a matriz de puntos"""
    # Obtener items
    items_result = await db.execute(
        select(VentaItem, Producto, MateriaLaboratorio)
        .outerjoin(Producto, VentaItem.producto_id == Producto.id)
        .outerjoin(MateriaLaboratorio, VentaItem.materia_laboratorio_id == MateriaLaboratorio.id)
        .where(VentaItem.venta_id == venta.id)
    )
    
    buffer = io.BytesIO()

    elements = []

    # Estilos con fuente monoespaciada (Courier) optimizados para térmica 75mm
    title_style = ParagraphStyle(
        'CustomTitle',
        fontName='Courier-Bold',
        fontSize=14,
        alignment=TA_CENTER,
        spaceAfter=2,
        textDecoration='underline'
    )

    normal_style = ParagraphStyle(
        'CustomNormal',
        fontName='Courier',
        fontSize=8,
        alignment=TA_CENTER,
        spaceAfter=1
    )

    # Encabezado (usar nombre de la empresa si está disponible)
    elements.append(Paragraph(empresa.nombre or 'Luz Brill', title_style))
    if empresa.telefono:
        elements.append(Paragraph(empresa.telefono, normal_style))
    elements.append(Spacer(1, 3*mm))
    
    # Número de nota (alineado a la derecha)
    nota_style = ParagraphStyle('Nota', fontName='Courier-Bold', fontSize=10, alignment=TA_RIGHT)
    elements.append(Paragraph(f'<b>NOTA NRO: {venta.id}</b>', nota_style))
    elements.append(Spacer(1, 2*mm))
    
    # Información del cliente
    info_data = [
        ['Razon Social:', f"{cliente.nombre} {cliente.apellido or ''}".strip()],
        ['Dirección:', cliente.direccion or '0'],
        ['Telefono:', cliente.telefono or '0'],
        ['Ruc:', cliente.ruc or '0'],
        ['Fecha de Venta:', venta.creado_en.strftime('%d/%m/%Y %I:%M %p')],
        ['Tipo Comprob:', venta.tipo_pago.value if venta.tipo_pago else 'CONTADO']
    ]
    
    # Ajuste de columnas para ancho de 75mm (rollo térmico)
    info_table = Table(info_data, colWidths=[30*mm, 40*mm])
    info_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Courier-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('LINEABOVE', (0, 0), (-1, 0), 2, colors.black),
        ('LINEBELOW', (0, -1), (-1, -1), 2, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 3*mm))
    
    # Items
    items_data = [['Cod', 'Cant', 'Descripción', 'IVA', 'Precio', 'Total']]
    
    for item_row in items_result.all():
        item, producto, materia = item_row
        
        if producto:
            codigo = producto.codigo_barra or str(producto.id)
            descripcion = producto.nombre
        elif materia:
            codigo = f"LAB-{materia.id}"
            descripcion = f"{materia.nombre} - {materia.descripcion or ''}"
        else:
            codigo = 'N/A'
            descripcion = 'Item'
        
        items_data.append([
            codigo[:10],  # Limitar código
            f"{item.cantidad:.2f}",
            descripcion[:30],  # Limitar descripción para caber en 75mm
            '10',
            f"{int(item.precio_unitario):,}",
            f"{int(item.total):,}"
        ])
    
    # Column widths recalculadas para 75mm total
    items_table = Table(items_data, colWidths=[12*mm, 10*mm, 30*mm, 6*mm, 8*mm, 9*mm])
    items_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, 0), 'Courier-Bold'),
        ('FONTNAME', (0, 1), (-1, -1), 'Courier'),
        ('FONTSIZE', (0, 0), (-1, -1), 7),
        ('LINEABOVE', (0, 0), (-1, 0), 2, colors.black),
        ('LINEBELOW', (0, 0), (-1, 0), 2, colors.black),
        ('LINEBELOW', (0, -1), (-1, -1), 1, colors.grey),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),  # Código
        ('ALIGN', (1, 0), (1, -1), 'CENTER'),  # Cantidad
        ('ALIGN', (2, 0), (2, -1), 'LEFT'),  # Descripción
        ('ALIGN', (3, 0), (3, -1), 'CENTER'),  # IVA
        ('ALIGN', (4, 0), (-1, -1), 'RIGHT'),  # Precio y Total
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
    ]))
    elements.append(items_table)
    elements.append(Spacer(1, 3*mm))
    
    # Totales
    subtotal_sin_descuento = float(venta.total + venta.descuento)
    descuento_porcentaje = float(cliente.descuento_porcentaje) if cliente.descuento_porcentaje else 0
    total_letras = numero_a_letras(int(venta.total)) + ' Guaraníes'
    
    totales_data = [
        ['Subtotal:', f"{int(subtotal_sin_descuento):,}"]
    ]
    
    if venta.descuento > 0:
        totales_data.append([f'Descuento ({descuento_porcentaje}%):', f"-{int(venta.descuento):,}"])
    
    totales_data.append(['En Letras:', total_letras.upper()])
    totales_data.append(['TOTAL A PAGAR:', f'Gs. {int(venta.total):,}'])
    
    # Totales: ajustar anchos para rollo 75mm
    totales_table = Table(totales_data, colWidths=[50*mm, 25*mm])
    totales_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -2), 'Courier-Bold'),
        ('FONTNAME', (0, -1), (-1, -1), 'Courier-Bold'),
        ('FONTSIZE', (0, 0), (-1, -2), 8),
        ('FONTSIZE', (0, -1), (-1, -1), 10),
        ('LINEABOVE', (0, 0), (-1, 0), 2, colors.black),
        ('LINEABOVE', (0, -1), (-1, -1), 2, colors.black),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
    ]))
    elements.append(totales_table)
    elements.append(Spacer(1, 5*mm))
    
    # Firma
    firma_style = ParagraphStyle('Firma', fontName='Courier', fontSize=8, alignment=TA_CENTER)
    elements.append(Paragraph('_' * 40, firma_style))
    elements.append(Paragraph('<b>Firma Cliente</b>', firma_style))
    elements.append(Spacer(1, 2*mm))
    
    footer_style = ParagraphStyle('Footer', fontName='Courier', fontSize=7, alignment=TA_CENTER, fontStyle='italic')
    elements.append(Paragraph('<b>Favor Conferir Su Mercaderia !!! No Aceptamos Reclamos Posteriores.</b>', footer_style))
    
    # Calcular altura dinámica según cantidad de filas para rollo continuo
    rows_count = max(1, len(items_data))
    per_row_mm = 6  # espacio por fila
    base_mm = 50    # espacio fijo para encabezado/totales
    totals_mm = 30
    page_height = (base_mm + rows_count * per_row_mm + totals_mm) * mm

    # Crear documento con ancho 75mm y altura calculada
    page_width = 75 * mm
    custom_size = (page_width, page_height)

    doc = SimpleDocTemplate(
        buffer,
        pagesize=custom_size,
        leftMargin=4*mm,
        rightMargin=4*mm,
        topMargin=4*mm,
        bottomMargin=4*mm
    )

    doc.build(elements)
    buffer.seek(0)
    return buffer.read()


async def generar_pdf_factura(venta: Venta, cliente: Cliente, vendedor: Usuario, empresa: Empresa, db: AsyncSession) -> bytes:
    """Genera PDF de factura con formato similar a matriz de puntos"""
    # Obtener items
    items_result = await db.execute(
        select(VentaItem, Producto, MateriaLaboratorio)
        .outerjoin(Producto, VentaItem.producto_id == Producto.id)
        .outerjoin(MateriaLaboratorio, VentaItem.materia_laboratorio_id == MateriaLaboratorio.id)
        .where(VentaItem.venta_id == venta.id)
    )
    
    buffer = io.BytesIO()
    # Tamaño de papel: 240mm x 200mm para factura
    page_width = 240 * mm
    page_height = 200 * mm
    custom_size = (page_width, page_height)
    
    doc = SimpleDocTemplate(
        buffer, 
        pagesize=custom_size,
        leftMargin=12*mm,
        rightMargin=12*mm,
        topMargin=10*mm,
        bottomMargin=10*mm
    )
    
    elements = []
    
    # Estilos
    title_style = ParagraphStyle(
        'CustomTitle',
        fontName='Courier-Bold',
        fontSize=16,
        alignment=TA_LEFT,
        spaceAfter=1,
        textDecoration='underline'
    )
    
    normal_bold = ParagraphStyle(
        'NormalBold',
        fontName='Courier-Bold',
        fontSize=9,
        alignment=TA_LEFT
    )
    
    # Header con empresa y número de factura
    header_data = [
        [
            Paragraph(f'<b>{empresa.nombre or "Luz Brill S.A."}</b>', title_style),
            Paragraph('<b>FACTURA</b>', ParagraphStyle('FacturaTit', fontName='Courier-Bold', fontSize=14, alignment=TA_CENTER))
        ],
        [
            Paragraph(f'<b>RUC:</b> {empresa.ruc}', normal_bold),
            Paragraph(f'<b>N° {venta.id:07d}</b>', ParagraphStyle('FacturaNum', fontName='Courier-Bold', fontSize=11, alignment=TA_CENTER))
        ],
        [
            Paragraph(f'<b>{empresa.direccion}</b>', normal_bold),
            ''
        ],
        [
            Paragraph(f'<b>Tel:</b> {empresa.telefono}', normal_bold),
            ''
        ]
    ]
    
    header_table = Table(header_data, colWidths=[140*mm, 56*mm])
    header_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('BOX', (1, 0), (1, 1), 2, colors.black),
        ('ALIGN', (1, 0), (1, 1), 'CENTER'),
        ('VALIGN', (1, 0), (1, 1), 'MIDDLE'),
    ]))
    elements.append(header_table)
    elements.append(Spacer(1, 4*mm))
    
    # Información del cliente
    cliente_data = [
        [
            f'Señor(es): {cliente.nombre} {cliente.apellido or ""}',
            f'Fecha: {venta.creado_en.strftime("%d/%m/%Y")}'
        ],
        [
            f'Dirección: {cliente.direccion or ""}',
            f'RUC: {cliente.ruc}'
        ],
        [
            f'Teléfono: {cliente.telefono or ""}',
            f'Condición: {venta.tipo_pago.value if venta.tipo_pago else "CONTADO"}'
        ]
    ]
    
    cliente_table = Table(cliente_data, colWidths=[140*mm, 56*mm])
    cliente_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Courier-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('BOX', (0, 0), (-1, -1), 2, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
    ]))
    elements.append(cliente_table)
    elements.append(Spacer(1, 3*mm))
    
    # Items
    items_data = [['Cant.', 'Descripción', 'P. Unit.', 'Exenta', 'IVA 5%', 'IVA 10%']]
    subtotal_iva_10 = 0
    
    for item_row in items_result.all():
        item, producto, materia = item_row
        
        if producto:
            descripcion = producto.nombre
        elif materia:
            descripcion = f"{materia.nombre} - {materia.descripcion or ''}"
        else:
            descripcion = 'Item'
        
        subtotal_iva_10 += float(item.total)
        
        items_data.append([
            f"{item.cantidad:.2f}",
            descripcion[:40],  # Limitar descripción
            f"{int(item.precio_unitario):,}",
            '0',
            '0',
            f"{int(item.total):,}"
        ])
    
    items_table = Table(items_data, colWidths=[18*mm, 95*mm, 25*mm, 18*mm, 18*mm, 25*mm])
    items_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, 0), 'Courier-Bold'),
        ('FONTNAME', (0, 1), (-1, -1), 'Courier'),
        ('FONTSIZE', (0, 0), (-1, -1), 7),
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('LINEABOVE', (0, 0), (-1, 0), 2, colors.black),
        ('LINEBELOW', (0, 0), (-1, 0), 2, colors.black),
        ('LINEBELOW', (0, -1), (-1, -1), 1, colors.grey),
        ('ALIGN', (0, 0), (0, -1), 'CENTER'),  # Cantidad
        ('ALIGN', (1, 0), (1, -1), 'LEFT'),  # Descripción
        ('ALIGN', (2, 0), (-1, -1), 'RIGHT'),  # Resto a la derecha
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
    ]))
    elements.append(items_table)
    elements.append(Spacer(1, 3*mm))
    
    # Totales
    base_imponible = float(venta.total) / 1.10
    iva_10 = float(venta.total) - base_imponible
    descuento_porcentaje = float(cliente.descuento_porcentaje) if cliente.descuento_porcentaje else 0
    total_letras = numero_a_letras(int(venta.total))
    
    totales_data = [
        ['Subtotal IVA 10%:', f"Gs. {int(subtotal_iva_10 + venta.descuento):,}"]
    ]
    
    if venta.descuento > 0:
        totales_data.append([f'Descuento ({descuento_porcentaje}%):', f"-Gs. {int(venta.descuento):,}"])
    
    totales_data.extend([
        ['Liquidación IVA 10%:', f"Gs. {round(iva_10, 0):,.0f}"],
        ['Son:', f"{total_letras.upper()} GUARANÍES"],
        ['TOTAL A PAGAR:', f'Gs. {int(venta.total):,}']
    ])
    
    totales_table = Table(totales_data, colWidths=[100*mm, 80*mm])
    totales_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -2), 'Courier-Bold'),
        ('FONTNAME', (0, -1), (-1, -1), 'Courier-Bold'),
        ('FONTSIZE', (0, 0), (-1, -2), 8),
        ('FONTSIZE', (0, -1), (-1, -1), 10),
        ('BOX', (0, 0), (-1, -1), 2, colors.black),
        ('LINEABOVE', (0, -2), (-1, -2), 1, colors.black),
        ('LINEABOVE', (0, -1), (-1, -1), 2, colors.black),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
    ]))
    elements.append(totales_table)
    elements.append(Spacer(1, 8*mm))
    
    # Firmas
    firmas_data = [
        ['Firma y Sello Empresa', 'Firma Cliente']
    ]
    firmas_table = Table(firmas_data, colWidths=[90*mm, 90*mm])
    firmas_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Courier'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('LINEABOVE', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
    ]))
    elements.append(firmas_table)
    
    doc.build(elements)
    buffer.seek(0)
    return buffer.read()


# ==================== FUNCIONARIOS ====================
@api_router.post("/funcionarios", response_model=FuncionarioResponse)
async def crear_funcionario(data: FuncionarioCreate, db: AsyncSession = Depends(get_db)):
    funcionario = Funcionario(**data.model_dump())
    db.add(funcionario)
    await db.commit()
    await db.refresh(funcionario)
    return funcionario

@api_router.get("/funcionarios")
async def listar_funcionarios(empresa_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(Funcionario).where(Funcionario.empresa_id == empresa_id, Funcionario.activo == True)
        .order_by(Funcionario.nombre)
    )
    funcionarios = result.scalars().all()
    
    # Get current month adelantos for each funcionario
    response = []
    
    for funcionario in funcionarios:
        # Calculate total adelantos this month
        adelantos_result = await db.execute(
            select(func_sql.coalesce(func_sql.sum(AdelantoSalario.monto), 0))
            .where(
                AdelantoSalario.funcionario_id == funcionario.id,
                func_sql.extract('year', AdelantoSalario.creado_en) == today_paraguay().year,
                func_sql.extract('month', AdelantoSalario.creado_en) == today_paraguay().month
            )
        )
        total_adelantos = adelantos_result.scalar() or Decimal('0')
        salario_restante = (funcionario.salario_base or Decimal('0')) - total_adelantos
        
        func_dict = FuncionarioResponse.model_validate(funcionario).model_dump()
        func_dict['total_adelantos_mes'] = float(total_adelantos)
        func_dict['salario_restante'] = float(salario_restante)
        response.append(func_dict)
    
    return response

@api_router.get("/funcionarios/{funcionario_id}", response_model=FuncionarioResponse)
async def obtener_funcionario(funcionario_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Funcionario).where(Funcionario.id == funcionario_id))
    funcionario = result.scalar_one_or_none()
    if not funcionario:
        raise HTTPException(status_code=404, detail="Funcionario no encontrado")
    return funcionario

@api_router.put("/funcionarios/{funcionario_id}", response_model=FuncionarioResponse)
async def actualizar_funcionario(funcionario_id: int, data: dict, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Funcionario).where(Funcionario.id == funcionario_id))
    funcionario = result.scalar_one_or_none()
    if not funcionario:
        raise HTTPException(status_code=404, detail="Funcionario no encontrado")
    
    for key, value in data.items():
        if hasattr(funcionario, key) and key != 'id':
            setattr(funcionario, key, value)
    
    await db.commit()
    await db.refresh(funcionario)
    return funcionario

@api_router.delete("/funcionarios/{funcionario_id}")
async def eliminar_funcionario(funcionario_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Funcionario).where(Funcionario.id == funcionario_id))
    funcionario = result.scalar_one_or_none()
    if not funcionario:
        raise HTTPException(status_code=404, detail="Funcionario no encontrado")
    
    funcionario.activo = False
    await db.commit()
    return {"message": "Funcionario desactivado"}

# ==================== ADELANTOS ====================
@api_router.post("/funcionarios/{funcionario_id}/adelantos", response_model=AdelantoSalarioResponse)
async def crear_adelanto(funcionario_id: int, data: dict, db: AsyncSession = Depends(get_db)):
    # Verify funcionario exists
    func_result = await db.execute(select(Funcionario).where(Funcionario.id == funcionario_id))
    funcionario = func_result.scalar_one_or_none()
    if not funcionario:
        raise HTTPException(status_code=404, detail="Funcionario no encontrado")
    
    monto = data.get('monto')
    if not monto:
        raise HTTPException(status_code=400, detail="El monto es requerido")
    
    adelanto = AdelantoSalario(
        funcionario_id=funcionario_id,
        monto=Decimal(str(monto))
    )
    db.add(adelanto)
    await db.commit()
    await db.refresh(adelanto)
    return adelanto

@api_router.get("/funcionarios/{funcionario_id}/adelantos", response_model=List[AdelantoSalarioResponse])
async def listar_adelantos(funcionario_id: int, periodo: Optional[str] = None, db: AsyncSession = Depends(get_db)):
    query = select(AdelantoSalario).where(AdelantoSalario.funcionario_id == funcionario_id)
    
    if periodo:
        year, month = map(int, periodo.split('-'))
        start_date = datetime(year, month, 1, tzinfo=PARAGUAY_TZ)
        if month == 12:
            end_date = datetime(year + 1, 1, 1, tzinfo=PARAGUAY_TZ)
        else:
            end_date = datetime(year, month + 1, 1, tzinfo=PARAGUAY_TZ)
        query = query.where(
            AdelantoSalario.creado_en >= start_date,
            AdelantoSalario.creado_en < end_date
        )
    
    result = await db.execute(query.order_by(AdelantoSalario.creado_en.desc()))
    return result.scalars().all()

# ==================== VEHICULOS ====================
@api_router.post("/vehiculos", response_model=VehiculoResponse)
async def crear_vehiculo(data: VehiculoCreate, db: AsyncSession = Depends(get_db)):
    vehiculo = Vehiculo(**data.model_dump())
    db.add(vehiculo)
    await db.commit()
    await db.refresh(vehiculo)
    return vehiculo

@api_router.get("/vehiculos", response_model=List[VehiculoResponse])
async def listar_vehiculos(empresa_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Vehiculo).where(Vehiculo.empresa_id == empresa_id))
    return result.scalars().all()

@api_router.put("/vehiculos/{vehiculo_id}", response_model=VehiculoResponse)
async def actualizar_vehiculo(vehiculo_id: int, data: dict, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Vehiculo).where(Vehiculo.id == vehiculo_id))
    vehiculo = result.scalar_one_or_none()
    if not vehiculo:
        raise HTTPException(status_code=404, detail="Vehículo no encontrado")
    
    for key, value in data.items():
        if hasattr(vehiculo, key) and key != 'id':
            setattr(vehiculo, key, value)
    
    await db.commit()
    await db.refresh(vehiculo)
    return vehiculo

@api_router.delete("/vehiculos/{vehiculo_id}")
async def eliminar_vehiculo(vehiculo_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Vehiculo).where(Vehiculo.id == vehiculo_id))
    vehiculo = result.scalar_one_or_none()
    if not vehiculo:
        raise HTTPException(status_code=404, detail="Vehículo no encontrado")
    
    await db.delete(vehiculo)
    await db.commit()
    return {"message": "Vehículo eliminado"}

# ==================== ENTREGAS ====================
@api_router.post("/entregas", response_model=EntregaResponse)
async def crear_entrega(data: EntregaCreate, db: AsyncSession = Depends(get_db)):
    entrega = Entrega(**data.model_dump())
    db.add(entrega)
    await db.commit()
    await db.refresh(entrega)
    return entrega

@api_router.get("/entregas", response_model=List[EntregaConDetalles])
async def listar_entregas(
    empresa_id: int,
    fecha_desde: Optional[str] = None,
    fecha_hasta: Optional[str] = None,
    vehiculo_id: Optional[int] = None,
    responsable_id: Optional[int] = None,
    estado: Optional[str] = None,
    db: AsyncSession = Depends(get_db)
):
    query = (
        select(Entrega, Venta, Cliente, Vehiculo, Usuario)
        .join(Venta, Entrega.venta_id == Venta.id)
        .join(Cliente, Venta.cliente_id == Cliente.id)
        .outerjoin(Vehiculo, Entrega.vehiculo_id == Vehiculo.id)
        .outerjoin(Usuario, Entrega.responsable_usuario_id == Usuario.id)
        .where(Venta.empresa_id == empresa_id)
    )
    
    if fecha_desde:
        query = query.where(Entrega.fecha_entrega >= datetime.fromisoformat(fecha_desde))
    if fecha_hasta:
        query = query.where(Entrega.fecha_entrega <= datetime.fromisoformat(fecha_hasta))
    if vehiculo_id:
        query = query.where(Entrega.vehiculo_id == vehiculo_id)
    if responsable_id:
        query = query.where(Entrega.responsable_usuario_id == responsable_id)
    if estado:
        query = query.where(Entrega.estado == EstadoEntrega(estado))
    
    result = await db.execute(query.order_by(Entrega.id.desc()))
    entregas = []
    for row in result.all():
        entrega, venta, cliente, vehiculo, usuario = row
        
        # Load venta items with product/materia names
        items_result = await db.execute(
            select(VentaItem).where(VentaItem.venta_id == venta.id)
        )
        items = []
        for item in items_result.scalars().all():
            item_dict = {
                'cantidad': item.cantidad,
                'precio_unitario': float(item.precio_unitario),
                'total': float(item.total),
                'producto_nombre': None,
                'materia_nombre': None
            }
            
            # Get product name
            if item.producto_id:
                prod_result = await db.execute(select(Producto).where(Producto.id == item.producto_id))
                producto = prod_result.scalar_one_or_none()
                if producto:
                    item_dict['producto_nombre'] = producto.nombre
            
            # Get materia name
            elif item.materia_laboratorio_id:
                mat_result = await db.execute(select(MateriaLaboratorio).where(MateriaLaboratorio.id == item.materia_laboratorio_id))
                materia = mat_result.scalar_one_or_none()
                if materia:
                    item_dict['materia_nombre'] = materia.nombre
            
            items.append(item_dict)
        
        entrega_dict = EntregaResponse.model_validate(entrega).model_dump()
        entrega_dict['cliente_nombre'] = f"{cliente.nombre} {cliente.apellido or ''}"
        entrega_dict['cliente_telefono'] = cliente.telefono
        entrega_dict['cliente_direccion'] = cliente.direccion
        entrega_dict['vehiculo_chapa'] = vehiculo.chapa if vehiculo else None
        entrega_dict['responsable_nombre'] = f"{usuario.nombre} {usuario.apellido or ''}" if usuario else None
        entrega_dict['items'] = items
        entregas.append(EntregaConDetalles(**entrega_dict))
    
    return entregas

@api_router.put("/entregas/{entrega_id}/asignar", response_model=EntregaResponse)
async def asignar_entrega(entrega_id: int, data: AsignarEntrega, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Entrega).where(Entrega.id == entrega_id))
    entrega = result.scalar_one_or_none()
    if not entrega:
        raise HTTPException(status_code=404, detail="Entrega no encontrada")
    
    # Actualizar vehículo y responsable
    entrega.vehiculo_id = data.vehiculo_id
    entrega.responsable_usuario_id = data.responsable_usuario_id
    
    # Si estaba PENDIENTE, cambiar a EN_CAMINO
    if entrega.estado == EstadoEntrega.PENDIENTE:
        entrega.estado = EstadoEntrega.EN_CAMINO
    
    await db.commit()
    await db.refresh(entrega)
    return entrega

@api_router.put("/entregas/{entrega_id}/estado")
async def actualizar_estado_entrega(entrega_id: int, estado: str, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Entrega).where(Entrega.id == entrega_id))
    entrega = result.scalar_one_or_none()
    if not entrega:
        raise HTTPException(status_code=404, detail="Entrega no encontrada")
    
    entrega.estado = EstadoEntrega(estado)
    await db.commit()
    return {"message": "Estado actualizado"}

@api_router.delete("/entregas/{entrega_id}")
async def eliminar_entrega(entrega_id: int, db: AsyncSession = Depends(get_db)):
    """Delete delivery order"""
    result = await db.execute(select(Entrega).where(Entrega.id == entrega_id))
    entrega = result.scalar_one_or_none()
    if not entrega:
        raise HTTPException(status_code=404, detail="Entrega no encontrada")
    
    await db.delete(entrega)
    await db.commit()
    return {"message": "Entrega eliminada"}

# ==================== FACTURAS ====================
@api_router.post("/facturas", response_model=FacturaResponse)
async def crear_factura(data: FacturaCreate, db: AsyncSession = Depends(get_db)):
    factura = Factura(**data.model_dump(), creado_en=now_paraguay())
    db.add(factura)
    await db.commit()
    await db.refresh(factura)
    return factura

@api_router.get("/facturas", response_model=List[FacturaResponse])
async def listar_facturas(empresa_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(Factura)
        .join(Venta, Factura.venta_id == Venta.id)
        .where(Venta.empresa_id == empresa_id)
        .order_by(Factura.creado_en.desc())
    )
    return result.scalars().all()

# ==================== PREFERENCIAS ====================
@api_router.post("/preferencias", response_model=PreferenciaUsuarioResponse)
async def crear_actualizar_preferencias(data: PreferenciaUsuarioCreate, db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(PreferenciaUsuario).where(PreferenciaUsuario.usuario_id == data.usuario_id)
    )
    pref = result.scalar_one_or_none()
    
    if pref:
        pref.tema = data.tema
        pref.color_primario = data.color_primario
    else:
        pref = PreferenciaUsuario(**data.model_dump())
        db.add(pref)
    
    await db.commit()
    await db.refresh(pref)
    return pref

@api_router.get("/preferencias/{usuario_id}", response_model=PreferenciaUsuarioResponse)
async def obtener_preferencias(usuario_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(PreferenciaUsuario).where(PreferenciaUsuario.usuario_id == usuario_id)
    )
    pref = result.scalar_one_or_none()
    if not pref:
        return PreferenciaUsuarioResponse(
            id=0,
            usuario_id=usuario_id,
            tema="light",
            color_primario="#0044CC"
        )
    return pref

# ==================== COTIZACION DIVISAS ====================
@api_router.get("/cotizacion", response_model=CotizacionDivisa)
async def obtener_cotizacion():
    """Get USD/PYG and BRL/PYG exchange rates"""
    global MANUAL_CURRENCY_RATES
    
    # If manual mode is set, return manual rates
    if MANUAL_CURRENCY_RATES['manual'] and MANUAL_CURRENCY_RATES['usd_pyg']:
        return CotizacionDivisa(
            usd_pyg=Decimal(str(MANUAL_CURRENCY_RATES['usd_pyg'])),
            brl_pyg=Decimal(str(MANUAL_CURRENCY_RATES['brl_pyg'])),
            manual=True,
            fecha_actualizacion=MANUAL_CURRENCY_RATES['updated_at'] or now_paraguay()
        )
    
    try:
        async with httpx.AsyncClient() as client:
            response = await client.get(
                "https://api.exchangerate-api.com/v4/latest/USD",
                timeout=10.0
            )
            data = response.json()
            
            usd_pyg = Decimal(str(data['rates'].get('PYG', 7500)))
            brl_rate = Decimal(str(data['rates'].get('BRL', 5.0)))
            brl_pyg = usd_pyg / brl_rate
            
            return CotizacionDivisa(
                usd_pyg=usd_pyg,
                brl_pyg=brl_pyg.quantize(Decimal('0.01')),
                manual=False,
                fecha_actualizacion=now_paraguay()
            )
    except Exception as e:
        logger.error(f"Error fetching exchange rates: {e}")
        # Return default or cached values
        if MANUAL_CURRENCY_RATES['usd_pyg']:
            return CotizacionDivisa(
                usd_pyg=Decimal(str(MANUAL_CURRENCY_RATES['usd_pyg'])),
                brl_pyg=Decimal(str(MANUAL_CURRENCY_RATES['brl_pyg'])),
                manual=True,
                fecha_actualizacion=MANUAL_CURRENCY_RATES['updated_at'] or now_paraguay()
            )
        return CotizacionDivisa(
            usd_pyg=Decimal('7500'),
            brl_pyg=Decimal('1500'),
            manual=True,
            fecha_actualizacion=now_paraguay()
        )

@api_router.post("/cotizacion/manual")
async def establecer_cotizacion_manual(data: dict):
    """Set manual exchange rates"""
    global MANUAL_CURRENCY_RATES
    
    MANUAL_CURRENCY_RATES['usd_pyg'] = data.get('usd_pyg')
    MANUAL_CURRENCY_RATES['brl_pyg'] = data.get('brl_pyg')
    MANUAL_CURRENCY_RATES['manual'] = data.get('manual', True)
    MANUAL_CURRENCY_RATES['updated_at'] = now_paraguay()
    
    return {"message": "Cotización manual establecida", "data": MANUAL_CURRENCY_RATES}

@api_router.post("/cotizacion/auto")
async def activar_cotizacion_automatica():
    """Switch back to automatic exchange rates"""
    global MANUAL_CURRENCY_RATES
    MANUAL_CURRENCY_RATES['manual'] = False
    return {"message": "Cotización automática activada"}

# ==================== DASHBOARD ====================
@api_router.get("/dashboard/stats", response_model=DashboardStats)
async def obtener_estadisticas_dashboard(empresa_id: int, db: AsyncSession = Depends(get_db)):
    today = today_paraguay()
    today_start, today_end = get_day_range_paraguay(today)
    
    ventas_hoy_result = await db.execute(
        select(
            func_sql.coalesce(func_sql.sum(Venta.total), 0),
            func_sql.count(Venta.id)
        )
        .where(
            Venta.empresa_id == empresa_id,
            Venta.estado == EstadoVenta.CONFIRMADA,
            Venta.creado_en >= today_start,
            Venta.creado_en <= today_end
        )
    )
    ventas_row = ventas_hoy_result.first()
    ventas_hoy = ventas_row[0] if ventas_row else Decimal('0')
    cantidad_ventas = ventas_row[1] if ventas_row else 0
    
    # Ventas del mes corriente
    month_start = datetime.combine(today.replace(day=1), datetime.min.time()).replace(tzinfo=PARAGUAY_TZ)
    month_end = datetime.combine(today, datetime.max.time()).replace(tzinfo=PARAGUAY_TZ)
    
    ventas_mes_result = await db.execute(
        select(
            func_sql.coalesce(func_sql.sum(Venta.total), 0),
            func_sql.count(Venta.id)
        )
        .where(
            Venta.empresa_id == empresa_id,
            Venta.estado == EstadoVenta.CONFIRMADA,
            Venta.creado_en >= month_start,
            Venta.creado_en <= month_end
        )
    )
    ventas_mes_row = ventas_mes_result.first()
    ventas_mes = ventas_mes_row[0] if ventas_mes_row else Decimal('0')
    cantidad_ventas_mes = ventas_mes_row[1] if ventas_mes_row else 0
    
    deliverys_result = await db.execute(
        select(func_sql.count(Entrega.id))
        .join(Venta, Entrega.venta_id == Venta.id)
        .where(
            Venta.empresa_id == empresa_id,
            Entrega.estado == EstadoEntrega.PENDIENTE
        )
    )
    deliverys_pendientes = deliverys_result.scalar() or 0
    
    stock_bajo_result = await db.execute(
        select(Producto.id, Producto.nombre, func_sql.sum(StockActual.cantidad), StockActual.alerta_minima)
        .join(StockActual, Producto.id == StockActual.producto_id)
        .where(
            Producto.empresa_id == empresa_id,
            Producto.activo == True,
            StockActual.alerta_minima.isnot(None)
        )
        .group_by(Producto.id, Producto.nombre, StockActual.alerta_minima)
        .having(func_sql.sum(StockActual.cantidad) <= StockActual.alerta_minima)
    )
    productos_bajo_stock = [
        StockBajo(
            producto_id=row[0],
            producto_nombre=row[1],
            stock_total=row[2] or 0,
            alerta_minima=row[3] or 0
        )
        for row in stock_bajo_result.all()
    ]
    
    ventas_por_hora_result = await db.execute(
        select(
            func_sql.extract('hour', Venta.creado_en).label('hora'),
            func_sql.count(Venta.id),
            func_sql.coalesce(func_sql.sum(Venta.total), 0)
        )
        .where(
            Venta.empresa_id == empresa_id,
            Venta.estado == EstadoVenta.CONFIRMADA,
            Venta.creado_en >= today_start,
            Venta.creado_en <= today_end
        )
        .group_by(func_sql.extract('hour', Venta.creado_en))
        .order_by('hora')
    )
    
    # Obtener unidades por hora
    unidades_por_hora_result = await db.execute(
        select(
            func_sql.extract('hour', Venta.creado_en).label('hora'),
            func_sql.coalesce(func_sql.sum(VentaItem.cantidad), 0).label('unidades')
        )
        .join(VentaItem, VentaItem.venta_id == Venta.id)
        .where(
            Venta.empresa_id == empresa_id,
            Venta.estado == EstadoVenta.CONFIRMADA,
            Venta.creado_en >= today_start,
            Venta.creado_en <= today_end
        )
        .group_by(func_sql.extract('hour', Venta.creado_en))
    )
    unidades_dict = {int(row[0]): int(row[1] or 0) for row in unidades_por_hora_result.all()}
    
    ventas_por_hora = [
        VentasPorHora(
            hora=int(row[0]),
            cantidad=row[1],
            monto=row[2],
            unidades=unidades_dict.get(int(row[0]), 0)
        )
        for row in ventas_por_hora_result.all()
    ]
    
    alto_stock_result = await db.execute(
        select(Producto.id, Producto.nombre, func_sql.sum(StockActual.cantidad))
        .join(StockActual, Producto.id == StockActual.producto_id)
        .where(Producto.empresa_id == empresa_id, Producto.activo == True)
        .group_by(Producto.id, Producto.nombre)
        .order_by(func_sql.sum(StockActual.cantidad).desc())
        .limit(5)
    )
    productos_alto_stock = [
        {"producto_id": row[0], "producto_nombre": row[1], "stock_total": row[2] or 0}
        for row in alto_stock_result.all()
    ]
    
    return DashboardStats(
        ventas_hoy=ventas_hoy,
        cantidad_ventas_hoy=cantidad_ventas,
        ventas_mes=ventas_mes,
        cantidad_ventas_mes=cantidad_ventas_mes,
        deliverys_pendientes=deliverys_pendientes,
        productos_stock_bajo=len(productos_bajo_stock),
        creditos_por_vencer=0,
        ventas_por_hora=ventas_por_hora,
        productos_bajo_stock=productos_bajo_stock,
        productos_alto_stock=productos_alto_stock
    )

@api_router.get("/dashboard/ventas-periodo")
async def obtener_ventas_por_periodo(
    empresa_id: int,
    periodo: str = "dia",
    tipo_pago: Optional[str] = None,  # 'credito', 'contado', o None para todos
    db: AsyncSession = Depends(get_db)
):
    """
    Obtener ventas agrupadas por período
    periodo: dia, semana, mes, trimestre, semestre, anio
    tipo_pago: credito (solo CREDITO), contado (EFECTIVO, CHEQUE, TRANSFERENCIA, TARJETA), None (todos)
    """
    now = now_paraguay()
    today = now.date()
    
    # Build base filters
    base_filters = [
        Venta.empresa_id == empresa_id,
        Venta.estado == EstadoVenta.CONFIRMADA
    ]
    
    # Add tipo_pago filter
    if tipo_pago == 'credito':
        base_filters.append(Venta.tipo_pago == TipoPago.CREDITO)
    elif tipo_pago == 'contado':
        base_filters.append(Venta.tipo_pago.in_([
            TipoPago.EFECTIVO,
            TipoPago.CHEQUE,
            TipoPago.TRANSFERENCIA,
            TipoPago.TARJETA
        ]))
    
    if periodo == "dia":
        # Ventas por hora del día actual en zona horaria de Paraguay
        today_start = datetime.combine(today, datetime.min.time()).replace(tzinfo=PARAGUAY_TZ)
        today_end = datetime.combine(today, datetime.max.time()).replace(tzinfo=PARAGUAY_TZ)
        
        base_filters.extend([
            Venta.creado_en >= today_start,
            Venta.creado_en <= today_end
        ])
        
        # Para PostgreSQL: convertir a zona horaria de Paraguay antes de extraer la hora
        # func_sql.timezone('America/Asuncion', ...) convierte a hora local
        hora_paraguay = func_sql.extract(
            'hour', 
            func_sql.timezone(text("'America/Asuncion'"), Venta.creado_en)
        )
        
        result = await db.execute(
            select(
                hora_paraguay.label('hora'),
                func_sql.count(Venta.id).label('cantidad'),
                func_sql.coalesce(func_sql.sum(Venta.total), 0).label('monto'),
                func_sql.coalesce(func_sql.sum(
                    select(func_sql.sum(VentaItem.cantidad))
                    .where(VentaItem.venta_id == Venta.id)
                    .correlate(Venta)
                    .scalar_subquery()
                ), 0).label('unidades')
            )
            .where(and_(*base_filters))
            .group_by(hora_paraguay)
            .order_by('hora')
        )
        return [
            {"label": f"{int(row[0])}:00", "cantidad": row[1], "monto": float(row[2]), "unidades": int(row[3] or 0)}
            for row in result.all()
        ]
    
    elif periodo == "semana":
        # Últimos 7 días
        fecha_inicio = today - timedelta(days=6)
        fecha_inicio_dt = datetime.combine(fecha_inicio, datetime.min.time()).replace(tzinfo=PARAGUAY_TZ)
        
        base_filters.append(Venta.creado_en >= fecha_inicio_dt)
        
        # Primero obtenemos los datos agregados
        result = await db.execute(
            select(
                func_sql.date(Venta.creado_en).label('fecha'),
                func_sql.count(Venta.id).label('cantidad'),
                func_sql.coalesce(func_sql.sum(Venta.total), 0).label('monto')
            )
            .where(and_(*base_filters))
            .group_by(func_sql.date(Venta.creado_en))
            .order_by('fecha')
        )
        ventas_dict = {row[0]: {"cantidad": row[1], "monto": float(row[2])} for row in result.all()}
        
        # Obtener unidades por fecha
        unidades_result = await db.execute(
            select(
                func_sql.date(Venta.creado_en).label('fecha'),
                func_sql.coalesce(func_sql.sum(VentaItem.cantidad), 0).label('unidades')
            )
            .join(VentaItem, VentaItem.venta_id == Venta.id)
            .where(and_(*base_filters))
            .group_by(func_sql.date(Venta.creado_en))
        )
        unidades_dict = {row[0]: int(row[1] or 0) for row in unidades_result.all()}
        
        # Asegurar que todos los días estén representados
        return [
            {
                "label": (fecha_inicio + timedelta(days=i)).strftime("%d/%m"),
                "cantidad": ventas_dict.get(fecha_inicio + timedelta(days=i), {}).get("cantidad", 0),
                "monto": ventas_dict.get(fecha_inicio + timedelta(days=i), {}).get("monto", 0),
                "unidades": unidades_dict.get(fecha_inicio + timedelta(days=i), 0)
            }
            for i in range(7)
        ]
    
    elif periodo == "mes":
        # Último mes (30 días) agrupado por día
        fecha_inicio = today - timedelta(days=29)
        fecha_inicio_dt = datetime.combine(fecha_inicio, datetime.min.time()).replace(tzinfo=PARAGUAY_TZ)
        
        base_filters.append(Venta.creado_en >= fecha_inicio_dt)
        
        result = await db.execute(
            select(
                func_sql.date(Venta.creado_en).label('fecha'),
                func_sql.count(Venta.id).label('cantidad'),
                func_sql.coalesce(func_sql.sum(Venta.total), 0).label('monto')
            )
            .where(and_(*base_filters))
            .group_by(func_sql.date(Venta.creado_en))
            .order_by('fecha')
        )
        ventas_dict = {row[0]: {"cantidad": row[1], "monto": float(row[2])} for row in result.all()}
        
        # Obtener unidades por fecha
        unidades_result = await db.execute(
            select(
                func_sql.date(Venta.creado_en).label('fecha'),
                func_sql.coalesce(func_sql.sum(VentaItem.cantidad), 0).label('unidades')
            )
            .join(VentaItem, VentaItem.venta_id == Venta.id)
            .where(and_(*base_filters))
            .group_by(func_sql.date(Venta.creado_en))
        )
        unidades_dict = {row[0]: int(row[1] or 0) for row in unidades_result.all()}
        
        return [
            {
                "label": (fecha_inicio + timedelta(days=i)).strftime("%d/%m"),
                "cantidad": ventas_dict.get(fecha_inicio + timedelta(days=i), {}).get("cantidad", 0),
                "monto": ventas_dict.get(fecha_inicio + timedelta(days=i), {}).get("monto", 0),
                "unidades": unidades_dict.get(fecha_inicio + timedelta(days=i), 0)
            }
            for i in range(30)
        ]
    
    elif periodo == "trimestre":
        # Últimos 3 meses agrupado por semana
        fecha_inicio = today - timedelta(days=90)
        fecha_inicio_dt = datetime.combine(fecha_inicio, datetime.min.time()).replace(tzinfo=PARAGUAY_TZ)
        
        base_filters.append(Venta.creado_en >= fecha_inicio_dt)
        
        result = await db.execute(
            select(
                func_sql.extract('week', Venta.creado_en).label('semana'),
                func_sql.extract('year', Venta.creado_en).label('anio'),
                func_sql.count(Venta.id).label('cantidad'),
                func_sql.coalesce(func_sql.sum(Venta.total), 0).label('monto')
            )
            .where(and_(*base_filters))
            .group_by('semana', 'anio')
            .order_by('anio', 'semana')
        )
        
        # Obtener unidades por semana
        unidades_result = await db.execute(
            select(
                func_sql.extract('week', Venta.creado_en).label('semana'),
                func_sql.extract('year', Venta.creado_en).label('anio'),
                func_sql.coalesce(func_sql.sum(VentaItem.cantidad), 0).label('unidades')
            )
            .join(VentaItem, VentaItem.venta_id == Venta.id)
            .where(and_(*base_filters))
            .group_by('semana', 'anio')
        )
        unidades_dict = {(int(row[0]), int(row[1])): int(row[2] or 0) for row in unidades_result.all()}
        
        return [
            {
                "label": f"Sem {int(row[0])}",
                "cantidad": row[2],
                "monto": float(row[3]),
                "unidades": unidades_dict.get((int(row[0]), int(row[1])), 0)
            }
            for row in result.all()
        ]
    
    elif periodo == "semestre":
        # Últimos 6 meses agrupado por mes
        fecha_inicio = today - timedelta(days=180)
        fecha_inicio_dt = datetime.combine(fecha_inicio, datetime.min.time()).replace(tzinfo=PARAGUAY_TZ)
        
        base_filters.append(Venta.creado_en >= fecha_inicio_dt)
        
        result = await db.execute(
            select(
                func_sql.extract('month', Venta.creado_en).label('mes'),
                func_sql.extract('year', Venta.creado_en).label('anio'),
                func_sql.count(Venta.id).label('cantidad'),
                func_sql.coalesce(func_sql.sum(Venta.total), 0).label('monto')
            )
            .where(and_(*base_filters))
            .group_by('mes', 'anio')
            .order_by('anio', 'mes')
        )
        
        # Obtener unidades por mes
        unidades_result = await db.execute(
            select(
                func_sql.extract('month', Venta.creado_en).label('mes'),
                func_sql.extract('year', Venta.creado_en).label('anio'),
                func_sql.coalesce(func_sql.sum(VentaItem.cantidad), 0).label('unidades')
            )
            .join(VentaItem, VentaItem.venta_id == Venta.id)
            .where(and_(*base_filters))
            .group_by('mes', 'anio')
        )
        unidades_dict = {(int(row[0]), int(row[1])): int(row[2] or 0) for row in unidades_result.all()}
        
        meses = ["Ene", "Feb", "Mar", "Abr", "May", "Jun", "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]
        return [
            {
                "label": meses[int(row[0]) - 1],
                "cantidad": row[2],
                "monto": float(row[3]),
                "unidades": unidades_dict.get((int(row[0]), int(row[1])), 0)
            }
            for row in result.all()
        ]
    
    elif periodo == "anio":
        # Último año agrupado por mes
        fecha_inicio = today - timedelta(days=365)
        fecha_inicio_dt = datetime.combine(fecha_inicio, datetime.min.time()).replace(tzinfo=PARAGUAY_TZ)
        
        base_filters.append(Venta.creado_en >= fecha_inicio_dt)
        
        result = await db.execute(
            select(
                func_sql.extract('month', Venta.creado_en).label('mes'),
                func_sql.extract('year', Venta.creado_en).label('anio'),
                func_sql.count(Venta.id).label('cantidad'),
                func_sql.coalesce(func_sql.sum(Venta.total), 0).label('monto')
            )
            .where(and_(*base_filters))
            .group_by('mes', 'anio')
            .order_by('anio', 'mes')
        )
        
        # Obtener unidades por mes
        unidades_result = await db.execute(
            select(
                func_sql.extract('month', Venta.creado_en).label('mes'),
                func_sql.extract('year', Venta.creado_en).label('anio'),
                func_sql.coalesce(func_sql.sum(VentaItem.cantidad), 0).label('unidades')
            )
            .join(VentaItem, VentaItem.venta_id == Venta.id)
            .where(and_(*base_filters))
            .group_by('mes', 'anio')
        )
        unidades_dict = {(int(row[0]), int(row[1])): int(row[2] or 0) for row in unidades_result.all()}
        
        meses = ["Ene", "Feb", "Mar", "Abr", "May", "Jun", "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]
        return [
            {
                "label": meses[int(row[0]) - 1],
                "cantidad": row[2],
                "monto": float(row[3]),
                "unidades": unidades_dict.get((int(row[0]), int(row[1])), 0)
            }
            for row in result.all()
        ]
    
    return []

# ==================== ALERTAS ====================
@api_router.get("/alertas", response_model=List[Alerta])
async def obtener_alertas(empresa_id: int, db: AsyncSession = Depends(get_db)):
    alertas = []
    today = today_paraguay()
    
    fecha_limite = today + timedelta(days=30)
    productos_vencimiento = await db.execute(
        select(Producto)
        .where(
            Producto.empresa_id == empresa_id,
            Producto.activo == True,
            Producto.fecha_vencimiento.isnot(None),
            Producto.fecha_vencimiento <= fecha_limite
        )
    )
    for prod in productos_vencimiento.scalars().all():
        dias = (prod.fecha_vencimiento - today).days
        nivel = "danger" if dias <= 7 else "warning"
        alertas.append(Alerta(
            tipo="vencimiento_producto",
            mensaje=f"{prod.nombre} vence en {dias} días",
            nivel=nivel,
            referencia_id=prod.id
        ))
    
    vehiculos_result = await db.execute(
        select(Vehiculo)
        .where(
            Vehiculo.empresa_id == empresa_id,
            or_(
                and_(
                    Vehiculo.vencimiento_habilitacion.isnot(None),
                    Vehiculo.vencimiento_habilitacion <= fecha_limite
                ),
                and_(
                    Vehiculo.vencimiento_cedula_verde.isnot(None),
                    Vehiculo.vencimiento_cedula_verde <= fecha_limite
                )
            )
        )
    )
    for veh in vehiculos_result.scalars().all():
        if veh.vencimiento_habilitacion and veh.vencimiento_habilitacion <= fecha_limite:
            dias = (veh.vencimiento_habilitacion - today).days
            nivel = "danger" if dias <= 7 else "warning"
            alertas.append(Alerta(
                tipo="vencimiento_habilitacion",
                mensaje=f"Habilitación de {veh.chapa} vence en {dias} días",
                nivel=nivel,
                referencia_id=veh.id
            ))
        if veh.vencimiento_cedula_verde and veh.vencimiento_cedula_verde <= fecha_limite:
            dias = (veh.vencimiento_cedula_verde - today).days
            nivel = "danger" if dias <= 7 else "warning"
            alertas.append(Alerta(
                tipo="vencimiento_cedula_verde",
                mensaje=f"Cédula verde de {veh.chapa} vence en {dias} días",
                nivel=nivel,
                referencia_id=veh.id
            ))
    
    stock_bajo = await db.execute(
        select(Producto.id, Producto.nombre, func_sql.sum(StockActual.cantidad), StockActual.alerta_minima)
        .join(StockActual, Producto.id == StockActual.producto_id)
        .where(
            Producto.empresa_id == empresa_id,
            Producto.activo == True,
            StockActual.alerta_minima.isnot(None)
        )
        .group_by(Producto.id, Producto.nombre, StockActual.alerta_minima)
        .having(func_sql.sum(StockActual.cantidad) <= StockActual.alerta_minima)
    )
    for row in stock_bajo.all():
        alertas.append(Alerta(
            tipo="stock_bajo",
            mensaje=f"{row[1]} tiene stock bajo ({row[2]} unidades)",
            nivel="warning",
            referencia_id=row[0]
        ))
    
    return alertas

# ==================== REPORTES PDF ====================
def crear_pdf_reporte(titulo, subtitulo, columnas, datos, totales=None, col_widths=None):
    """Genera un PDF con tabla de datos"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=18, spaceAfter=6, alignment=TA_CENTER)
    elements.append(Paragraph(titulo, title_style))
    
    # Subtitle
    subtitle_style = ParagraphStyle('CustomSubtitle', parent=styles['Normal'], fontSize=10, spaceAfter=20, alignment=TA_CENTER, textColor=colors.grey)
    elements.append(Paragraph(subtitulo, subtitle_style))
    
    # Table
    table_data = [columnas] + datos
    
    if totales:
        table_data.append(totales)
    
    table = Table(table_data, repeatRows=1, colWidths=col_widths)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0044CC')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ALIGN', (-1, 1), (-1, -1), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5F5')]),
    ]))
    
    if totales:
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#E0E0E0')),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ]))
    
    elements.append(table)
    
    # Footer
    elements.append(Spacer(1, 20))
    footer_style = ParagraphStyle('Footer', parent=styles['Normal'], fontSize=8, textColor=colors.grey, alignment=TA_CENTER)
    elements.append(Paragraph(f"Generado el {now_paraguay().strftime('%d/%m/%Y %H:%M')} - Luz Brill ERP", footer_style))
    
    doc.build(elements)
    buffer.seek(0)
    return buffer.getvalue()

def crear_excel_reporte(titulo, subtitulo, columnas, datos, totales=None):
    """Genera un Excel (.xlsx) con tabla de datos estilizada"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte"

    num_cols = len(columnas)
    last_col = get_column_letter(num_cols)

    # Title row (row 1)
    ws.merge_cells(f"A1:{last_col}1")
    title_cell = ws["A1"]
    title_cell.value = titulo
    title_cell.font = Font(name="Calibri", size=16, bold=True, color="0044CC")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Subtitle row (row 2)
    ws.merge_cells(f"A2:{last_col}2")
    sub_cell = ws["A2"]
    sub_cell.value = subtitulo
    sub_cell.font = Font(name="Calibri", size=10, color="808080")
    sub_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    # Empty row 3 (spacer)
    ws.row_dimensions[3].height = 6

    # Headers (row 4)
    HEADER_ROW = 4
    header_fill = PatternFill(start_color="0044CC", end_color="0044CC", fill_type="solid")
    header_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, col_name in enumerate(columnas, 1):
        cell = ws.cell(row=HEADER_ROW, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    ws.row_dimensions[HEADER_ROW].height = 20

    # Data rows
    for row_offset, row_data in enumerate(datos, 1):
        row_num = HEADER_ROW + row_offset
        bg = "FFFFFF" if row_offset % 2 == 1 else "F5F5F5"
        fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=value)
            cell.fill = fill
            cell.font = Font(name="Calibri", size=9)
            cell.border = border
            cell.alignment = Alignment(horizontal="left" if col_idx < num_cols else "right", vertical="center")

    # Totals row
    if totales:
        tot_row = HEADER_ROW + len(datos) + 1
        tot_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        tot_font = Font(name="Calibri", size=9, bold=True)
        for col_idx, value in enumerate(totales, 1):
            cell = ws.cell(row=tot_row, column=col_idx, value=value)
            cell.fill = tot_fill
            cell.font = tot_font
            cell.border = border
            cell.alignment = Alignment(horizontal="right", vertical="center")

    # Auto-size columns
    for col_idx in range(1, num_cols + 1):
        max_len = len(str(columnas[col_idx - 1]))
        for row_data in datos:
            if col_idx <= len(row_data):
                max_len = max(max_len, len(str(row_data[col_idx - 1])))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 45)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()

EXCEL_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

@api_router.get("/reportes/ventas")
async def reporte_ventas(
    empresa_id: int,
    fecha_desde: str,
    fecha_hasta: str,
    tipo_pago: str = None,
    estado: str = None,
    cliente_id: int = None,
    db: AsyncSession = Depends(get_db)
):
    """Genera reporte PDF de ventas por rango de fechas con filtros"""
    try:
        fecha_ini = datetime.fromisoformat(fecha_desde)
        fecha_fin = datetime.fromisoformat(fecha_hasta) + timedelta(days=1)
    except ValueError:
        raise HTTPException(status_code=400, detail="Formato de fecha inválido. Use YYYY-MM-DD")
    
    query = (
        select(Venta, Cliente)
        .join(Cliente, Venta.cliente_id == Cliente.id)
        .where(
            Venta.empresa_id == empresa_id,
            Venta.creado_en >= fecha_ini,
            Venta.creado_en < fecha_fin
        )
    )
    
    # Apply cliente_id filter if provided
    if cliente_id:
        query = query.where(Venta.cliente_id == cliente_id)
    
    # Apply estado filter (default: CONFIRMADA)
    if estado:
        try:
            query = query.where(Venta.estado == EstadoVenta[estado])
        except KeyError:
            raise HTTPException(status_code=400, detail=f"Estado '{estado}' no válido")
    else:
        query = query.where(Venta.estado == EstadoVenta.CONFIRMADA)
    
    # Apply tipo_pago filter - support CONTADO (all non-credit) and CREDITO
    if tipo_pago:
        if tipo_pago == 'CONTADO':
            # CONTADO = all payment types except CREDITO
            query = query.where(Venta.tipo_pago.in_([TipoPago.EFECTIVO, TipoPago.TARJETA, TipoPago.TRANSFERENCIA, TipoPago.CHEQUE]))
        elif tipo_pago == 'CREDITO':
            query = query.where(Venta.tipo_pago == TipoPago.CREDITO)
        else:
            # Support specific payment types if provided
            try:
                query = query.where(Venta.tipo_pago == TipoPago[tipo_pago])
            except KeyError:
                raise HTTPException(status_code=400, detail=f"Tipo de pago '{tipo_pago}' no válido")
    
    query = query.order_by(Venta.creado_en)
    result = await db.execute(query)
    ventas = result.all()
    
    if not ventas:
        raise HTTPException(status_code=404, detail="No se encontraron ventas para los filtros seleccionados")
    
    columnas = ['ID', 'Fecha', 'Cliente', 'Tipo Pago', 'Total', 'Costo', 'Ganancia']
    datos = []
    total_general = Decimal('0')
    costo_general = Decimal('0')
    ganancia_general = Decimal('0')
    
    for venta, cliente in ventas:
        total_general += venta.total
        costo_v = venta.costo_total or Decimal('0')
        ganancia_v = venta.ganancia or Decimal('0')
        costo_general += costo_v
        ganancia_general += ganancia_v
        datos.append([
            str(venta.id),
            venta.creado_en.strftime('%d/%m/%Y'),
            f"{cliente.nombre} {cliente.apellido or ''}".strip()[:25],
            venta.tipo_pago.value if venta.tipo_pago else 'EFECTIVO',
            f"{float(venta.total):,.0f}",
            f"{float(costo_v):,.0f}",
            f"{float(ganancia_v):,.0f}"
        ])
    
    totales = ['', '', '', 'TOTAL:', f"{float(total_general):,.0f}", f"{float(costo_general):,.0f}", f"{float(ganancia_general):,.0f}"]
    
    # Build subtitle with filters
    subtitle = f"Período: {fecha_desde} al {fecha_hasta}"
    if tipo_pago:
        if tipo_pago == 'CONTADO':
            subtitle += " | Tipo: CONTADO"
        elif tipo_pago == 'CREDITO':
            subtitle += " | Tipo: CRÉDITO"
        else:
            subtitle += f" | Pago: {tipo_pago}"
    if estado:
        subtitle += f" | Estado: {estado}"
    subtitle += f" | Total ventas: {len(datos)}"
    
    pdf_bytes = crear_pdf_reporte(
        "Reporte de Ventas",
        subtitle,
        columnas,
        datos,
        totales
    )
    
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=reporte_ventas_{fecha_desde}_{fecha_hasta}.pdf"}
    )

@api_router.get("/reportes/ventas/excel")
async def reporte_ventas_excel(
    empresa_id: int,
    fecha_desde: str,
    fecha_hasta: str,
    tipo_pago: str = None,
    estado: str = None,
    cliente_id: int = None,
    db: AsyncSession = Depends(get_db)
):
    """Genera reporte Excel de ventas por rango de fechas con filtros"""
    try:
        fecha_ini = datetime.fromisoformat(fecha_desde)
        fecha_fin = datetime.fromisoformat(fecha_hasta) + timedelta(days=1)
    except ValueError:
        raise HTTPException(status_code=400, detail="Formato de fecha inválido. Use YYYY-MM-DD")

    query = (
        select(Venta, Cliente)
        .join(Cliente, Venta.cliente_id == Cliente.id)
        .where(
            Venta.empresa_id == empresa_id,
            Venta.creado_en >= fecha_ini,
            Venta.creado_en < fecha_fin
        )
    )
    if cliente_id:
        query = query.where(Venta.cliente_id == cliente_id)
    if estado:
        try:
            query = query.where(Venta.estado == EstadoVenta[estado])
        except KeyError:
            raise HTTPException(status_code=400, detail=f"Estado '{estado}' no válido")
    else:
        query = query.where(Venta.estado == EstadoVenta.CONFIRMADA)
    if tipo_pago:
        if tipo_pago == 'CONTADO':
            query = query.where(Venta.tipo_pago.in_([TipoPago.EFECTIVO, TipoPago.TARJETA, TipoPago.TRANSFERENCIA, TipoPago.CHEQUE]))
        elif tipo_pago == 'CREDITO':
            query = query.where(Venta.tipo_pago == TipoPago.CREDITO)
        else:
            try:
                query = query.where(Venta.tipo_pago == TipoPago[tipo_pago])
            except KeyError:
                raise HTTPException(status_code=400, detail=f"Tipo de pago '{tipo_pago}' no válido")
    query = query.order_by(Venta.creado_en)
    result = await db.execute(query)
    ventas = result.all()

    if not ventas:
        raise HTTPException(status_code=404, detail="No se encontraron ventas para los filtros seleccionados")

    columnas = ['ID', 'Fecha', 'Cliente', 'Tipo Pago', 'Total', 'Costo', 'Ganancia']
    datos = []
    total_general = Decimal('0')
    costo_general = Decimal('0')
    ganancia_general = Decimal('0')
    for venta, cliente in ventas:
        total_general += venta.total
        costo_v = venta.costo_total or Decimal('0')
        ganancia_v = venta.ganancia or Decimal('0')
        costo_general += costo_v
        ganancia_general += ganancia_v
        datos.append([
            str(venta.id),
            venta.creado_en.strftime('%d/%m/%Y'),
            f"{cliente.nombre} {cliente.apellido or ''}".strip(),
            venta.tipo_pago.value if venta.tipo_pago else 'EFECTIVO',
            float(venta.total),
            float(costo_v),
            float(ganancia_v)
        ])
    totales = ['', '', '', 'TOTAL:', float(total_general), float(costo_general), float(ganancia_general)]

    subtitle = f"Período: {fecha_desde} al {fecha_hasta}"
    if tipo_pago:
        subtitle += f" | Tipo: {tipo_pago}"
    if estado:
        subtitle += f" | Estado: {estado}"
    subtitle += f" | Total ventas: {len(datos)}"

    excel_bytes = crear_excel_reporte("Reporte de Ventas", subtitle, columnas, datos, totales)
    return Response(
        content=excel_bytes,
        media_type=EXCEL_MIME,
        headers={"Content-Disposition": f"attachment; filename=reporte_ventas_{fecha_desde}_{fecha_hasta}.xlsx"}
    )

@api_router.get("/reportes/historial-ventas")
async def reporte_historial_ventas(
    empresa_id: int,
    fecha_desde: str = None,
    fecha_hasta: str = None,
    cliente_id: int = None,
    usuario_id: int = None,
    estado: str = None,
    monto_min: float = None,
    monto_max: float = None,
    db: AsyncSession = Depends(get_db)
):
    """Genera reporte PDF del historial de ventas con filtros aplicados"""
    query = (
        select(Venta, Cliente)
        .join(Cliente, Venta.cliente_id == Cliente.id)
        .where(Venta.empresa_id == empresa_id)
        .order_by(Venta.creado_en.desc())
    )
    
    # Apply filters
    if fecha_desde:
        query = query.where(Venta.creado_en >= datetime.fromisoformat(fecha_desde))
    if fecha_hasta:
        query = query.where(Venta.creado_en <= datetime.fromisoformat(fecha_hasta) + timedelta(days=1))
    if cliente_id:
        query = query.where(Venta.cliente_id == cliente_id)
    if usuario_id:
        query = query.where(Venta.usuario_id == usuario_id)
    if estado:
        try:
            query = query.where(Venta.estado == EstadoVenta[estado])
        except KeyError:
            raise HTTPException(status_code=400, detail=f"Estado '{estado}' no válido")
    if monto_min:
        query = query.where(Venta.total >= monto_min)
    if monto_max:
        query = query.where(Venta.total <= monto_max)
    
    result = await db.execute(query)
    ventas = result.all()
    
    if not ventas:
        # Return empty PDF
        subtitle = "Sin registros para los filtros aplicados"
        pdf_bytes = crear_pdf_reporte("Historial de Ventas", subtitle, ['Info'], [['No se encontraron ventas']], None)
        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={"Content-Disposition": "attachment; filename=historial_ventas_vacio.pdf"}
        )
    
    # Build table data
    columnas = ['ID', 'Fecha', 'Cliente', 'Tipo Pago', 'Estado', 'Total', 'Costo', 'Ganancia']
    datos = []
    total_general = Decimal('0')
    costo_general = Decimal('0')
    ganancia_general = Decimal('0')
    cantidad_confirmadas = 0
    
    for venta, cliente in ventas:
        es_confirmada = venta.estado == EstadoVenta.CONFIRMADA
        if es_confirmada:
            total_general += venta.total
            costo_general += venta.costo_total or Decimal('0')
            ganancia_general += venta.ganancia or Decimal('0')
            cantidad_confirmadas += 1
        costo_v = venta.costo_total or Decimal('0')
        ganancia_v = venta.ganancia or Decimal('0')
        estado_text = venta.estado.value if venta.estado else 'N/A'
        datos.append([
            str(venta.id),
            venta.creado_en.strftime('%d/%m/%Y %H:%M'),
            f"{cliente.nombre} {cliente.apellido or ''}".strip()[:20],
            venta.tipo_pago.value if venta.tipo_pago else 'EFECTIVO',
            estado_text,
            f"{float(venta.total):,.0f}",
            f"{float(costo_v) if es_confirmada else 0:,.0f}",
            f"{float(ganancia_v) if es_confirmada else 0:,.0f}"
        ])
    
    # Build subtitle with filters and summary
    subtitle_parts = []
    if fecha_desde:
        subtitle_parts.append(f"Desde: {fecha_desde}")
    if fecha_hasta:
        subtitle_parts.append(f"Hasta: {fecha_hasta}")
    if estado:
        subtitle_parts.append(f"Estado: {estado}")
    
    subtitle = " | ".join(subtitle_parts) if subtitle_parts else "Todas las ventas"
    subtitle += f" | Total registros: {len(ventas)} | Confirmadas: {cantidad_confirmadas}"
    
    totales = ['', '', '', '', 'TOTAL CONFIRMADO:', f"{float(total_general):,.0f}", f"{float(costo_general):,.0f}", f"{float(ganancia_general):,.0f}"]
    
    pdf_bytes = crear_pdf_reporte("Historial de Ventas", subtitle, columnas, datos, totales)
    
    filename = f"historial_ventas_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@api_router.get("/reportes/historial-ventas/excel")
async def reporte_historial_ventas_excel(
    empresa_id: int,
    fecha_desde: str = None,
    fecha_hasta: str = None,
    cliente_id: int = None,
    usuario_id: int = None,
    estado: str = None,
    monto_min: float = None,
    monto_max: float = None,
    db: AsyncSession = Depends(get_db)
):
    """Genera reporte Excel del historial de ventas con filtros aplicados"""
    query = (
        select(Venta, Cliente)
        .join(Cliente, Venta.cliente_id == Cliente.id)
        .where(Venta.empresa_id == empresa_id)
        .order_by(Venta.creado_en.desc())
    )
    if fecha_desde:
        query = query.where(Venta.creado_en >= datetime.fromisoformat(fecha_desde))
    if fecha_hasta:
        query = query.where(Venta.creado_en <= datetime.fromisoformat(fecha_hasta) + timedelta(days=1))
    if cliente_id:
        query = query.where(Venta.cliente_id == cliente_id)
    if usuario_id:
        query = query.where(Venta.usuario_id == usuario_id)
    if estado:
        try:
            query = query.where(Venta.estado == EstadoVenta[estado])
        except KeyError:
            raise HTTPException(status_code=400, detail=f"Estado '{estado}' no válido")
    if monto_min:
        query = query.where(Venta.total >= monto_min)
    if monto_max:
        query = query.where(Venta.total <= monto_max)

    result = await db.execute(query)
    ventas = result.all()

    columnas = ['ID', 'Fecha', 'Cliente', 'Tipo Pago', 'Estado', 'Total', 'Costo', 'Ganancia']
    datos = []
    total_general = Decimal('0')
    costo_general = Decimal('0')
    ganancia_general = Decimal('0')
    cantidad_confirmadas = 0

    for venta, cliente in ventas:
        es_confirmada = venta.estado == EstadoVenta.CONFIRMADA
        if es_confirmada:
            total_general += venta.total
            costo_general += venta.costo_total or Decimal('0')
            ganancia_general += venta.ganancia or Decimal('0')
            cantidad_confirmadas += 1
        costo_v = venta.costo_total or Decimal('0')
        ganancia_v = venta.ganancia or Decimal('0')
        datos.append([
            str(venta.id),
            venta.creado_en.strftime('%d/%m/%Y %H:%M'),
            f"{cliente.nombre} {cliente.apellido or ''}".strip(),
            venta.tipo_pago.value if venta.tipo_pago else 'EFECTIVO',
            venta.estado.value if venta.estado else 'N/A',
            float(venta.total),
            float(costo_v) if es_confirmada else 0,
            float(ganancia_v) if es_confirmada else 0
        ])

    subtitle_parts = []
    if fecha_desde:
        subtitle_parts.append(f"Desde: {fecha_desde}")
    if fecha_hasta:
        subtitle_parts.append(f"Hasta: {fecha_hasta}")
    if estado:
        subtitle_parts.append(f"Estado: {estado}")
    subtitle = " | ".join(subtitle_parts) if subtitle_parts else "Todas las ventas"
    subtitle += f" | Total registros: {len(ventas)} | Confirmadas: {cantidad_confirmadas}"

    totales = ['', '', '', '', 'TOTAL CONF.:', float(total_general), float(costo_general), float(ganancia_general)]
    excel_bytes = crear_excel_reporte("Historial de Ventas", subtitle, columnas, datos, totales)
    filename = f"historial_ventas_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return Response(
        content=excel_bytes,
        media_type=EXCEL_MIME,
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@api_router.get("/reportes/stock")
async def reporte_stock(
    empresa_id: int,
    fecha_desde: str = None,
    fecha_hasta: str = None,
    solo_alertas: str = None,
    db: AsyncSession = Depends(get_db)
):
    """Genera reporte PDF de stock actual con filtros"""
    try:
        # Get products with aggregated stock and the minimum alerta_minima
        result = await db.execute(
            select(
                Producto, 
                func_sql.coalesce(func_sql.sum(StockActual.cantidad), 0).label('stock_total'),
                func_sql.min(StockActual.alerta_minima).label('alerta_minima_val')
            )
            .outerjoin(StockActual, Producto.id == StockActual.producto_id)
            .where(Producto.empresa_id == empresa_id, Producto.activo == True)
            .group_by(Producto.id)
            .order_by(Producto.nombre)
        )
        productos = result.all()
        
        if not productos:
            raise HTTPException(status_code=404, detail="No se encontraron productos activos en el inventario")
        
        # A4 width 595pt - 60pt margins = 535pt available
        # ID=35, Código=70, Producto=190, Stock=45, P.Venta=95, P.Costo=95 → total=530
        col_widths_stock = [35, 70, 190, 45, 95, 95]
        columnas = ['ID', 'Código', 'Producto', 'Stock', 'Precio Venta', 'Precio Costo']
        datos = []
        
        for producto, stock, alerta_minima_val in productos:
            # Use alerta_minima from StockActual if available, else use stock_minimo from Producto, else default to 10
            stock_minimo = alerta_minima_val if alerta_minima_val is not None else (producto.stock_minimo if producto.stock_minimo is not None else 10)
            es_alerta = stock <= stock_minimo
            
            # Filter by solo_alertas if specified
            if solo_alertas == 'true' and not es_alerta:
                continue
            
            estado = " (!)" if es_alerta else ""
            datos.append([
                str(producto.id),
                producto.codigo_barra or '-',
                f"{producto.nombre}{estado}",
                str(int(stock)),
                f"{float(producto.precio_venta):,.0f}",
                f"{float(producto.precio_costo or 0):,.0f}"
            ])
        
        # Build subtitle
        subtitle = f"Fecha: {today_paraguay().strftime('%d/%m/%Y')}"
        if solo_alertas == 'true':
            subtitle += " | Solo stock bajo"
        if not datos:
            subtitle += " | Sin registros"
        else:
            subtitle += f" | Total productos: {len(datos)}"
        
        pdf_bytes = crear_pdf_reporte(
            "Reporte de Stock Actual",
            subtitle,
            columnas,
            datos,
            col_widths=col_widths_stock
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al generar reporte de stock: {str(e)}")
    
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=reporte_stock_{today_paraguay().isoformat()}.pdf"}
    )

@api_router.get("/reportes/stock/excel")
async def reporte_stock_excel(
    empresa_id: int,
    fecha_desde: str = None,
    fecha_hasta: str = None,
    solo_alertas: str = None,
    db: AsyncSession = Depends(get_db)
):
    """Genera reporte Excel de stock actual con filtros"""
    try:
        result = await db.execute(
            select(
                Producto,
                func_sql.coalesce(func_sql.sum(StockActual.cantidad), 0).label('stock_total'),
                func_sql.min(StockActual.alerta_minima).label('alerta_minima_val')
            )
            .outerjoin(StockActual, Producto.id == StockActual.producto_id)
            .where(Producto.empresa_id == empresa_id, Producto.activo == True)
            .group_by(Producto.id)
            .order_by(Producto.nombre)
        )
        productos = result.all()

        if not productos:
            raise HTTPException(status_code=404, detail="No se encontraron productos activos en el inventario")

        columnas = ['ID', 'Código', 'Producto', 'Stock', 'Precio Venta', 'Precio Costo', 'Alerta']
        datos = []
        for producto, stock, alerta_minima_val in productos:
            stock_minimo = alerta_minima_val if alerta_minima_val is not None else (producto.stock_minimo if producto.stock_minimo is not None else 10)
            es_alerta = stock <= stock_minimo
            if solo_alertas == 'true' and not es_alerta:
                continue
            datos.append([
                producto.id,
                producto.codigo_barra or '-',
                producto.nombre,
                int(stock),
                float(producto.precio_venta),
                float(producto.precio_costo or 0),
                'Stock bajo' if es_alerta else 'OK'
            ])

        subtitle = f"Fecha: {today_paraguay().strftime('%d/%m/%Y')}"
        if solo_alertas == 'true':
            subtitle += " | Solo stock bajo"
        subtitle += f" | Total productos: {len(datos)}"

        excel_bytes = crear_excel_reporte("Reporte de Stock Actual", subtitle, columnas, datos)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al generar reporte de stock: {str(e)}")

    return Response(
        content=excel_bytes,
        media_type=EXCEL_MIME,
        headers={"Content-Disposition": f"attachment; filename=reporte_stock_{today_paraguay().isoformat()}.xlsx"}
    )

@api_router.get("/reportes/deudas-proveedores")
async def reporte_deudas_proveedores(
    empresa_id: int,
    fecha_desde: str = None,
    fecha_hasta: str = None,
    estado: str = None,
    db: AsyncSession = Depends(get_db)
):
    """Genera reporte PDF de deudas a proveedores con filtros"""
    try:
        query = (
            select(DeudaProveedor, Proveedor)
            .join(Proveedor, DeudaProveedor.proveedor_id == Proveedor.id)
            .where(Proveedor.empresa_id == empresa_id)
        )
        
        # Filter by estado (default: PENDIENTE if not specified)
        if estado:
            if estado == 'TODOS':
                # No filter, show all
                pass
            elif estado == 'PENDIENTE':
                query = query.where(DeudaProveedor.pagado == False)
            elif estado == 'PAGADO':
                query = query.where(DeudaProveedor.pagado == True)
            else:
                raise HTTPException(status_code=400, detail=f"Estado '{estado}' no válido. Use PENDIENTE, PAGADO o TODOS")
        else:
            # Default to PENDIENTE if estado is None
            query = query.where(DeudaProveedor.pagado == False)
        
        # Filter by fecha_emision range
        if fecha_desde:
            try:
                fecha_ini = datetime.fromisoformat(fecha_desde)
                query = query.where(DeudaProveedor.fecha_emision >= fecha_ini.date())
            except ValueError:
                raise HTTPException(status_code=400, detail="Formato de fecha_desde inválido. Use YYYY-MM-DD")
        if fecha_hasta:
            try:
                fecha_fin = datetime.fromisoformat(fecha_hasta)
                query = query.where(DeudaProveedor.fecha_emision <= fecha_fin.date())
            except ValueError:
                raise HTTPException(status_code=400, detail="Formato de fecha_hasta inválido. Use YYYY-MM-DD")
        
        query = query.order_by(DeudaProveedor.fecha_limite)
        result = await db.execute(query)
        deudas = result.all()
        
        # If no data, return empty report instead of error
        if not deudas:
            # Create empty report
            columnas = ['Proveedor', 'Descripción', 'Monto', 'Emisión', 'Vencimiento']
            datos = []
            totales = ['', '', '0', '', '']
            
            subtitle = f"Fecha: {today_paraguay().strftime('%d/%m/%Y')}"
            if fecha_desde and fecha_hasta:
                subtitle += f" | Período: {fecha_desde} al {fecha_hasta}"
            if estado:
                subtitle += f" | Estado: {estado}"
            subtitle += " | Sin registros"
            
            pdf_bytes = crear_pdf_reporte(
                "Reporte de Deudas a Proveedores",
                subtitle,
                columnas,
                datos,
                totales
            )
            return Response(
                content=pdf_bytes,
                media_type="application/pdf",
                headers={"Content-Disposition": f"attachment; filename=reporte_deudas_proveedores_{today_paraguay().isoformat()}.pdf"}
            )
        
        columnas = ['Proveedor', 'Descripción', 'Monto', 'Emisión', 'Vencimiento']
        datos = []
        total_deuda = Decimal('0')
        
        for deuda, proveedor in deudas:
            total_deuda += deuda.monto
            vencido = "⚠️" if deuda.fecha_limite and deuda.fecha_limite < today_paraguay() else ""
            datos.append([
                proveedor.nombre[:25],
                (deuda.descripcion or '-')[:30],
                f"{float(deuda.monto):,.0f}",
                deuda.fecha_emision.strftime('%d/%m/%Y') if deuda.fecha_emision else '-',
                f"{deuda.fecha_limite.strftime('%d/%m/%Y') if deuda.fecha_limite else '-'}{vencido}"
            ])
        
        totales = ['', '', f"{float(total_deuda):,.0f}", '', '']
        
        # Build subtitle
        subtitle = f"Fecha: {today_paraguay().strftime('%d/%m/%Y')}"
        if fecha_desde and fecha_hasta:
            subtitle += f" | Período: {fecha_desde} al {fecha_hasta}"
        if estado:
            subtitle += f" | Estado: {estado}"
        subtitle += f" | Total deudas: {len(datos)}"
        
        pdf_bytes = crear_pdf_reporte(
            "Reporte de Deudas a Proveedores",
            subtitle,
            columnas,
            datos,
            totales
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al generar reporte de deudas: {str(e)}")
    
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=reporte_deudas_proveedores_{today_paraguay().isoformat()}.pdf"}
    )

@api_router.get("/reportes/deudas-proveedores/excel")
async def reporte_deudas_proveedores_excel(
    empresa_id: int,
    fecha_desde: str = None,
    fecha_hasta: str = None,
    estado: str = None,
    db: AsyncSession = Depends(get_db)
):
    """Genera reporte Excel de deudas a proveedores con filtros"""
    try:
        query = (
            select(DeudaProveedor, Proveedor)
            .join(Proveedor, DeudaProveedor.proveedor_id == Proveedor.id)
            .where(Proveedor.empresa_id == empresa_id)
        )
        if estado:
            if estado == 'TODOS':
                pass
            elif estado == 'PENDIENTE':
                query = query.where(DeudaProveedor.pagado == False)
            elif estado == 'PAGADO':
                query = query.where(DeudaProveedor.pagado == True)
            else:
                raise HTTPException(status_code=400, detail=f"Estado '{estado}' no válido")
        else:
            query = query.where(DeudaProveedor.pagado == False)
        if fecha_desde:
            try:
                query = query.where(DeudaProveedor.fecha_emision >= datetime.fromisoformat(fecha_desde).date())
            except ValueError:
                raise HTTPException(status_code=400, detail="Formato de fecha_desde inválido. Use YYYY-MM-DD")
        if fecha_hasta:
            try:
                query = query.where(DeudaProveedor.fecha_emision <= datetime.fromisoformat(fecha_hasta).date())
            except ValueError:
                raise HTTPException(status_code=400, detail="Formato de fecha_hasta inválido. Use YYYY-MM-DD")
        query = query.order_by(DeudaProveedor.fecha_limite)
        result = await db.execute(query)
        deudas = result.all()

        columnas = ['Proveedor', 'Descripción', 'Monto', 'Emisión', 'Vencimiento', 'Estado']
        datos = []
        total_deuda = Decimal('0')
        for deuda, proveedor in deudas:
            total_deuda += deuda.monto
            vencido = deuda.fecha_limite and deuda.fecha_limite < today_paraguay()
            datos.append([
                proveedor.nombre,
                deuda.descripcion or '-',
                float(deuda.monto),
                deuda.fecha_emision.strftime('%d/%m/%Y') if deuda.fecha_emision else '-',
                deuda.fecha_limite.strftime('%d/%m/%Y') if deuda.fecha_limite else '-',
                'VENCIDO' if vencido else ('PAGADO' if deuda.pagado else 'PENDIENTE')
            ])

        totales = ['', '', float(total_deuda), '', '', '']
        subtitle = f"Fecha: {today_paraguay().strftime('%d/%m/%Y')}"
        if fecha_desde and fecha_hasta:
            subtitle += f" | Período: {fecha_desde} al {fecha_hasta}"
        if estado:
            subtitle += f" | Estado: {estado}"
        subtitle += f" | Total deudas: {len(datos)}"

        excel_bytes = crear_excel_reporte("Reporte de Deudas a Proveedores", subtitle, columnas, datos, totales)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al generar reporte de deudas: {str(e)}")

    return Response(
        content=excel_bytes,
        media_type=EXCEL_MIME,
        headers={"Content-Disposition": f"attachment; filename=reporte_deudas_proveedores_{today_paraguay().isoformat()}.xlsx"}
    )

@api_router.get("/reportes/creditos-clientes")
async def reporte_creditos_clientes(
    empresa_id: int,
    fecha_desde: str = None,
    fecha_hasta: str = None,
    estado: str = None,
    db: AsyncSession = Depends(get_db)
):
    """Genera reporte PDF de créditos de clientes con filtros"""
    try:
        query = (
            select(CreditoCliente, Cliente)
            .join(Cliente, CreditoCliente.cliente_id == Cliente.id)
            .where(Cliente.empresa_id == empresa_id)
        )
        
        # Filter by estado (default: PENDIENTE if not specified)
        if estado:
            if estado == 'TODOS':
                # No filter, show all
                pass
            elif estado == 'PENDIENTE':
                query = query.where(CreditoCliente.pagado == False)
            elif estado == 'PAGADO':
                query = query.where(CreditoCliente.pagado == True)
            else:
                raise HTTPException(status_code=400, detail=f"Estado '{estado}' no válido. Use PENDIENTE, PAGADO o TODOS")
        else:
            # Default to PENDIENTE if estado is None
            query = query.where(CreditoCliente.pagado == False)
        
        # Filter by fecha_venta range
        if fecha_desde:
            try:
                fecha_ini = datetime.fromisoformat(fecha_desde)
                query = query.where(CreditoCliente.fecha_venta >= fecha_ini.date())
            except ValueError:
                raise HTTPException(status_code=400, detail="Formato de fecha_desde inválido. Use YYYY-MM-DD")
        if fecha_hasta:
            try:
                fecha_fin = datetime.fromisoformat(fecha_hasta)
                query = query.where(CreditoCliente.fecha_venta <= fecha_fin.date())
            except ValueError:
                raise HTTPException(status_code=400, detail="Formato de fecha_hasta inválido. Use YYYY-MM-DD")
        
        query = query.order_by(CreditoCliente.fecha_venta)
        result = await db.execute(query)
        creditos = result.all()
        
        # If no data, return empty report instead of error
        if not creditos:
            # Create empty report
            columnas = ['Cliente', 'Venta #', 'Original', 'Pagado', 'Pendiente', 'Fecha']
            datos = []
            totales = ['', 'TOTAL:', '0', '0', '0', '']
            
            subtitle = f"Fecha: {today_paraguay().strftime('%d/%m/%Y')}"
            if fecha_desde and fecha_hasta:
                subtitle += f" | Período: {fecha_desde} al {fecha_hasta}"
            if estado:
                subtitle += f" | Estado: {estado}"
            subtitle += " | Sin registros"
            
            pdf_bytes = crear_pdf_reporte(
                "Reporte de Créditos de Clientes",
                subtitle,
                columnas,
                datos,
                totales
            )
            return Response(
                content=pdf_bytes,
                media_type="application/pdf",
                headers={"Content-Disposition": f"attachment; filename=reporte_creditos_clientes_{today_paraguay().isoformat()}.pdf"}
            )
        
        columnas = ['Cliente', 'Venta #', 'Original', 'Pagado', 'Pendiente', 'Fecha']
        datos = []
        total_pendiente = Decimal('0')
        total_pagado = Decimal('0')
        total_original = Decimal('0')
        
        for credito, cliente in creditos:
            monto_pagado = credito.monto_original - credito.monto_pendiente
            total_pendiente += credito.monto_pendiente
            total_pagado += monto_pagado
            total_original += credito.monto_original
            datos.append([
                f"{cliente.nombre} {cliente.apellido or ''}".strip()[:25],
                str(credito.venta_id or '-'),
                f"{float(credito.monto_original):,.0f}",
                f"{float(monto_pagado):,.0f}",
                f"{float(credito.monto_pendiente):,.0f}",
                credito.fecha_venta.strftime('%d/%m/%Y') if credito.fecha_venta else '-'
            ])
        
        totales = ['', 'TOTAL:', f"{float(total_original):,.0f}", f"{float(total_pagado):,.0f}", f"{float(total_pendiente):,.0f}", '']
        
        # Build subtitle
        subtitle = f"Fecha: {today_paraguay().strftime('%d/%m/%Y')}"
        if fecha_desde and fecha_hasta:
            subtitle += f" | Período: {fecha_desde} al {fecha_hasta}"
        if estado:
            subtitle += f" | Estado: {estado}"
        subtitle += f" | Total créditos: {len(datos)}"
        
        pdf_bytes = crear_pdf_reporte(
            "Reporte de Créditos de Clientes",
            subtitle,
            columnas,
            datos,
            totales
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al generar reporte de créditos: {str(e)}")
    
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=reporte_creditos_clientes_{today_paraguay().isoformat()}.pdf"}
    )

@api_router.get("/reportes/creditos-clientes/excel")
async def reporte_creditos_clientes_excel(
    empresa_id: int,
    fecha_desde: str = None,
    fecha_hasta: str = None,
    estado: str = None,
    db: AsyncSession = Depends(get_db)
):
    """Genera reporte Excel de créditos de clientes con filtros"""
    try:
        query = (
            select(CreditoCliente, Cliente)
            .join(Cliente, CreditoCliente.cliente_id == Cliente.id)
            .where(Cliente.empresa_id == empresa_id)
        )
        if estado:
            if estado == 'TODOS':
                pass
            elif estado == 'PENDIENTE':
                query = query.where(CreditoCliente.pagado == False)
            elif estado == 'PAGADO':
                query = query.where(CreditoCliente.pagado == True)
            else:
                raise HTTPException(status_code=400, detail=f"Estado '{estado}' no válido")
        else:
            query = query.where(CreditoCliente.pagado == False)
        if fecha_desde:
            try:
                query = query.where(CreditoCliente.fecha_venta >= datetime.fromisoformat(fecha_desde).date())
            except ValueError:
                raise HTTPException(status_code=400, detail="Formato de fecha_desde inválido. Use YYYY-MM-DD")
        if fecha_hasta:
            try:
                query = query.where(CreditoCliente.fecha_venta <= datetime.fromisoformat(fecha_hasta).date())
            except ValueError:
                raise HTTPException(status_code=400, detail="Formato de fecha_hasta inválido. Use YYYY-MM-DD")
        query = query.order_by(CreditoCliente.fecha_venta)
        result = await db.execute(query)
        creditos = result.all()

        columnas = ['Cliente', 'Venta #', 'Monto Original', 'Pagado', 'Pendiente', 'Fecha']
        datos = []
        total_pendiente = Decimal('0')
        total_pagado = Decimal('0')
        total_original = Decimal('0')
        for credito, cliente in creditos:
            monto_pagado = credito.monto_original - credito.monto_pendiente
            total_pendiente += credito.monto_pendiente
            total_pagado += monto_pagado
            total_original += credito.monto_original
            datos.append([
                f"{cliente.nombre} {cliente.apellido or ''}".strip(),
                str(credito.venta_id or '-'),
                float(credito.monto_original),
                float(monto_pagado),
                float(credito.monto_pendiente),
                credito.fecha_venta.strftime('%d/%m/%Y') if credito.fecha_venta else '-'
            ])

        totales = ['', 'TOTAL:', float(total_original), float(total_pagado), float(total_pendiente), '']
        subtitle = f"Fecha: {today_paraguay().strftime('%d/%m/%Y')}"
        if fecha_desde and fecha_hasta:
            subtitle += f" | Período: {fecha_desde} al {fecha_hasta}"
        if estado:
            subtitle += f" | Estado: {estado}"
        subtitle += f" | Total créditos: {len(datos)}"

        excel_bytes = crear_excel_reporte("Reporte de Créditos de Clientes", subtitle, columnas, datos, totales)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al generar reporte de créditos: {str(e)}")

    return Response(
        content=excel_bytes,
        media_type=EXCEL_MIME,
        headers={"Content-Disposition": f"attachment; filename=reporte_creditos_clientes_{today_paraguay().isoformat()}.xlsx"}
    )

# ==================== CICLOS DE SALARIO ====================
@api_router.post("/ciclos-salario/generar")
async def generar_ciclo_salario(empresa_id: int, mes: str, db: AsyncSession = Depends(get_db)):
    """Genera ciclos de salario para todos los funcionarios activos del mes especificado (formato: YYYY-MM)"""
    # Check if cycles already exist for this month
    existing = await db.execute(
        select(CicloSalario).where(CicloSalario.periodo == mes)
    )
    if existing.scalar_one_or_none():
        raise HTTPException(status_code=400, detail=f"Ya existen ciclos de salario para {mes}")
    
    # Get all active funcionarios
    func_result = await db.execute(
        select(Funcionario).where(Funcionario.empresa_id == empresa_id, Funcionario.activo == True)
    )
    funcionarios = func_result.scalars().all()
    
    if not funcionarios:
        raise HTTPException(status_code=404, detail="No hay funcionarios activos")
    
    # Parse month
    año, mes_num = mes.split('-')
    fecha_inicio = date(int(año), int(mes_num), 1)
    if int(mes_num) == 12:
        fecha_fin = date(int(año) + 1, 1, 1) - timedelta(days=1)
    else:
        fecha_fin = date(int(año), int(mes_num) + 1, 1) - timedelta(days=1)
    
    ciclos_creados = []
    for funcionario in funcionarios:
        # Calculate adelantos for this funcionario in previous month
        # (Adelantos del mes anterior se descuentan del salario actual)
        mes_anterior = fecha_inicio - timedelta(days=1)
        periodo_anterior = mes_anterior.strftime('%Y-%m')
        
        adelantos_result = await db.execute(
            select(func_sql.coalesce(func_sql.sum(AdelantoSalario.monto), 0))
            .where(
                AdelantoSalario.funcionario_id == funcionario.id,
                func_sql.extract('year', AdelantoSalario.creado_en) == mes_anterior.year,
                func_sql.extract('month', AdelantoSalario.creado_en) == mes_anterior.month
            )
        )
        total_adelantos = adelantos_result.scalar() or Decimal('0')
        
        salario_neto = (funcionario.salario_base or Decimal('0')) - total_adelantos
        
        ciclo = CicloSalario(
            funcionario_id=funcionario.id,
            periodo=mes,
            fecha_inicio=fecha_inicio,
            fecha_fin=fecha_fin,
            salario_base=funcionario.salario_base or Decimal('0'),
            descuentos=total_adelantos,
            salario_neto=salario_neto,
            pagado=False
        )
        db.add(ciclo)
        ciclos_creados.append({
            "funcionario": f"{funcionario.nombre} {funcionario.apellido or ''}",
            "salario_base": float(funcionario.salario_base or 0),
            "adelantos": float(total_adelantos),
            "salario_neto": float(salario_neto)
        })
    
    await db.commit()
    
    return {
        "message": f"Ciclos de salario generados para {mes}",
        "ciclos_creados": len(ciclos_creados),
        "detalle": ciclos_creados
    }

@api_router.get("/ciclos-salario")
async def listar_ciclos_salario(empresa_id: int, periodo: Optional[str] = None, db: AsyncSession = Depends(get_db)):
    """Lista los ciclos de salario"""
    query = (
        select(CicloSalario, Funcionario)
        .join(Funcionario, CicloSalario.funcionario_id == Funcionario.id)
        .where(Funcionario.empresa_id == empresa_id)
        .order_by(CicloSalario.periodo.desc(), Funcionario.nombre)
    )
    
    if periodo:
        query = query.where(CicloSalario.periodo == periodo)
    
    result = await db.execute(query)
    ciclos = result.all()
    
    return [
        {
            "id": ciclo.id,
            "funcionario_id": ciclo.funcionario_id,
            "funcionario_nombre": f"{func.nombre} {func.apellido or ''}".strip(),
            "periodo": ciclo.periodo,
            "fecha_inicio": ciclo.fecha_inicio.isoformat() if ciclo.fecha_inicio else None,
            "fecha_fin": ciclo.fecha_fin.isoformat() if ciclo.fecha_fin else None,
            "salario_base": float(ciclo.salario_base),
            "descuentos": float(ciclo.descuentos),
            "salario_neto": float(ciclo.salario_neto),
            "pagado": ciclo.pagado,
            "fecha_pago": ciclo.fecha_pago.isoformat() if ciclo.fecha_pago else None
        }
        for ciclo, func in ciclos
    ]

@api_router.post("/ciclos-salario/{ciclo_id}/pagar")
async def pagar_ciclo_salario(ciclo_id: int, db: AsyncSession = Depends(get_db)):
    """Marca un ciclo de salario como pagado"""
    result = await db.execute(select(CicloSalario).where(CicloSalario.id == ciclo_id))
    ciclo = result.scalar_one_or_none()
    if not ciclo:
        raise HTTPException(status_code=404, detail="Ciclo no encontrado")
    
    if ciclo.pagado:
        raise HTTPException(status_code=400, detail="Este ciclo ya fue pagado")
    
    ciclo.pagado = True
    ciclo.fecha_pago = now_paraguay()
    
    await db.commit()
    return {"message": "Salario marcado como pagado", "ciclo_id": ciclo_id}

@api_router.get("/ciclos-salario/alertas")
async def alertas_salarios_pendientes(empresa_id: int, db: AsyncSession = Depends(get_db)):
    """Obtiene alertas de salarios pendientes de pago"""
    # Get current period
    hoy = today_paraguay()
    periodo_actual = hoy.strftime('%Y-%m')
    
    # Get unpaid cycles
    result = await db.execute(
        select(CicloSalario, Funcionario)
        .join(Funcionario, CicloSalario.funcionario_id == Funcionario.id)
        .where(
            Funcionario.empresa_id == empresa_id,
            CicloSalario.pagado == False
        )
        .order_by(CicloSalario.periodo)
    )
    ciclos_pendientes = result.all()
    
    alertas = []
    for ciclo, func in ciclos_pendientes:
        alertas.append({
            "tipo": "salario_pendiente",
            "mensaje": f"Salario pendiente: {func.nombre} {func.apellido or ''} - {ciclo.periodo}",
            "monto": float(ciclo.salario_neto),
            "funcionario_id": func.id,
            "ciclo_id": ciclo.id,
            "periodo": ciclo.periodo,
            "urgente": ciclo.periodo < periodo_actual
        })
    
    return alertas

# ==================== GASTOS OPERATIVOS ====================

@api_router.get("/gastos")
async def list_gastos(
    empresa_id: int,
    fecha_desde: Optional[str] = None,
    fecha_hasta: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
    _: dict = Depends(get_current_usuario),
):
    PARAGUAY_TZ = ZoneInfo("America/Asuncion")
    q = select(Gasto).where(Gasto.empresa_id == empresa_id)
    if fecha_desde:
        q = q.where(Gasto.fecha >= datetime.fromisoformat(fecha_desde).date())
    if fecha_hasta:
        q = q.where(Gasto.fecha <= datetime.fromisoformat(fecha_hasta).date())
    q = q.order_by(Gasto.fecha.desc(), Gasto.id.desc())
    result = await db.execute(q)
    return result.scalars().all()


@api_router.post("/gastos", response_model=GastoResponse, status_code=201)
async def create_gasto(
    gasto: GastoCreate,
    db: AsyncSession = Depends(get_db),
    _: dict = Depends(get_current_usuario),
):
    db_gasto = Gasto(**gasto.model_dump())
    db.add(db_gasto)
    await db.commit()
    await db.refresh(db_gasto)
    return db_gasto


@api_router.put("/gastos/{gasto_id}", response_model=GastoResponse)
async def update_gasto(
    gasto_id: int,
    gasto: GastoUpdate,
    db: AsyncSession = Depends(get_db),
    _: dict = Depends(get_current_usuario),
):
    result = await db.execute(select(Gasto).where(Gasto.id == gasto_id))
    db_gasto = result.scalar_one_or_none()
    if not db_gasto:
        raise HTTPException(status_code=404, detail="Gasto no encontrado")
    for field, value in gasto.model_dump().items():
        setattr(db_gasto, field, value)
    await db.commit()
    await db.refresh(db_gasto)
    return db_gasto


@api_router.delete("/gastos/{gasto_id}", status_code=204)
async def delete_gasto(
    gasto_id: int,
    db: AsyncSession = Depends(get_db),
    _: dict = Depends(get_current_usuario),
):
    result = await db.execute(select(Gasto).where(Gasto.id == gasto_id))
    db_gasto = result.scalar_one_or_none()
    if not db_gasto:
        raise HTTPException(status_code=404, detail="Gasto no encontrado")
    await db.delete(db_gasto)
    await db.commit()


# ==================== BUSINESS INTELLIGENCE ====================

@api_router.get("/bi/estadisticas-productos")
async def bi_estadisticas_productos(
    empresa_id: int,
    periodo: str = "semanal",  # semanal | mensual
    periodos: int = 8,
    db: AsyncSession = Depends(get_db)
):
    """Estadísticas de ventas por producto con promedio y comparativa de período."""
    PARAGUAY_TZ = ZoneInfo("America/Asuncion")
    now = datetime.now(PARAGUAY_TZ)
    if periodo == "semanal":
        delta = timedelta(weeks=periodos)
    else:
        delta = timedelta(days=periodos * 30)
    fecha_inicio = now - delta

    # Overall stats per product
    stats_q = (
        select(
            VentaItem.producto_id,
            Producto.nombre.label("producto_nombre"),
            func_sql.sum(VentaItem.cantidad).label("total_unidades"),
            func_sql.sum(VentaItem.total).label("total_ventas"),
            func_sql.sum(VentaItem.precio_costo * VentaItem.cantidad).label("total_costo"),
            func_sql.count(func_sql.distinct(Venta.id)).label("num_ventas"),
        )
        .join(Venta, VentaItem.venta_id == Venta.id)
        .join(Producto, VentaItem.producto_id == Producto.id)
        .where(
            Venta.empresa_id == empresa_id,
            Venta.estado == EstadoVenta.CONFIRMADA,
            Venta.creado_en >= fecha_inicio,
            VentaItem.producto_id.isnot(None),
        )
        .group_by(VentaItem.producto_id, Producto.nombre)
        .order_by(func_sql.sum(VentaItem.total).desc())
        .limit(200)
    )

    # Current period (latest period)
    if periodo == "semanal":
        current_inicio = now - timedelta(weeks=1)
        prev_inicio = now - timedelta(weeks=2)
        prev_fin = current_inicio
    else:
        import calendar as cal_mod
        first_this = datetime(now.year, now.month, 1, tzinfo=PARAGUAY_TZ)
        m = now.month - 1 if now.month > 1 else 12
        y = now.year if now.month > 1 else now.year - 1
        first_prev = datetime(y, m, 1, tzinfo=PARAGUAY_TZ)
        current_inicio = first_this
        prev_inicio = first_prev
        prev_fin = first_this

    cur_q = (
        select(VentaItem.producto_id,
               func_sql.sum(VentaItem.cantidad).label("unidades"),
               func_sql.sum(VentaItem.total).label("ventas"))
        .join(Venta, VentaItem.venta_id == Venta.id)
        .where(Venta.empresa_id == empresa_id, Venta.estado == EstadoVenta.CONFIRMADA,
               Venta.creado_en >= current_inicio, VentaItem.producto_id.isnot(None))
        .group_by(VentaItem.producto_id)
    )
    prev_q = (
        select(VentaItem.producto_id,
               func_sql.sum(VentaItem.cantidad).label("unidades"),
               func_sql.sum(VentaItem.total).label("ventas"))
        .join(Venta, VentaItem.venta_id == Venta.id)
        .where(Venta.empresa_id == empresa_id, Venta.estado == EstadoVenta.CONFIRMADA,
               Venta.creado_en >= prev_inicio, Venta.creado_en < prev_fin,
               VentaItem.producto_id.isnot(None))
        .group_by(VentaItem.producto_id)
    )
    stock_q = (
        select(StockActual.producto_id, func_sql.sum(StockActual.cantidad).label("stock_total"))
        .join(Almacen, StockActual.almacen_id == Almacen.id)
        .where(Almacen.empresa_id == empresa_id)
        .group_by(StockActual.producto_id)
    )

    stats_r, cur_r, prev_r, stock_r = await db.execute(stats_q), await db.execute(cur_q), await db.execute(prev_q), await db.execute(stock_q)
    cur_map = {r.producto_id: r for r in cur_r.all()}
    prev_map = {r.producto_id: r for r in prev_r.all()}
    stock_map = {r.producto_id: int(r.stock_total) for r in stock_r.all()}

    productos = []
    for row in stats_r.all():
        pid = row.producto_id
        cu = cur_map.get(pid)
        pv = prev_map.get(pid)
        cu_uni = int(cu.unidades) if cu else 0
        pv_uni = int(pv.unidades) if pv else 0
        if pv_uni > 0:
            variacion = round(((cu_uni - pv_uni) / pv_uni) * 100, 1)
        elif cu_uni > 0:
            variacion = 100.0
        else:
            variacion = 0.0
        total_u = int(row.total_unidades)
        total_v = float(row.total_ventas or 0)
        total_c = float(row.total_costo or 0)
        productos.append({
            "producto_id": pid,
            "producto_nombre": row.producto_nombre,
            "total_unidades": total_u,
            "total_ventas": total_v,
            "total_costo": total_c,
            "ganancia_bruta": round(total_v - total_c, 2),
            "num_ventas": int(row.num_ventas),
            "promedio_por_periodo": round(total_u / periodos, 1),
            "periodo_actual_unidades": cu_uni,
            "periodo_anterior_unidades": pv_uni,
            "variacion_pct": variacion,
            "stock_actual": stock_map.get(pid, 0),
        })

    # Build trend bars (one per period)
    tendencia = []
    for i in range(periodos - 1, -1, -1):
        if periodo == "semanal":
            t_ini = now - timedelta(weeks=i + 1)
            t_fin = now - timedelta(weeks=i)
            label = f"Sem -{i}" if i > 0 else "Esta sem."
        else:
            import calendar as cal_mod
            m_off = now.month - i - 1
            yr = now.year + (m_off // 12 if m_off >= 0 else (m_off - 11) // 12)
            mo = m_off % 12
            if mo <= 0:
                mo += 12
                yr -= 1
            days_m = cal_mod.monthrange(yr, mo)[1]
            t_ini = datetime(yr, mo, 1, tzinfo=PARAGUAY_TZ)
            t_fin = datetime(yr, mo, days_m, 23, 59, 59, tzinfo=PARAGUAY_TZ)
            label = f"{yr}-{mo:02d}" if i > 0 else f"{now.year}-{now.month:02d}"
        t_q = (
            select(func_sql.sum(VentaItem.total).label("ventas"),
                   func_sql.sum(VentaItem.cantidad).label("unidades"))
            .join(Venta, VentaItem.venta_id == Venta.id)
            .where(Venta.empresa_id == empresa_id, Venta.estado == EstadoVenta.CONFIRMADA,
                   Venta.creado_en >= t_ini, Venta.creado_en < t_fin,
                   VentaItem.producto_id.isnot(None))
        )
        t_r = await db.execute(t_q)
        t_row = t_r.one()
        tendencia.append({"label": label, "ventas": float(t_row.ventas or 0), "unidades": int(t_row.unidades or 0)})

    return {"periodo": periodo, "periodos": periodos, "productos": productos, "tendencia": tendencia}


@api_router.get("/bi/cierre")
async def bi_cierre(
    empresa_id: int,
    fecha_desde: str,
    fecha_hasta: str,
    db: AsyncSession = Depends(get_db)
):
    """Balance de cierre: ventas/ganancias vs compras de mercadería vs salarios."""
    PARAGUAY_TZ = ZoneInfo("America/Asuncion")
    try:
        f_ini = datetime.fromisoformat(fecha_desde).replace(tzinfo=PARAGUAY_TZ)
        f_fin = datetime.fromisoformat(fecha_hasta).replace(hour=23, minute=59, second=59, tzinfo=PARAGUAY_TZ)
    except ValueError:
        raise HTTPException(status_code=400, detail="Formato de fecha inválido (YYYY-MM-DD)")

    # Ventas confirmadas
    venta_q = (
        select(func_sql.sum(Venta.total).label("total_ventas"),
               func_sql.sum(Venta.ganancia).label("total_ganancia"),
               func_sql.sum(Venta.costo_total).label("total_costo_ventas"),
               func_sql.sum(Venta.iva).label("total_iva"),
               func_sql.count(Venta.id).label("cantidad_ventas"))
        .where(Venta.empresa_id == empresa_id, Venta.estado == EstadoVenta.CONFIRMADA,
               Venta.creado_en >= f_ini, Venta.creado_en <= f_fin)
    )

    # Compras de mercadería (entradas de stock con costo)
    compras_q = (
        select(func_sql.sum(MovimientoStock.costo_unitario * MovimientoStock.cantidad).label("total_compras"),
               func_sql.count(MovimientoStock.id).label("num_entradas"))
        .join(Almacen, MovimientoStock.almacen_id == Almacen.id)
        .where(Almacen.empresa_id == empresa_id,
               MovimientoStock.tipo == TipoMovimientoStock.ENTRADA,
               MovimientoStock.costo_unitario.isnot(None),
               MovimientoStock.creado_en >= f_ini, MovimientoStock.creado_en <= f_fin)
    )

    # Salarios pagados (ciclos)
    salarios_q = (
        select(func_sql.sum(CicloSalario.salario_neto).label("total_salarios"),
               func_sql.count(CicloSalario.id).label("num_ciclos"))
        .join(Funcionario, CicloSalario.funcionario_id == Funcionario.id)
        .where(Funcionario.empresa_id == empresa_id, CicloSalario.pagado == True,
               CicloSalario.fecha_pago >= f_ini, CicloSalario.fecha_pago <= f_fin)
    )

    # Deudas proveedores pendientes (registradas en el período)
    deudas_q = (
        select(func_sql.sum(DeudaProveedor.monto).label("total_deudas"),
               func_sql.count(DeudaProveedor.id).label("num_deudas"))
        .join(Proveedor, DeudaProveedor.proveedor_id == Proveedor.id)
        .where(Proveedor.empresa_id == empresa_id, DeudaProveedor.pagado == False)
    )

    # Monthly breakdown of sales — Python-side grouping for cross-DB compatibility
    ventas_mes_q = (
        select(Venta.creado_en, Venta.total, Venta.ganancia, Venta.costo_total)
        .where(Venta.empresa_id == empresa_id, Venta.estado == EstadoVenta.CONFIRMADA,
               Venta.creado_en >= f_ini, Venta.creado_en <= f_fin)
    )

    # Gastos operativos en el período
    gastos_q = (
        select(func_sql.sum(Gasto.monto).label("total_gastos"),
               func_sql.count(Gasto.id).label("num_gastos"))
        .where(Gasto.empresa_id == empresa_id,
               Gasto.fecha >= f_ini.date(), Gasto.fecha <= f_fin.date())
    )

    # Valor excedente de stock (FIFO): suma de tandas con unidades restantes × costo de compra
    # No filtrado por fecha — representa el estado actual del inventario como snapshot
    stock_valor_q = (
        select(
            func_sql.sum(
                MovimientoStock.cantidad_restante * MovimientoStock.costo_unitario
            ).label("valor_total"),
            func_sql.count(MovimientoStock.id).label("num_tandas"),
            func_sql.count(func_sql.distinct(MovimientoStock.producto_id)).label("num_productos"),
        )
        .join(Almacen, MovimientoStock.almacen_id == Almacen.id)
        .where(
            Almacen.empresa_id == empresa_id,
            MovimientoStock.tipo == TipoMovimientoStock.ENTRADA,
            MovimientoStock.cantidad_restante > 0,
            MovimientoStock.costo_unitario.isnot(None),
        )
    )

    vr = await db.execute(venta_q)
    cr = await db.execute(compras_q)
    sr = await db.execute(salarios_q)
    dr = await db.execute(deudas_q)
    vmr = await db.execute(ventas_mes_q)
    gr = await db.execute(gastos_q)
    svr = await db.execute(stock_valor_q)

    vrow = vr.one()
    crow = cr.one()
    srow = sr.one()
    drow = dr.one()
    grow = gr.one()
    svrow = svr.one()

    total_ventas = float(vrow.total_ventas or 0)
    total_ganancia = float(vrow.total_ganancia or 0)
    total_costo_ventas = float(vrow.total_costo_ventas or 0)
    total_iva = float(vrow.total_iva or 0)
    total_compras = float(crow.total_compras or 0)
    total_salarios = float(srow.total_salarios or 0)
    total_deudas_pendientes = float(drow.total_deudas or 0)
    total_gastos_op = float(grow.total_gastos or 0)
    valor_stock = float(svrow.valor_total or 0)
    num_tandas = int(svrow.num_tandas or 0)
    num_productos_stock = int(svrow.num_productos or 0)

    total_egresos = total_compras + total_salarios + total_gastos_op
    balance_neto = total_ventas - total_egresos

    # Group ventas by month in Python
    from collections import defaultdict
    mes_map = defaultdict(lambda: {"ventas": 0.0, "ganancia": 0.0, "costo": 0.0})
    for row in vmr.all():
        key = row.creado_en.strftime('%Y-%m') if row.creado_en else None
        if key:
            mes_map[key]["ventas"] += float(row.total or 0)
            mes_map[key]["ganancia"] += float(row.ganancia or 0)
            mes_map[key]["costo"] += float(row.costo_total or 0)
    meses = [{"mes": k, "ventas": v["ventas"], "ganancia": v["ganancia"], "costo": v["costo"]}
             for k, v in sorted(mes_map.items())]

    return {
        "periodo": {"desde": fecha_desde, "hasta": fecha_hasta},
        "ventas": {
            "total": total_ventas,
            "ganancia": total_ganancia,
            "costo": total_costo_ventas,
            "iva": total_iva,
            "cantidad": int(vrow.cantidad_ventas or 0),
        },
        "compras_mercaderia": {
            "total": total_compras,
            "num_entradas": int(crow.num_entradas or 0),
        },
        "salarios": {
            "total_pagado": total_salarios,
            "num_ciclos": int(srow.num_ciclos or 0),
        },
        "gastos_operativos": {
            "total": total_gastos_op,
            "cantidad": int(grow.num_gastos or 0),
        },
        "deudas_pendientes": {
            "total": total_deudas_pendientes,
            "cantidad": int(drow.num_deudas or 0),
        },
        "valor_stock_actual": {
            "total": valor_stock,
            "num_tandas": num_tandas,
            "num_productos": num_productos_stock,
        },
        "resumen": {
            "total_ingresos": total_ventas,
            "total_egresos": total_egresos,
            "balance_neto": balance_neto,
            "margen_bruto_pct": round((total_ganancia / total_ventas * 100) if total_ventas > 0 else 0, 1),
        },
        "evolucion_mensual": meses,
    }


@api_router.get("/bi/sugerencias-compra")
async def bi_sugerencias_compra(
    empresa_id: int,
    semanas_analisis: int = 4,
    semanas_cobertura: int = 4,
    db: AsyncSession = Depends(get_db)
):
    """Sugerencias de compra basadas en velocidad de venta histórica."""
    PARAGUAY_TZ = ZoneInfo("America/Asuncion")
    now = datetime.now(PARAGUAY_TZ)
    fecha_inicio = now - timedelta(weeks=semanas_analisis)

    # Sales velocity per product
    vel_q = (
        select(VentaItem.producto_id, Producto.nombre.label("producto_nombre"),
               func_sql.sum(VentaItem.cantidad).label("total_unidades"),
               func_sql.count(func_sql.distinct(Venta.id)).label("num_ventas"))
        .join(Venta, VentaItem.venta_id == Venta.id)
        .join(Producto, VentaItem.producto_id == Producto.id)
        .where(Venta.empresa_id == empresa_id, Venta.estado == EstadoVenta.CONFIRMADA,
               Venta.creado_en >= fecha_inicio, VentaItem.producto_id.isnot(None))
        .group_by(VentaItem.producto_id, Producto.nombre)
        .order_by(func_sql.sum(VentaItem.cantidad).desc())
    )

    # Current stock
    stock_q = (
        select(StockActual.producto_id, func_sql.sum(StockActual.cantidad).label("stock_total"),
               Producto.nombre.label("producto_nombre"), Producto.precio_costo.label("precio_costo"))
        .join(Almacen, StockActual.almacen_id == Almacen.id)
        .join(Producto, StockActual.producto_id == Producto.id)
        .where(Almacen.empresa_id == empresa_id)
        .group_by(StockActual.producto_id, Producto.nombre, Producto.precio_costo)
    )

    vel_r = await db.execute(vel_q)
    stock_r = await db.execute(stock_q)

    stock_map = {r.producto_id: {"stock": int(r.stock_total), "nombre": r.producto_nombre,
                                  "precio_costo": float(r.precio_costo or 0)}
                 for r in stock_r.all()}

    sugerencias = []
    for row in vel_r.all():
        pid = row.producto_id
        total_u = int(row.total_unidades)
        vel_semanal = round(total_u / semanas_analisis, 1)
        stock_recomendado = round(vel_semanal * semanas_cobertura)
        stock_actual = stock_map.get(pid, {}).get("stock", 0)
        compra_sugerida = max(0, stock_recomendado - stock_actual)
        precio_costo = stock_map.get(pid, {}).get("precio_costo", 0)
        if compra_sugerida > 0:
            sugerencias.append({
                "producto_id": pid,
                "producto_nombre": row.producto_nombre,
                "velocidad_semanal": vel_semanal,
                "stock_actual": stock_actual,
                "stock_recomendado": stock_recomendado,
                "compra_sugerida": int(compra_sugerida),
                "costo_estimado": round(compra_sugerida * precio_costo, 2),
                "cobertura_actual_semanas": round(stock_actual / vel_semanal, 1) if vel_semanal > 0 else None,
                "urgencia": "ALTA" if stock_actual <= 0 else ("MEDIA" if stock_actual < vel_semanal else "BAJA"),
            })

    sugerencias.sort(key=lambda x: (x["urgencia"] != "ALTA", x["urgencia"] != "MEDIA", -x["velocidad_semanal"]))
    costo_total = sum(s["costo_estimado"] for s in sugerencias)

    return {
        "parametros": {"semanas_analisis": semanas_analisis, "semanas_cobertura": semanas_cobertura},
        "sugerencias": sugerencias,
        "resumen": {
            "productos_con_necesidad": len(sugerencias),
            "costo_total_estimado": round(costo_total, 2),
            "urgentes": sum(1 for s in sugerencias if s["urgencia"] == "ALTA"),
            "media_urgencia": sum(1 for s in sugerencias if s["urgencia"] == "MEDIA"),
        }
    }


@api_router.get("/bi/exceso-compras")
async def bi_exceso_compras(
    empresa_id: int,
    semanas_referencia: int = 4,
    db: AsyncSession = Depends(get_db)
):
    """Detección de compras en exceso: productos con stock alto respecto a su velocidad de venta."""
    PARAGUAY_TZ = ZoneInfo("America/Asuncion")
    now = datetime.now(PARAGUAY_TZ)
    fecha_inicio = now - timedelta(weeks=semanas_referencia)

    # Sales velocity últimas N semanas
    vel_q = (
        select(VentaItem.producto_id,
               func_sql.sum(VentaItem.cantidad).label("total_unidades"))
        .join(Venta, VentaItem.venta_id == Venta.id)
        .where(Venta.empresa_id == empresa_id, Venta.estado == EstadoVenta.CONFIRMADA,
               Venta.creado_en >= fecha_inicio, VentaItem.producto_id.isnot(None))
        .group_by(VentaItem.producto_id)
    )

    # Entradas de stock recientes
    entradas_q = (
        select(MovimientoStock.producto_id,
               func_sql.sum(MovimientoStock.cantidad).label("total_comprado"),
               func_sql.sum(MovimientoStock.costo_unitario * MovimientoStock.cantidad).label("total_costo"))
        .join(Almacen, MovimientoStock.almacen_id == Almacen.id)
        .where(Almacen.empresa_id == empresa_id,
               MovimientoStock.tipo == TipoMovimientoStock.ENTRADA,
               MovimientoStock.creado_en >= fecha_inicio)
        .group_by(MovimientoStock.producto_id)
    )

    # Stock actual + product info
    stock_q = (
        select(StockActual.producto_id, func_sql.sum(StockActual.cantidad).label("stock_total"),
               Producto.nombre.label("nombre"), Producto.precio_venta.label("precio_venta"),
               Producto.precio_costo.label("precio_costo"))
        .join(Almacen, StockActual.almacen_id == Almacen.id)
        .join(Producto, StockActual.producto_id == Producto.id)
        .where(Almacen.empresa_id == empresa_id)
        .group_by(StockActual.producto_id, Producto.nombre, Producto.precio_venta, Producto.precio_costo)
    )

    vel_r = await db.execute(vel_q)
    ent_r = await db.execute(entradas_q)
    stk_r = await db.execute(stock_q)

    vel_map = {r.producto_id: int(r.total_unidades) for r in vel_r.all()}
    ent_map = {r.producto_id: {"comprado": int(r.total_comprado),
                                "costo": float(r.total_costo or 0)} for r in ent_r.all()}

    excesos = []
    sin_movimiento = []

    for row in stk_r.all():
        pid = row.producto_id
        stock = int(row.stock_total)
        vendido = vel_map.get(pid, 0)
        vel_semanal = round(vendido / semanas_referencia, 1)
        entrada = ent_map.get(pid, {})
        comprado = entrada.get("comprado", 0)
        costo_sobrante = float(row.precio_costo or 0)

        if stock <= 0:
            continue

        # Sin ventas recientes pero con stock
        if vendido == 0 and stock > 0:
            sin_movimiento.append({
                "producto_id": pid,
                "producto_nombre": row.nombre,
                "stock_actual": stock,
                "vendido_periodo": 0,
                "velocidad_semanal": 0,
                "dias_para_agotar": None,
                "valor_inmovilizado": round(stock * costo_sobrante, 2),
                "tipo": "SIN_VENTAS",
            })
        # Stock alto comparado con velocidad de venta
        elif vel_semanal > 0:
            semanas_para_agotar = stock / vel_semanal
            if semanas_para_agotar > semanas_referencia * 2:
                excesos.append({
                    "producto_id": pid,
                    "producto_nombre": row.nombre,
                    "stock_actual": stock,
                    "vendido_periodo": vendido,
                    "velocidad_semanal": vel_semanal,
                    "semanas_para_agotar": round(semanas_para_agotar, 1),
                    "comprado_periodo": comprado,
                    "exceso_vs_venta": max(0, comprado - vendido),
                    "valor_inmovilizado": round(stock * costo_sobrante, 2),
                    "tipo": "EXCESO_STOCK",
                })

    excesos.sort(key=lambda x: -x["valor_inmovilizado"])
    sin_movimiento.sort(key=lambda x: -x["valor_inmovilizado"])

    return {
        "parametros": {"semanas_referencia": semanas_referencia},
        "exceso_stock": excesos,
        "sin_ventas_recientes": sin_movimiento,
        "resumen": {
            "productos_exceso": len(excesos),
            "productos_sin_ventas": len(sin_movimiento),
            "valor_total_inmovilizado": round(
                sum(e["valor_inmovilizado"] for e in excesos) +
                sum(s["valor_inmovilizado"] for s in sin_movimiento), 2
            ),
        }
    }


@api_router.get("/bi/producto-evolucion")
async def bi_producto_evolucion(
    empresa_id: int,
    producto_id: int,
    periodo: str = "semanal",
    periodos: int = 12,
    db: AsyncSession = Depends(get_db)
):
    """Retorna la evolución por período (semana/mes) de un producto específico."""
    PARAGUAY_TZ = ZoneInfo("America/Asuncion")
    now = datetime.now(PARAGUAY_TZ)

    prod_q = select(Producto.nombre).where(Producto.id == producto_id)
    prod_r = await db.execute(prod_q)
    prod_nombre = prod_r.scalar_one_or_none()
    if not prod_nombre:
        raise HTTPException(status_code=404, detail="Producto no encontrado")

    tendencia = []
    for i in range(periodos - 1, -1, -1):
        if periodo == "semanal":
            t_ini = now - timedelta(weeks=i + 1)
            t_fin = now - timedelta(weeks=i)
            label = f"Sem -{i}" if i > 0 else "Esta sem."
        else:
            import calendar as cal_mod
            m_off = now.month - i - 1
            yr = now.year + (m_off // 12 if m_off >= 0 else (m_off - 11) // 12)
            mo = m_off % 12
            if mo <= 0:
                mo += 12
                yr -= 1
            days_m = cal_mod.monthrange(yr, mo)[1]
            t_ini = datetime(yr, mo, 1, tzinfo=PARAGUAY_TZ)
            t_fin = datetime(yr, mo, days_m, 23, 59, 59, tzinfo=PARAGUAY_TZ)
            label = f"{yr}-{mo:02d}" if i > 0 else f"{now.year}-{now.month:02d}"
        t_q = (
            select(
                func_sql.sum(VentaItem.total).label("ventas"),
                func_sql.sum(VentaItem.cantidad).label("unidades"),
                func_sql.sum(
                    VentaItem.total - (VentaItem.precio_costo * VentaItem.cantidad)
                ).label("ganancia"),
            )
            .join(Venta, VentaItem.venta_id == Venta.id)
            .where(
                Venta.empresa_id == empresa_id,
                Venta.estado == EstadoVenta.CONFIRMADA,
                VentaItem.producto_id == producto_id,
                Venta.creado_en >= t_ini,
                Venta.creado_en < t_fin,
            )
        )
        t_r = await db.execute(t_q)
        t_row = t_r.one()
        tendencia.append({
            "label": label,
            "ventas": float(t_row.ventas or 0),
            "unidades": int(t_row.unidades or 0),
            "ganancia": float(t_row.ganancia or 0),
        })

    return {
        "producto_id": producto_id,
        "producto_nombre": prod_nombre,
        "periodo": periodo,
        "periodos": periodos,
        "tendencia": tendencia,
    }


@api_router.get("/bi/estadisticas-productos/excel")
async def bi_estadisticas_productos_excel(
    empresa_id: int,
    periodo: str = "semanal",
    periodos: int = 8,
    metrica: str = "ventas",  # ventas | unidades | ganancia
    db: AsyncSession = Depends(get_db)
):
    """Exporta ranking completo de productos B.I. a Excel con gráfico embebido."""
    PARAGUAY_TZ = ZoneInfo("America/Asuncion")
    now = datetime.now(PARAGUAY_TZ)
    if periodo == "semanal":
        delta = timedelta(weeks=periodos)
    else:
        delta = timedelta(days=periodos * 30)
    fecha_inicio = now - delta

    if periodo == "semanal":
        current_inicio = now - timedelta(weeks=1)
        prev_inicio = now - timedelta(weeks=2)
        prev_fin = current_inicio
    else:
        import calendar as cal_mod
        first_this = datetime(now.year, now.month, 1, tzinfo=PARAGUAY_TZ)
        m = now.month - 1 if now.month > 1 else 12
        y = now.year if now.month > 1 else now.year - 1
        first_prev = datetime(y, m, 1, tzinfo=PARAGUAY_TZ)
        current_inicio = first_this
        prev_inicio = first_prev
        prev_fin = first_this

    stats_q = (
        select(
            VentaItem.producto_id,
            Producto.nombre.label("producto_nombre"),
            func_sql.sum(VentaItem.cantidad).label("total_unidades"),
            func_sql.sum(VentaItem.total).label("total_ventas"),
            func_sql.sum(VentaItem.precio_costo * VentaItem.cantidad).label("total_costo"),
            func_sql.count(func_sql.distinct(Venta.id)).label("num_ventas"),
        )
        .join(Venta, VentaItem.venta_id == Venta.id)
        .join(Producto, VentaItem.producto_id == Producto.id)
        .where(
            Venta.empresa_id == empresa_id,
            Venta.estado == EstadoVenta.CONFIRMADA,
            Venta.creado_en >= fecha_inicio,
            VentaItem.producto_id.isnot(None),
        )
        .group_by(VentaItem.producto_id, Producto.nombre)
        .order_by(func_sql.sum(VentaItem.total).desc())
    )
    cur_q = (
        select(VentaItem.producto_id,
               func_sql.sum(VentaItem.cantidad).label("unidades"),
               func_sql.sum(VentaItem.total).label("ventas"))
        .join(Venta, VentaItem.venta_id == Venta.id)
        .where(Venta.empresa_id == empresa_id, Venta.estado == EstadoVenta.CONFIRMADA,
               Venta.creado_en >= current_inicio, VentaItem.producto_id.isnot(None))
        .group_by(VentaItem.producto_id)
    )
    prev_q = (
        select(VentaItem.producto_id,
               func_sql.sum(VentaItem.cantidad).label("unidades"),
               func_sql.sum(VentaItem.total).label("ventas"))
        .join(Venta, VentaItem.venta_id == Venta.id)
        .where(Venta.empresa_id == empresa_id, Venta.estado == EstadoVenta.CONFIRMADA,
               Venta.creado_en >= prev_inicio, Venta.creado_en < prev_fin,
               VentaItem.producto_id.isnot(None))
        .group_by(VentaItem.producto_id)
    )
    stock_q = (
        select(StockActual.producto_id, func_sql.sum(StockActual.cantidad).label("stock_total"))
        .join(Almacen, StockActual.almacen_id == Almacen.id)
        .where(Almacen.empresa_id == empresa_id)
        .group_by(StockActual.producto_id)
    )

    stats_r = await db.execute(stats_q)
    cur_r = await db.execute(cur_q)
    prev_r = await db.execute(prev_q)
    stock_r = await db.execute(stock_q)
    cur_map = {r.producto_id: r for r in cur_r.all()}
    prev_map = {r.producto_id: r for r in prev_r.all()}
    stock_map = {r.producto_id: int(r.stock_total) for r in stock_r.all()}

    productos = []
    for row in stats_r.all():
        pid = row.producto_id
        cu = cur_map.get(pid)
        pv = prev_map.get(pid)
        cu_uni = int(cu.unidades) if cu else 0
        pv_uni = int(pv.unidades) if pv else 0
        if pv_uni > 0:
            variacion = round(((cu_uni - pv_uni) / pv_uni) * 100, 1)
        elif cu_uni > 0:
            variacion = 100.0
        else:
            variacion = 0.0
        total_u = int(row.total_unidades)
        total_v = float(row.total_ventas or 0)
        total_c = float(row.total_costo or 0)
        ganancia = round(total_v - total_c, 0)
        prom = round(total_u / periodos, 1)
        stock = stock_map.get(pid, 0)
        productos.append({
            "nombre": row.producto_nombre,
            "ventas": total_v,
            "ganancia": ganancia,
            "unidades": total_u,
            "prom": prom,
            "stock": stock,
            "variacion": variacion,
        })

    # Sort by chosen metric
    sort_key = {"ventas": "ventas", "ganancia": "ganancia", "unidades": "unidades"}.get(metrica, "ventas")
    productos.sort(key=lambda x: -x[sort_key])

    metrica_label = {"ventas": "Ventas (₲)", "ganancia": "Ganancia (₲)", "unidades": "Unidades"}[metrica]
    periodo_label = "semana" if periodo == "semanal" else "mes"
    titulo = f"B.I. — Ranking de Productos por {metrica_label}"
    subtitulo = f"Período: últimos {periodos} {periodo_label}s | Ordenado por: {metrica_label}"

    # Build workbook manually (with chart)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ranking Productos"

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Title
    ws.merge_cells("A1:H1")
    t = ws["A1"]
    t.value = titulo
    t.font = Font(name="Calibri", size=14, bold=True, color="0044CC")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    # Subtitle
    ws.merge_cells("A2:H2")
    s = ws["A2"]
    s.value = subtitulo
    s.font = Font(name="Calibri", size=9, color="808080")
    s.alignment = Alignment(horizontal="center")

    # Headers row 4
    headers = ["#", "Producto", "Ventas (₲)", "Ganancia (₲)", "Unidades",
               f"Prom. uds/{periodo_label}", "Stock actual", "Variación %"]
    hfill = PatternFill(start_color="0044CC", end_color="0044CC", fill_type="solid")
    hfont = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=ci, value=h)
        cell.fill = hfill
        cell.font = hfont
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    ws.row_dimensions[4].height = 20

    # Data rows
    for ri, p in enumerate(productos, 1):
        rn = 4 + ri
        bg = "FFFFFF" if ri % 2 == 1 else "F5F7FF"
        fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
        row_vals = [ri, p["nombre"], p["ventas"], p["ganancia"], p["unidades"],
                    p["prom"], p["stock"], p["variacion"]]
        for ci, val in enumerate(row_vals, 1):
            cell = ws.cell(row=rn, column=ci, value=val)
            cell.fill = fill
            cell.font = Font(name="Calibri", size=9)
            cell.border = border
            cell.alignment = Alignment(horizontal="right" if ci > 1 else "center", vertical="center")
        
    # Column widths
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 35
    for col in ["C", "D", "E", "F", "G", "H"]:
        ws.column_dimensions[col].width = 16

    # Totals row
    tot_rn = 4 + len(productos) + 1
    tot_fill = PatternFill(start_color="E8EDF5", end_color="E8EDF5", fill_type="solid")
    tot_font = Font(name="Calibri", size=9, bold=True)
    totals = [
        "", "TOTAL",
        sum(p["ventas"] for p in productos),
        sum(p["ganancia"] for p in productos),
        sum(p["unidades"] for p in productos),
        "", "", ""
    ]
    for ci, val in enumerate(totals, 1):
        cell = ws.cell(row=tot_rn, column=ci, value=val)
        cell.fill = tot_fill
        cell.font = tot_font
        cell.border = border
        cell.alignment = Alignment(horizontal="right", vertical="center")

    # Embedded BarChart (top 20)
    chart_data_count = min(20, len(productos))
    chart_col = {"ventas": 3, "ganancia": 4, "unidades": 5}[metrica]
    chart = BarChart()
    chart.type = "bar"  # horizontal
    chart.title = f"Top {chart_data_count} — {metrica_label}"
    chart.y_axis.title = "Producto"
    chart.x_axis.title = metrica_label
    chart.style = 10
    chart.width = 22
    chart.height = 14

    data_ref = Reference(ws, min_col=chart_col, min_row=4, max_row=4 + chart_data_count)
    cats_ref = Reference(ws, min_col=2, min_row=5, max_row=4 + chart_data_count)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.series[0].graphicalProperties.solidFill = "0044CC"

    # Place chart to the right of data (column J, row 4)
    ws.add_chart(chart, "J4")

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    today_str = now.strftime("%Y%m%d")
    return Response(
        content=buffer.getvalue(),
        media_type=EXCEL_MIME,
        headers={"Content-Disposition": f"attachment; filename=bi_productos_{metrica}_{today_str}.xlsx"}
    )


# ==================== SYNC PERMISSIONS ====================
@api_router.post("/sistema/sync-permisos")
async def sync_permisos(db: AsyncSession = Depends(get_db)):
    """Add any missing permissions and assign them to all ADMIN roles. Safe to run on production."""
    all_permisos = [
        ("gastos.ver", "Ver gastos operativos"),
        ("gastos.gestionar", "Gestionar gastos operativos"),
    ]

    added_perms = []
    for clave, desc in all_permisos:
        existing = await db.execute(select(Permiso).where(Permiso.clave == clave))
        if not existing.scalar_one_or_none():
            perm = Permiso(clave=clave, descripcion=desc)
            db.add(perm)
            await db.flush()
            added_perms.append(clave)

    # Assign new permissions to all ADMIN roles
    assigned = 0
    for clave in added_perms:
        perm_result = await db.execute(select(Permiso).where(Permiso.clave == clave))
        perm = perm_result.scalar_one_or_none()
        if not perm:
            continue
        admin_roles = await db.execute(select(Rol).where(Rol.nombre == "ADMIN"))
        for rol in admin_roles.scalars().all():
            existing_rp = await db.execute(
                select(RolPermiso).where(RolPermiso.rol_id == rol.id, RolPermiso.permiso_id == perm.id)
            )
            if not existing_rp.scalar_one_or_none():
                db.add(RolPermiso(rol_id=rol.id, permiso_id=perm.id))
                assigned += 1

    await db.commit()
    return {
        "permisos_added": added_perms,
        "rol_permisos_assigned": assigned,
        "message": "Sincronización de permisos completada" if added_perms else "No hay permisos nuevos para agregar"
    }


# ==================== SEED DATA ====================
@api_router.post("/seed")
async def seed_data(db: AsyncSession = Depends(get_db)):
    """Create initial data for testing"""
    existing = await db.execute(select(Empresa))
    if existing.scalar_one_or_none():
        return {"message": "Data already seeded"}
    
    empresa = Empresa(
        nombre="Luz Brill S.A.",
        ruc="80012345-6",
        direccion="Asunción, Paraguay",
        telefono="021-123456",
        email="contacto@luzbrill.com"
    )
    db.add(empresa)
    await db.flush()
    
    admin = Usuario(
        empresa_id=empresa.id,
        email="admin@luzbrill.com",
        password_hash=hash_password("admin123"),
        nombre="Administrador",
        apellido="Sistema"
    )
    db.add(admin)
    
    # Create comprehensive permissions
    permisos_data = [
        # Dashboard
        ("dashboard.ver", "Ver dashboard"),
        # Ventas
        ("ventas.crear", "Crear ventas"),
        ("ventas.ver", "Ver ventas"),
        ("ventas.anular", "Anular ventas"),
        ("ventas.modificar_precio", "Modificar precios en ventas"),
        ("ventas.aplicar_descuento", "Aplicar descuentos"),
        ("ventas.imprimir_boleta", "Imprimir boletas"),
        ("ventas.imprimir_factura", "Imprimir facturas"),
        # Productos
        ("productos.crear", "Crear productos"),
        ("productos.editar", "Editar productos"),
        ("productos.eliminar", "Eliminar productos"),
        ("productos.modificar_precio", "Modificar precios de productos"),
        ("productos.ver_costo", "Ver costos de productos"),
        # Stock
        ("stock.ver", "Ver inventario"),
        ("stock.entrada", "Registrar entradas de stock"),
        ("stock.salida", "Registrar salidas de stock"),
        ("stock.ajustar", "Ajustar stock"),
        ("stock.traspasar", "Traspasar entre almacenes"),
        ("stock.configurar_alertas", "Configurar alertas de stock"),
        # Clientes
        ("clientes.crear", "Crear clientes"),
        ("clientes.editar", "Editar clientes"),
        ("clientes.eliminar", "Eliminar clientes"),
        ("clientes.ver_creditos", "Ver créditos de clientes"),
        ("clientes.modificar_descuento", "Modificar descuento de cliente"),
        # Proveedores
        ("proveedores.crear", "Crear proveedores"),
        ("proveedores.editar", "Editar proveedores"),
        ("proveedores.eliminar", "Eliminar proveedores"),
        ("proveedores.gestionar_deudas", "Gestionar deudas a proveedores"),
        # Funcionarios
        ("funcionarios.ver", "Ver funcionarios"),
        ("funcionarios.crear", "Crear funcionarios"),
        ("funcionarios.editar", "Editar funcionarios"),
        ("funcionarios.ver_salarios", "Ver salarios"),
        ("funcionarios.adelantos", "Registrar adelantos de salario"),
        ("funcionarios.pagar_salarios", "Marcar salarios como pagados"),
        # Delivery
        ("delivery.ver", "Ver entregas"),
        ("delivery.crear", "Crear entregas"),
        ("delivery.actualizar_estado", "Actualizar estado de entregas"),
        # Flota
        ("flota.ver", "Ver flota"),
        ("flota.gestionar", "Gestionar vehículos"),
        # Laboratorio
        ("laboratorio.crear", "Crear materias de laboratorio"),
        ("laboratorio.ver", "Ver materias de laboratorio"),
        # Sistema
        ("usuarios.ver", "Ver usuarios"),
        ("usuarios.gestionar", "Gestionar usuarios"),
        ("roles.gestionar", "Gestionar roles y permisos"),
        ("sistema.configurar", "Configurar sistema"),
        ("reportes.ver", "Ver reportes"),
        ("reportes.exportar", "Exportar reportes"),
        # Business Intelligence
        ("bi.ver", "Ver Business Intelligence"),
        # Additional view permissions
        ("productos.ver", "Ver productos"),
        ("proveedores.ver", "Ver proveedores"),
        ("clientes.ver", "Ver clientes"),
        ("facturas.ver", "Ver facturas"),
        ("ventas.ver_historial", "Ver historial de ventas"),
        # Gastos operativos
        ("gastos.ver", "Ver gastos operativos"),
        ("gastos.gestionar", "Gestionar gastos operativos"),
    ]
    for clave, desc in permisos_data:
        db.add(Permiso(clave=clave, descripcion=desc))
    
    # Create roles
    for rol_nombre in ["ADMIN", "GERENTE", "VENDEDOR", "DELIVERY"]:
        db.add(Rol(empresa_id=empresa.id, nombre=rol_nombre, descripcion=f"Rol {rol_nombre}"))
    
    await db.flush()  # Flush to get IDs
    
    # Assign all permissions to ADMIN role
    admin_rol_result = await db.execute(select(Rol).where(Rol.nombre == "ADMIN", Rol.empresa_id == empresa.id))
    admin_rol = admin_rol_result.scalar_one_or_none()
    
    if admin_rol:
        permisos_result = await db.execute(select(Permiso))
        all_permisos = permisos_result.scalars().all()
        for permiso in all_permisos:
            db.add(RolPermiso(rol_id=admin_rol.id, permiso_id=permiso.id))
    
    # Assign permissions to GERENTE role
    gerente_rol_result = await db.execute(select(Rol).where(Rol.nombre == "GERENTE", Rol.empresa_id == empresa.id))
    gerente_rol = gerente_rol_result.scalar_one_or_none()
    gerente_permisos = [
        "dashboard.ver",
        "ventas.crear", "ventas.ver", "ventas.anular", "ventas.modificar_precio", "ventas.aplicar_descuento",
        "ventas.imprimir_boleta", "ventas.imprimir_factura", "ventas.ver_historial",
        "productos.ver", "productos.crear", "productos.editar", "productos.modificar_precio",
        "stock.ver", "stock.entrada", "stock.salida", "stock.traspasar",
        "clientes.ver", "clientes.crear", "clientes.editar", "clientes.ver_creditos",
        "proveedores.ver", "proveedores.gestionar_deudas",
        "funcionarios.ver", "funcionarios.ver_salarios", "funcionarios.adelantos",
        "delivery.ver", "delivery.crear", "delivery.actualizar_estado",
        "flota.ver", "laboratorio.ver", "laboratorio.crear",
        "reportes.ver", "reportes.exportar",
        "bi.ver",
        "gastos.ver", "gastos.gestionar"
    ]
    if gerente_rol:
        for perm_clave in gerente_permisos:
            perm_result = await db.execute(select(Permiso).where(Permiso.clave == perm_clave))
            perm = perm_result.scalar_one_or_none()
            if perm:
                db.add(RolPermiso(rol_id=gerente_rol.id, permiso_id=perm.id))
    
    # Assign permissions to VENDEDOR role
    vendedor_rol_result = await db.execute(select(Rol).where(Rol.nombre == "VENDEDOR", Rol.empresa_id == empresa.id))
    vendedor_rol = vendedor_rol_result.scalar_one_or_none()
    vendedor_permisos = [
        "dashboard.ver",
        "ventas.crear", "ventas.ver", "ventas.imprimir_boleta", "ventas.imprimir_factura", "ventas.ver_historial",
        "productos.ver", "stock.ver",
        "clientes.ver", "clientes.crear",
        "delivery.ver", "delivery.crear",
        "laboratorio.ver"
    ]
    if vendedor_rol:
        for perm_clave in vendedor_permisos:
            perm_result = await db.execute(select(Permiso).where(Permiso.clave == perm_clave))
            perm = perm_result.scalar_one_or_none()
            if perm:
                db.add(RolPermiso(rol_id=vendedor_rol.id, permiso_id=perm.id))
    
    # Assign permissions to DELIVERY role
    delivery_rol_result = await db.execute(select(Rol).where(Rol.nombre == "DELIVERY", Rol.empresa_id == empresa.id))
    delivery_rol = delivery_rol_result.scalar_one_or_none()
    delivery_permisos = [
        "delivery.ver", "delivery.actualizar_estado",
        "flota.ver"
    ]
    if delivery_rol:
        for perm_clave in delivery_permisos:
            perm_result = await db.execute(select(Permiso).where(Permiso.clave == perm_clave))
            perm = perm_result.scalar_one_or_none()
            if perm:
                db.add(RolPermiso(rol_id=delivery_rol.id, permiso_id=perm.id))
    
    # Assign admin user to ADMIN role
    admin.rol_id = admin_rol.id if admin_rol else None
    
    # Create categories
    categorias = ["Pinturas", "Herramientas", "Materiales", "Accesorios"]
    for cat in categorias:
        db.add(Categoria(empresa_id=empresa.id, nombre=cat))
    
    # Create brands
    marcas = ["Alba", "Sherwin Williams", "Sinteplast", "Tersuave", "3M"]
    for marca in marcas:
        db.add(Marca(empresa_id=empresa.id, nombre=marca))
    
    # Create warehouses
    db.add(Almacen(empresa_id=empresa.id, nombre="Depósito Principal", ubicacion="Planta Baja"))
    db.add(Almacen(empresa_id=empresa.id, nombre="Tienda", ubicacion="Local Comercial"))
    
    # Create occasional client
    db.add(Cliente(
        empresa_id=empresa.id,
        nombre="Cliente",
        apellido="Ocasional",
        ruc="00000000-0"
    ))
    
    await db.commit()
    return {"message": "Data seeded successfully", "empresa_id": empresa.id}

@api_router.post("/reset-database")
async def reset_database(db: AsyncSession = Depends(get_db)):
    """
    PELIGRO: Elimina y recrea todas las tablas
    Solo usar en desarrollo o primera configuración
    """
    try:
        # Drop all tables
        async with engine.begin() as conn:
            await conn.run_sync(Base.metadata.drop_all)
            await conn.run_sync(Base.metadata.create_all)
        
        logger.info("Database reset completed")
        
        # Create default empresa
        empresa = Empresa(
            ruc="12345678901",
            razon_social="Luz Brill S.A.",
            nombre_comercial="Luz Brill",
            direccion="Av. Principal 123",
            telefono="123456789",
            email="contacto@luzbrill.com",
            estado=True
        )
        db.add(empresa)
        await db.flush()
        
        # Create admin role
        rol_admin = Rol(
            empresa_id=empresa.id,
            nombre="ADMIN",
            descripcion="Administrador del sistema",
            estado=True
        )
        db.add(rol_admin)
        await db.flush()
        
        # Create admin user
        admin = Usuario(
            empresa_id=empresa.id,
            email="admin@luzbrill.com",
            password_hash=hash_password("admin123"),
            nombre="Admin",
            apellido="Sistema",
            telefono="123456789",
            activo=True
        )
        db.add(admin)
        await db.flush()
        
        # Assign role to user via UsuarioRol
        usuario_rol = UsuarioRol(
            usuario_id=admin.id,
            rol_id=rol_admin.id
        )
        db.add(usuario_rol)
        
        # Create permissions
        permisos_data = [
            ("ventas.crear", "Crear ventas"),
            ("ventas.ver", "Ver ventas"),
            ("ventas.anular", "Anular ventas"),
            ("productos.crear", "Crear productos"),
            ("productos.editar", "Editar productos"),
            ("productos.eliminar", "Eliminar productos"),
            ("stock.ajustar", "Ajustar stock"),
            ("clientes.crear", "Crear clientes"),
            ("clientes.editar", "Editar clientes"),
            ("funcionarios.ver", "Ver funcionarios"),
            ("funcionarios.editar", "Editar funcionarios"),
            ("usuarios.gestionar", "Gestionar usuarios"),
            ("sistema.configurar", "Configurar sistema"),
            ("gastos.ver", "Ver gastos operativos"),
            ("gastos.gestionar", "Gestionar gastos operativos"),
        ]
        for clave, desc in permisos_data:
            permiso = Permiso(clave=clave, descripcion=desc, estado=True)
            db.add(permiso)
            await db.flush()
            
            # Assign all permissions to admin role
            rol_permiso = RolPermiso(rol_id=rol_admin.id, permiso_id=permiso.id)
            db.add(rol_permiso)
        
        # Create other roles
        for rol_nombre in ["GERENTE", "VENDEDOR", "DELIVERY"]:
            db.add(Rol(
                empresa_id=empresa.id,
                nombre=rol_nombre,
                descripcion=f"Rol {rol_nombre}",
                estado=True
            ))
        
        # Create categories
        categorias = ["Pinturas", "Herramientas", "Materiales", "Accesorios"]
        for cat in categorias:
            db.add(Categoria(empresa_id=empresa.id, nombre=cat, estado=True))
        
        # Create brands
        marcas = ["Alba", "Sherwin Williams", "Sinteplast", "Tersuave"]
        for marca in marcas:
            db.add(Marca(empresa_id=empresa.id, nombre=marca, estado=True))
        
        # Create warehouses
        db.add(Almacen(empresa_id=empresa.id, nombre="Depósito Principal", ubicacion="Planta Baja", estado=True))
        db.add(Almacen(empresa_id=empresa.id, nombre="Tienda", ubicacion="Local Comercial", estado=True))
        
        # Create occasional client
        db.add(Cliente(
            empresa_id=empresa.id,
            nombre="Cliente",
            apellido="Ocasional",
            ruc="00000000-0",
            estado=True
        ))
        
        await db.commit()
        
        return {
            "status": "success",
            "message": "Base de datos reseteada y poblada exitosamente",
            "credentials": {
                "email": "admin@luzbrill.com",
                "password": "admin123"
            },
            "empresa_id": empresa.id
        }
    except Exception as e:
        logger.error(f"Error al resetear base de datos: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error al resetear base de datos: {str(e)}")

# Include router
app.include_router(api_router)

# Health check endpoint para debugging
@app.get("/health")
async def health_check():
    return {
        "status": "ok",
        "docs_dir": str(DOCS_DIR),
        "docs_dir_exists": DOCS_DIR.exists(),
        "backend_url": os.getenv("BACKEND_URL", "NOT_SET")
    }

# Endpoint temporal para migrar tabla documentos_temporales
@app.post("/api/admin/migrate-documentos-temporales")
async def migrate_documentos_temporales():
    """
    Migra la tabla documentos_temporales para usar TIMESTAMP WITH TIME ZONE
    """
    try:
        async with engine.begin() as conn:
            # Primero eliminar datos existentes si los hay (temporales de todas formas)
            await conn.execute(text("DROP TABLE IF EXISTS documentos_temporales CASCADE"))
            
            # Recrear la tabla con el nuevo esquema
            await conn.run_sync(Base.metadata.create_all)
            
        logger.info("Tabla documentos_temporales migrada exitosamente")
        return {
            "status": "success",
            "message": "Tabla documentos_temporales migrada a TIMESTAMP WITH TIME ZONE"
        }
    except Exception as e:
        logger.error(f"Error al migrar tabla: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Error en migración: {str(e)}")

# Startup event
@app.on_event("startup")
async def startup():
    await init_db()
    # Auto-migration: add cantidad_restante column to movimientos_stock if missing
    try:
        async with engine.begin() as conn:
            db_url = str(engine.url)
            if 'postgresql' in db_url:
                await conn.execute(text(
                    "ALTER TABLE movimientos_stock ADD COLUMN IF NOT EXISTS cantidad_restante INTEGER"
                ))
            else:
                try:
                    await conn.execute(text(
                        "ALTER TABLE movimientos_stock ADD COLUMN cantidad_restante INTEGER"
                    ))
                except Exception:
                    pass  # Column already exists in SQLite
        logger.info("Stock migration: cantidad_restante column ensured")
    except Exception as e:
        logger.warning(f"Stock migration warning (may be harmless): {e}")

    # Auto-migration: ensure bi.ver permission exists and is assigned to ADMIN roles
    try:
        async with async_session_maker() as db:
            # Insert permission if missing
            existing = await db.execute(select(Permiso).where(Permiso.clave == "bi.ver"))
            if not existing.scalar_one_or_none():
                bi_perm = Permiso(clave="bi.ver", descripcion="Ver Business Intelligence")
                db.add(bi_perm)
                await db.flush()
                # Assign to every ADMIN rol found
                admin_roles = await db.execute(select(Rol).where(Rol.nombre == "ADMIN"))
                for admin_rol in admin_roles.scalars().all():
                    db.add(RolPermiso(rol_id=admin_rol.id, permiso_id=bi_perm.id))
                # Assign to every GERENTE rol found
                gerente_roles = await db.execute(select(Rol).where(Rol.nombre == "GERENTE"))
                for g_rol in gerente_roles.scalars().all():
                    db.add(RolPermiso(rol_id=g_rol.id, permiso_id=bi_perm.id))
                await db.commit()
                logger.info("BI migration: bi.ver permission created and assigned")
            else:
                logger.info("BI migration: bi.ver already exists, skipping")
    except Exception as e:
        logger.warning(f"BI permission migration warning: {e}")

    logger.info("Database initialized")

@app.on_event("shutdown")
async def shutdown():
    await engine.dispose()
    logger.info("Database connection closed")
